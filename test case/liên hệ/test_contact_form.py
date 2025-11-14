import os
import time
import random
import string
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as xlImage
from PIL import Image as PILImage
import io

def random_string(length):
    """Tạo chuỗi ngẫu nhiên"""
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for _ in range(length))

def save_test_report(test_results, log_file, test_data=None, success_message=None, screenshot_paths=None):
    """Lưu báo cáo kiểm thử và xuất ra Excel
    
    Args:
        test_results: Kết quả kiểm thử
        log_file: Đường dẫn file log
        test_data: Dữ liệu test
        success_message: Thông báo thành công
        screenshot_paths: Danh sách đường dẫn ảnh chụp màn hình
    """
    if screenshot_paths is None:
        screenshot_paths = []
    # Tạo báo cáo văn bản
    report = f"""
==================================================
📝 BÁO CÁO KIỂM THỬ TỰ ĐỘNG - FORM LIÊN HỆ
==================================================
Thời gian bắt đầu: {test_results['start_time']}
Thời gian kết thúc: {test_results['end_time']}
Trạng thái: {'✅ THÀNH CÔNG' if test_results['status'] == 'PASSED' else '❌ THẤT BẠI'}

📋 THÔNG TIN GỬI TIN NHẮN:
"""
    
    if test_data:
        for key, value in test_data.items():
            report += f"• {key.capitalize()}: {value}\n"
    
    if success_message:
        report += f"\n📨 Thông báo phản hồi: {success_message}\n"
    
    report += "\n📋 CHI TIẾT CÁC BƯỚC KIỂM THỬ:\n"
    
    for step in test_results['steps']:
        status_icon = "✅" if step['status'] == 'PASSED' else "❌"
        report += f"""
--- Bước {step['step']}: {step['action']} {status_icon}
   • Mong đợi: {step['expected']}
   • Thực tế: {step['actual']}
   • Trạng thái: {'Thành công' if step['status'] == 'PASSED' else 'Thất bại'}
"""
    
    report += f"""
==================================================
📂 File log đầy đủ: {os.path.abspath(log_file)}
==================================================
"""
    
    # Lưu báo cáo văn bản
    report_file = log_file.replace('.txt', '_report.txt')
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write(report)
    
    # Tạo báo cáo Excel
    excel_file = log_file.replace('.txt', '_report.xlsx')
    
    # Tạo dữ liệu cho báo cáo Excel
    test_cases = [
        {
            'STT': 1,
            'Tên Test Case': 'Gửi thông tin liên hệ với dữ liệu hợp lệ',
            'Mục đích': 'Kiểm tra chức năng gửi thông tin liên hệ với dữ liệu hợp lệ',
            'Dữ liệu đầu vào': 'Họ tên: ' + test_data.get('name', '') + '\n' +
                             'Email: ' + test_data.get('email', '') + '\n' +
                             'Số điện thoại: ' + test_data.get('phone', '') + '\n' +
                             'Nội dung: ' + test_data.get('message', ''),
            'Kết quả mong đợi': 'Gửi thông tin thành công và hiển thị thông báo xác nhận',
            'Kết quả thực tế': success_message or 'Đã gửi thông tin liên hệ',
            'Trạng thái': 'Passed' if test_results['status'] == 'PASSED' else 'Failed',
            'Ghi chú': 'Kiểm thử tự động bằng Selenium',
            'Thời gian chạy': test_results['start_time'],
            'Ảnh chụp': 'Xem ảnh bên dưới' if screenshot_paths else 'Không có ảnh'
        }
    ]
    
    # Tạo DataFrame từ dữ liệu test cases
    df = pd.DataFrame(test_cases)
    
    # Sắp xếp lại cột
    df = df[['STT', 'Tên Test Case', 'Mục đích', 'Dữ liệu đầu vào', 
             'Kết quả mong đợi', 'Kết quả thực tế', 'Trạng thái', 
             'Ghi chú', 'Thời gian chạy']]
    
    # Lưu vào Excel
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Kết quả kiểm thử')
        
        # Lấy workbook và worksheet
        workbook = writer.book
        worksheet = writer.sheets['Kết quả kiểm thử']
        
        # Định dạng header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Áp dụng định dạng cho hàng đầu tiên (header)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Điều chỉnh độ rộng cột
        column_widths = {
            'A': 5,    # STT
            'B': 35,   # Tên Test Case
            'C': 25,   # Mục đích
            'D': 40,   # Dữ liệu đầu vào
            'E': 30,   # Kết quả mong đợi
            'F': 30,   # Kết quả thực tế
            'G': 15,   # Trạng thái
            'H': 20,   # Ghi chú
            'I': 20,   # Thời gian chạy
            'J': 15    # Ảnh chụp
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Tự động điều chỉnh chiều cao hàng
        for row in worksheet.iter_rows():
            max_length = 0
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_height = min(100, (max_length // 30 + 1) * 15)
            worksheet.row_dimensions[row[0].row].height = adjusted_height
        
        # Định dạng wrap text cho các ô dài
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical='top')
        
        # Thêm ảnh vào báo cáo
        if screenshot_paths:
            # Tạo sheet mới cho ảnh
            img_sheet = workbook.create_sheet(title="Ảnh chụp màn hình")
            img_sheet.column_dimensions['A'].width = 20
            img_sheet.column_dimensions['B'].width = 40
            
            # Thêm tiêu đề
            img_sheet['A1'] = 'STT'
            img_sheet['B1'] = 'Mô tả ảnh'
            img_sheet['C1'] = 'Hình ảnh'
            
            # Định dạng tiêu đề
            for cell in img_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Thêm ảnh vào sheet
            for idx, img_path in enumerate(screenshot_paths, start=2):
                if os.path.exists(img_path):
                    try:
                        # Thêm thông tin ảnh
                        img_sheet[f'A{idx}'] = idx - 1
                        img_sheet[f'B{idx}'] = f'Ảnh {idx-1} - {os.path.basename(img_path)}'
                        
                        # Mở và điều chỉnh kích thước ảnh
                        img = PILImage.open(img_path)
                        # Giảm kích thước ảnh nếu cần
                        max_width = 800
                        max_height = 600
                        img.thumbnail((max_width, max_height), PILImage.Resampling.LANCZOS)
                        
                        # Lưu ảnh tạm
                        temp_img = io.BytesIO()
                        img.save(temp_img, format='PNG')
                        temp_img.seek(0)
                        
                        # Thêm ảnh vào sheet
                        img_obj = xlImage(temp_img)
                        img_obj.anchor = f'C{idx}'
                        img_sheet.add_image(img_obj)
                        
                        # Điều chỉnh chiều cao hàng cho phù hợp
                        img_sheet.row_dimensions[idx].height = img.height * 0.8
                        
                    except Exception as e:
                        print(f"Không thể thêm ảnh {img_path}: {str(e)}")
        
        # Thêm filter
        for sheet in [worksheet, img_sheet if 'img_sheet' in locals() else None]:
            if sheet:
                sheet.auto_filter.ref = sheet.dimensions
    
    print(report)
    print(f"\n📊 Đã lưu báo cáo Excel: {os.path.abspath(excel_file)}")
    return report_file, excel_file

def test_contact_form():
    # Tạo thư mục kết quả test
    current_dir = os.path.dirname(os.path.abspath(__file__))
    ket_qua_test_dir = os.path.join(current_dir, "kết quả test")
    if not os.path.exists(ket_qua_test_dir):
        os.makedirs(ket_qua_test_dir)
    
    # Tạo thư mục screenshots nếu chưa tồn tại
    screenshots_dir = os.path.join(current_dir, "screenshots")
    if not os.path.exists(screenshots_dir):
        os.makedirs(screenshots_dir)
    
    # Tạo file log trong thư mục kết quả test
    log_file = os.path.join(ket_qua_test_dir, f"test_log_{int(time.time())}.txt")
    
    def log_message(message):
        """Ghi log và in ra console"""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        print(log_entry)
        with open(log_file, 'a', encoding='utf-8') as f:
            f.write(log_entry + "\n")
    
    # Tổng hợp kết quả kiểm thử
    test_results = {
        'steps': [],
        'start_time': time.strftime("%Y-%m-%d %H:%M:%S"),
        'status': 'PASSED'
    }
    
    def add_test_step(step_num, action, expected, actual, status='PASSED'):
        """Thêm bước kiểm thử vào kết quả"""
        test_results['steps'].append({
            'step': step_num,
            'action': action,
            'expected': expected,
            'actual': actual,
            'status': status
        })
        if status == 'FAILED':
            test_results['status'] = 'FAILED'
    
    try:
        # Bước 1: Khởi tạo trình duyệt
        log_message("🔄 Bước 1: Đang khởi tạo trình duyệt Chrome...")
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), 
                                options=chrome_options)
        
        # Bước 2: Mở trang liên hệ
        log_message("🌐 Bước 2: Đang mở trang liên hệ...")
        driver.get("http://localhost/webbansach/lien-he.php")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "form"))
        )
        add_test_step(2, "Mở trang liên hệ", 
                     "Trang liên hệ được tải thành công", 
                     "Trang đã được tải thành công")
        
        # Tạo dữ liệu test
        test_data = {
            'name': 'Nguyễn Văn A',
            'email': f'test_{random_string(8)}@example.com',
            'phone': '09' + ''.join(random.choices(string.digits, k=8)),
            'message': 'Đây là tin nhắn kiểm thử tự động. Xin vui lòng bỏ qua.'
        }
        
        # Bước 3: Điền thông tin liên hệ
        log_message("📝 Bước 3: Đang điền thông tin liên hệ...")
        try:
            # Điền tên
            name_field = driver.find_element(By.ID, 'con_name')
            name_field.send_keys(test_data['name'])
            
            # Điền email
            email_field = driver.find_element(By.ID, 'con_email')
            email_field.send_keys(test_data['email'])
            
            # Điền số điện thoại
            phone_field = driver.find_element(By.ID, 'con_phone')
            phone_field.send_keys(test_data['phone'])
            
            # Điền nội dung
            message_field = driver.find_element(By.ID, 'con_message')
            message_field.send_keys(test_data['message'])
            
            # Chụp màn hình sau khi điền form
            screenshot_path = os.path.join(screenshots_dir, f"form_filled_{int(time.time())}.png")
            driver.save_screenshot(screenshot_path)
            log_message(f"📸 Đã lưu ảnh form đã điền: {screenshot_path}")
            
            add_test_step(3, "Điền thông tin liên hệ", 
                         "Điền đầy đủ thông tin vào form", 
                         "Đã điền đầy đủ thông tin vào form")
            
        except Exception as e:
            error_msg = f"Lỗi khi điền form: {str(e)}"
            log_message(f"❌ {error_msg}")
            add_test_step(3, "Điền thông tin liên hệ", 
                         "Điền đầy đủ thông tin vào form", 
                         error_msg, 'FAILED')
            raise
        
        # Bước 4: Tìm và nhấn nút gửi tin nhắn
        log_message("🚀 Bước 4: Đang tìm và nhấn nút gửi tin nhắn...")
        success_message = None
        
        try:
            # Tìm nút submit bằng XPath chính xác với class và nội dung
            submit_xpath = "//button[@type='submit' and contains(@class, 'btn-outlined--primary') and contains(@class, 'btn-contact')]"
            submit_buttons = driver.find_elements(By.XPATH, submit_xpath)
            
            # Nếu không tìm thấy, thử tìm bằng nội dung
            if not submit_buttons:
                submit_buttons = driver.find_elements(By.XPATH, 
                    "//button[contains(., 'Gửi Tin Nhắn')] | " +
                    "//button[contains(@class, 'btn-contact')] | " +
                    "//button[contains(text(), 'Gửi Tin Nhắn')]")
            
            if not submit_buttons:
                raise Exception("Không tìm thấy nút gửi tin nhắn")
                
            submit_button = submit_buttons[0]
            
            # Cuộn đến nút gửi để đảm bảo nó hiển thị
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", submit_button)
            time.sleep(1)  # Chờ một chút để hoàn thành cuộn
            
            # Làm nổi bật nút gửi
            original_style = submit_button.get_attribute("style")
            driver.execute_script("arguments[0].style.border='3px solid red';", submit_button)
            
            # Chụp màn hình trước khi nhấn nút gửi
            screenshot_path = os.path.join(screenshots_dir, f"before_submit_{int(time.time())}.png")
            driver.save_screenshot(screenshot_path)
            log_message(f"📸 Đã lưu ảnh trước khi gửi: {screenshot_path}")
            
            # Đặt lại style
            driver.execute_script(f"arguments[0].style.border='{original_style}';", submit_button)
            
            # Nhấn nút gửi bằng JavaScript để tránh các vấn đề với Selenium
            driver.execute_script("arguments[0].click();", submit_button)
            log_message("✅ Đã nhấn nút gửi tin nhắn")
            
            # Chờ cho đến khi có thông báo hoặc chuyển trang
            try:
                # Chờ tối đa 10 giây cho thông báo thành công
                success_alert = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, 'alert-success'))
                )
                success_message = success_alert.text.strip()
                log_message(f"✅ Thông báo thành công: {success_message}")
                
                # Tạo tên file ảnh với thời gian
                timestamp = int(time.time())
                
                # Chụp ảnh thông báo thành công
                success_screenshot = os.path.join(screenshots_dir, f"success_alert_{timestamp}.png")
                success_alert.screenshot(success_screenshot)
                log_message(f"📸 Đã lưu ảnh thông báo thành công: {success_screenshot}")
                
                # Chờ một chút để đảm bảo thông báo hiển thị đầy đủ
                time.sleep(1)
                
                # Chụp toàn màn hình để có bối cảnh đầy đủ
                full_page_screenshot = os.path.join(screenshots_dir, f"success_page_{timestamp}.png")
                driver.save_screenshot(full_page_screenshot)
                log_message(f"📸 Đã lưu ảnh toàn màn hình sau khi gửi: {full_page_screenshot}")
                
                # Lưu đường dẫn ảnh để thêm vào báo cáo
                test_results['screenshot_paths'] = [success_screenshot, full_page_screenshot]
                
            except Exception as e:
                log_message(f"⚠ Không tìm thấy thông báo thành công: {str(e)}")
                
                # Kiểm tra xem có thông báo lỗi không
                try:
                    error_alert = driver.find_element(By.CLASS_NAME, 'alert-danger')
                    error_msg = error_alert.text.strip()
                    log_message(f"❌ Thông báo lỗi: {error_msg}")
                    raise Exception(f"Gửi tin nhắn thất bại: {error_msg}")
                except:
                    # Nếu không có thông báo lỗi, có thể đã chuyển hướng
                    if "lien-he" not in driver.current_url and "contact" not in driver.current_url.lower():
                        success_message = "Gửi tin nhắn thành công (đã chuyển hướng)"
                        log_message(f"✅ {success_message}")
                    else:
                        # Chụp màn hình để kiểm tra
                        error_screenshot = os.path.join(screenshots_dir, f"unknown_status_{int(time.time())}.png")
                        driver.save_screenshot(error_screenshot)
                        log_message(f"⚠ Không xác định trạng thái, đã lưu ảnh: {error_screenshot}")
                        raise Exception("Không xác định được trạng thái gửi tin nhắn")
            
            add_test_step(4, "Nhấn nút gửi tin nhắn", 
                         "Nhấn nút gửi tin nhắn và xác nhận thành công", 
                         f"Đã nhấn nút gửi tin nhắn. {success_message if success_message else 'Không có thông báo xác nhận'}")
            
        except Exception as e:
            error_msg = f"Lỗi khi nhấn nút gửi: {str(e)}"
            log_message(f"❌ {error_msg}")
            # Chụp màn hình lỗi
            error_screenshot = os.path.join(screenshots_dir, f"error_submit_{int(time.time())}.png")
            driver.save_screenshot(error_screenshot)
            log_message(f"📸 Đã lưu ảnh lỗi: {error_screenshot}")
            add_test_step(4, "Nhấn nút gửi tin nhắn", 
                         "Nhấn nút gửi tin nhắn thành công", 
                         error_msg, 'FAILED')
            raise
        
        # Bước 5: Xác minh kết quả
        log_message("🔄 Bước 5: Đang xác minh kết quả...")
        try:
            # Nếu chưa có thông báo thành công, thử tìm lại
            if not success_message:
                try:
                    success_element = WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.CLASS_NAME, 'alert-success'))
                    )
                    success_message = success_element.text.strip()
                    log_message(f"✅ Xác nhận thông báo thành công: {success_message}")
                except:
                    # Kiểm tra URL nếu có thông báo thành công
                    if "status=success" in driver.current_url or "success=true" in driver.current_url.lower():
                        success_message = "Gửi tin nhắn thành công (xác nhận từ URL)"
                        log_message(f"✅ {success_message}")
            
            # Nếu vẫn không có thông báo, kiểm tra xem form đã được xóa chưa
            if not success_message:
                try:
                    # Kiểm tra xem các trường đã bị xóa chưa (dấu hiệu gửi thành công)
                    name_field = driver.find_element(By.ID, 'con_name')
                    if not name_field.get_attribute('value'):
                        success_message = "Gửi tin nhắn thành công (form đã được xóa)"
                        log_message(f"✅ {success_message}")
                except:
                    pass
            
            # Chụp màn hình sau khi gửi
            screenshot_path = os.path.join(screenshots_dir, f"after_submit_{int(time.time())}.png")
            driver.save_screenshot(screenshot_path)
            log_message(f"📸 Đã lưu ảnh sau khi gửi: {screenshot_path}")
            
            # Kiểm tra thông báo thành công
            success = False
            actual_result = ""
            
            if "status=success" in driver.current_url:
                success = True
                actual_result = "URL chứa tham số status=success"
            elif driver.find_elements(By.CLASS_NAME, 'alert-success'):
                success = True
                actual_result = "Tìm thấy thông báo thành công trên trang"
            elif driver.find_elements(By.CLASS_NAME, 'alert-danger'):
                actual_result = "Tìm thấy thông báo lỗi trên trang"
            else:
                actual_result = "Không xác định được trạng thái gửi tin nhắn"
            
            if success:
                log_message("✅ Gửi tin nhắn thành công!")
                add_test_step(5, "Xác minh kết quả", 
                             "Hiển thị thông báo gửi thành công", 
                             actual_result, 'PASSED')
                return True
            else:
                error_msg = f"Gửi tin nhắn thất bại: {actual_result}"
                log_message(f"❌ {error_msg}")
                add_test_step(5, "Xác minh kết quả", 
                             "Hiển thị thông báo gửi thành công", 
                             error_msg, 'FAILED')
                return False
                
        except Exception as e:
            error_msg = f"Lỗi khi xác minh kết quả: {str(e)}"
            log_message(f"❌ {error_msg}")
            # Chụp màn hình lỗi
            error_screenshot = os.path.join(screenshots_dir, f"error_verification_{int(time.time())}.png")
            driver.save_screenshot(error_screenshot)
            log_message(f"📸 Đã lưu ảnh lỗi: {error_screenshot}")
            
            add_test_step(5, "Xác minh kết quả", 
                         "Hiển thị thông báo gửi thành công", 
                         error_msg, 'FAILED')
            return False
                
    except Exception as e:
        log_message(f"❌ Có lỗi xảy ra: {str(e)}")
        test_results['status'] = 'FAILED'
        return False
    finally:
        # Lưu báo cáo
        test_results['end_time'] = time.strftime("%Y-%m-%d %H:%M:%S")
        screenshot_paths = test_results.get('screenshot_paths', [])
        
        # Đảm bảo Excel được lưu vào folder "kết quả test"
        current_dir = os.path.dirname(os.path.abspath(__file__))
        ket_qua_test_dir = os.path.join(current_dir, "kết quả test")
        if not os.path.exists(ket_qua_test_dir):
            os.makedirs(ket_qua_test_dir)
        
        # Tạo log_file trong folder "kết quả test" nếu chưa có
        if not log_file or not log_file.startswith(ket_qua_test_dir):
            log_file = os.path.join(ket_qua_test_dir, f"test_log_{int(time.time())}.txt")
        
        report_file, excel_file = save_test_report(
            test_results, 
            log_file, 
            test_data, 
            success_message,
            screenshot_paths=screenshot_paths
        )
        log_message(f"📄 Đã lưu báo cáo kiểm thử: {report_file}")
        log_message(f"📊 Đã lưu báo cáo Excel: {excel_file}")
        
        # Đóng trình duyệt
        try:
            driver.quit()
            log_message("👋 Đã đóng trình duyệt")
        except:
            pass
        
        # In đường dẫn đến file báo cáo
        print(f"\n📄 Báo cáo chi tiết đã được lưu tại: {os.path.abspath(report_file)}")

def install_required_packages():
    """Cài đặt các gói Python cần thiết"""
    import sys
    import subprocess
    import pkg_resources
    
    required = {
        'selenium', 'webdriver-manager', 'pandas', 
        'openpyxl', 'webdriver-manager'
    }
    
    installed = {pkg.key for pkg in pkg_resources.working_set}
    missing = required - installed
    
    if missing:
        print(f"🔧 Đang cài đặt các gói còn thiếu: {', '.join(missing)}")
        python = sys.executable
        subprocess.check_call([python, '-m', 'pip', 'install', *missing], stdout=subprocess.DEVNULL)
        print("✅ Đã cài đặt xong các gói cần thiết")

if __name__ == "__main__":
    print("="*50)
    print("🚀 BẮT ĐẦU KIỂM THỬ CHỨC NĂNG LIÊN HỆ")
    print("="*50)
    
    # Cài đặt các gói cần thiết
    install_required_packages()
    
    # Tạo thư mục screenshots nếu chưa tồn tại
    if not os.path.exists('screenshots'):
        os.makedirs('screenshots')
    
    # Chạy test
    result = test_contact_form()
    
    # In kết quả
    if result:
        print("\n🎉 KIỂM THỬ HOÀN THÀNH THÀNH CÔNG!")
    else:
        print("\n❌ CÓ LỖI XẢY RA TRONG QUÁ TRÌNH KIỂM THỬ.")
    
    print("\nKết thúc chương trình.")
