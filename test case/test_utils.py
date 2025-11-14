import os
import time
import io
from datetime import datetime
from PIL import Image as PILImage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage

class TestReporter:
    def __init__(self, test_name):
        self.test_name = test_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Kết quả kiểm thử"
        self.current_row = 1
        self.screenshot_count = 0
        self.temp_files = []  # To keep track of temporary files
        self.test_steps = []
        self.setup_worksheet()
        
    def setup_worksheet(self):
        # Thiết lập tiêu đề và định dạng cho bảng tính
        self.ws = self.wb.active
        self.ws.title = self.test_name[:31]  # Giới hạn độ dài tiêu đề cho Excel
        
        # Định dạng cột
        self.ws.column_dimensions['A'].width = 15  # Thời gian
        self.ws.column_dimensions['B'].width = 40  # Bước thực hiện
        self.ws.column_dimensions['C'].width = 30  # Dữ liệu đầu vào
        self.ws.column_dimensions['D'].width = 30  # Output/Kết quả thực tế
        self.ws.column_dimensions['E'].width = 30  # Kết quả mong đợi
        self.ws.column_dimensions['F'].width = 15  # Trạng thái (Pass/Fail)
        self.ws.column_dimensions['G'].width = 20  # Hình ảnh
        
        # Tạo tiêu đề báo cáo
        self.ws.merge_cells('A1:G1')
        title_cell = self.ws.cell(row=1, column=1, value=f'BÁO CÁO KIỂM THỬ: {self.test_name.upper()}')
        title_cell.font = Font(size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Thêm tiêu đề cột
        headers = ['Thời gian', 'Bước thực hiện', 'Input', 'Output', 'Kết quả mong đợi', 'Trạng thái', 'Hình ảnh']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        self.current_row = 3
    
    def add_step(self, description, status='PASS', input_data='', output='', expected='', screenshot_path=None, notes=''):
        """Thêm một bước kiểm thử vào báo cáo
        
        Args:
            description: Mô tả bước thực hiện
            status: Trạng thái (PASS/FAIL/WARNING/ERROR/INFO)
            input_data: Dữ liệu đầu vào
            output: Kết quả thực tế/Output
            expected: Kết quả mong đợi
            screenshot_path: Đường dẫn đến ảnh chụp màn hình (nếu có)
            notes: Ghi chú bổ sung
        """
        timestamp = datetime.now().strftime('%H:%M:%S')
        
        # Lưu thông tin bước vào danh sách
        step_info = {
            'timestamp': timestamp,
            'description': description,
            'input_data': input_data,
            'output': output,
            'expected': expected,
            'status': status,
            'screenshot_path': screenshot_path,
            'notes': notes
        }
        self.test_steps.append(step_info)
        
        # Thêm dữ liệu vào bảng tính
        row_data = [timestamp, description, input_data, output, expected, status, '']
        
        for col_num, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Định dạng ô trạng thái
            if col_num == 6:  # Cột trạng thái
                if status == 'PASS':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100', bold=True)
                    cell.value = 'PASS'
                elif status == 'FAIL':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006', bold=True)
                    cell.value = 'FAIL'
                elif status == 'WARNING':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    cell.font = Font(color='9C5700', bold=True)
                    cell.value = 'WARNING'
                elif status == 'ERROR':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='FF0000', bold=True)
                    cell.value = 'FAIL'
                else:  # INFO
                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                    cell.font = Font(color='000000')
                    cell.value = 'INFO'
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thêm border cho các ô
        for col in range(1, 8):
            cell = self.ws.cell(row=self.current_row, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Chèn ảnh nếu có
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                self._add_screenshot_to_cell(screenshot_path, self.current_row, 7)
            except Exception as e:
                print(f"Lỗi khi chèn ảnh vào bước: {str(e)}")
                # Ghi path vào cột hình ảnh nếu không chèn được
                self.ws.cell(row=self.current_row, column=7, value=f"Ảnh: {os.path.basename(screenshot_path)}")
        
        self.current_row += 1
    
    def _add_screenshot_to_cell(self, image_path, row, col):
        """Chèn ảnh vào một ô cụ thể trong Excel"""
        try:
            # Mở và resize ảnh
            img = PILImage.open(image_path)
            
            # Tính toán kích thước mới (max width 200px, max height 150px)
            max_width = 200
            max_height = 150
            
            # Tính aspect ratio
            img_ratio = img.height / img.width
            
            if img.width > max_width:
                new_width = max_width
                new_height = int(new_width * img_ratio)
            else:
                new_width = img.width
                new_height = img.height
            
            if new_height > max_height:
                new_height = max_height
                new_width = int(new_height / img_ratio)
            
            # Resize ảnh
            img_resized = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # Lưu vào file tạm
            temp_dir = os.path.join(os.getcwd(), 'temp_screenshots')
            os.makedirs(temp_dir, exist_ok=True)
            temp_img_path = os.path.join(temp_dir, f"excel_step_{row}_{self.screenshot_count}.png")
            self.screenshot_count += 1
            img_resized.save(temp_img_path, 'PNG', quality=95)
            self.temp_files.append(temp_img_path)
            
            # Thêm ảnh vào Excel
            img_excel = XLImage(temp_img_path)
            img_excel.width = new_width
            img_excel.height = new_height
            
            # Position image in column G (7th column)
            col_letter = get_column_letter(col)
            img_anchor = f'{col_letter}{row}'
            img_excel.anchor = img_anchor
            self.ws.add_image(img_excel)
            
            # Điều chỉnh row height để fit ảnh
            row_height = min(int(new_height * 0.75), 150)
            self.ws.row_dimensions[row].height = row_height
            
            # Điều chỉnh column width
            col_width = (new_width / 7) + 2
            self.ws.column_dimensions[col_letter].width = min(max(col_width, 20), 30)
            
        except Exception as e:
            print(f"Lỗi khi chèn ảnh vào ô: {str(e)}")
            raise
            
    def add_final_screenshot(self, screenshot, description="Ảnh chụp màn hình cuối cùng"):
        """Thêm ảnh chụp màn hình cuối cùng vào báo cáo"""
        if not screenshot:
            return
            
        try:
            # Lưu ảnh vào file tạm
            temp_dir = os.path.join(os.getcwd(), 'temp_screenshots')
            os.makedirs(temp_dir, exist_ok=True)
            temp_img_path = os.path.join(temp_dir, f'final_{self.screenshot_count}.png')
            self.screenshot_count += 1
            
            # Lưu ảnh
            if isinstance(screenshot, PILImage.Image):
                screenshot.save(temp_img_path, 'PNG')
            else:
                # Nếu là file path
                import shutil
                shutil.copy(screenshot, temp_img_path)
            
            # Thêm bước với ảnh
            self.add_step(
                description=description,
                status='INFO',
                input_data='',
                output='Đã chụp ảnh màn hình cuối cùng',
                expected='Có ảnh chụp màn hình',
                screenshot_path=temp_img_path
            )
            
            # Thêm vào danh sách file tạm để xóa sau
            self.temp_files.append(temp_img_path)
                
        except Exception as e:
            print(f"Lỗi khi xử lý ảnh chụp màn hình: {str(e)}")
    
    def save_report(self, folder_path, final_screenshot=None):
        """Lưu báo cáo ra file Excel"""
        # Thêm ảnh chụp màn hình cuối cùng nếu có
        if final_screenshot:
            self.add_final_screenshot(final_screenshot)
            
        # Tạo thư mục nếu chưa tồn tại
        os.makedirs(folder_path, exist_ok=True)
        
        # Tạo tên file báo cáo
        report_name = f"ket_qua_{self.test_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(folder_path, report_name)
            
        # Tự động điều chỉnh độ rộng cột
        for col_idx in range(1, 8):  # Chúng ta có 7 cột từ A đến G
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Bỏ qua nếu cột đã được merge
            is_merged = False
            for merged_range in self.ws.merged_cells.ranges:
                if column_letter in [get_column_letter(c) for c in range(merged_range.min_col, merged_range.max_col + 1)]:
                    is_merged = True
                    break
            
            if is_merged:
                continue
                
            # Tìm độ dài lớn nhất của nội dung trong cột
            for row in self.ws.iter_rows(min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value and not isinstance(cell, MergedCell):
                    try:
                        cell_value = str(cell.value)
                        if '\n' in cell_value:
                            max_line_length = max(len(line) for line in cell_value.split('\n'))
                            if max_line_length > max_length:
                                max_length = max_line_length
                        else:
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                    except:
                        pass
            
            # Điều chỉnh độ rộng cột (nhưng không override nếu đã set cho ảnh)
            if max_length > 0 and col_idx != 7:  # Bỏ qua cột G (hình ảnh) vì đã set riêng
                adjusted_width = (max_length + 2) * 1.2
                # Giới hạn độ rộng tối đa cho từng cột
                max_widths = {'A': 15, 'B': 50, 'C': 40, 'D': 40, 'E': 40, 'F': 15, 'G': 25}
                max_width = max_widths.get(column_letter, 30)
                current_width = self.ws.column_dimensions[column_letter].width or 8.43
                self.ws.column_dimensions[column_letter].width = min(max(adjusted_width, current_width), max_width)
        
        # Đặt chế độ in vừa trang
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0
        
        # Tạo thư mục nếu chưa tồn tại
        os.makedirs(folder_path, exist_ok=True)
        
        # Tạo tên file báo cáo
        report_name = f"ket_qua_{self.test_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(folder_path, report_name)
        
        try:
            # Lưu file
            self.wb.save(report_path)
            return report_path
        finally:
            # Dọn dẹp file tạm
            for temp_file in self.temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    print(f"Không thể xóa file tạm {temp_file}: {str(e)}")
            
            # Xóa thư mục tạm nếu rỗng
            temp_dir = os.path.join(os.getcwd(), 'temp_screenshots')
            try:
                if os.path.exists(temp_dir) and not os.listdir(temp_dir):
                    os.rmdir(temp_dir)
            except Exception as e:
                print(f"Không thể xóa thư mục tạm: {str(e)}")

def tao_thu_muc_ket_qua(test_name, base_dir=None):
    """Tạo thư mục lưu kết quả kiểm thử
    
    Args:
        test_name: Tên test case
        base_dir: Thư mục gốc để tạo folder "kết quả test". 
                  Nếu None, sẽ tự động tìm thư mục chứa file đang gọi hàm này.
    """
    import inspect
    import os
    
    # Nếu không chỉ định base_dir, tìm thư mục chứa file đang gọi hàm này
    if base_dir is None:
        # Lấy frame của hàm đang gọi
        frame = inspect.currentframe()
        try:
            # Lấy frame của hàm caller (người gọi tao_thu_muc_ket_qua)
            caller_frame = frame.f_back
            caller_file = caller_frame.f_globals.get('__file__')
            if caller_file:
                base_dir = os.path.dirname(os.path.abspath(caller_file))
            else:
                # Fallback: sử dụng thư mục hiện tại
                base_dir = os.getcwd()
        finally:
            del frame
    
    # Tạo folder "kết quả test" trong thư mục base_dir
    thu_muc_goc = os.path.join(base_dir, "kết quả test")
    os.makedirs(thu_muc_goc, exist_ok=True)
    
    now = datetime.now()
    thu_muc_ket_qua = os.path.join(thu_muc_goc, f"{test_name}_{now.strftime('%Y%m%d_%H%M%S')}")
    os.makedirs(thu_muc_ket_qua, exist_ok=True)
    return thu_muc_ket_qua

def chup_man_hinh(driver):
    """Chụp màn hình và trả về đối tượng ảnh"""
    try:
        screenshot = driver.get_screenshot_as_png()
        return PILImage.open(io.BytesIO(screenshot))
    except Exception as e:
        print(f"Lỗi khi chụp màn hình: {str(e)}")
        return None

def highlight_element(driver, element):
    """Tô sáng phần tử để dễ nhận biết"""
    try:
        original_style = element.get_attribute('style')
        driver.execute_script("arguments[0].style.border='3px solid red';", element)
        time.sleep(0.5)
        return original_style
    except Exception as e:
        print(f"Không thể tô sáng phần tử: {str(e)}")
        return ""

def in_thong_bao(reporter, noi_dung, status='PASS', input_data='', output='', expected='', screenshot_path=None, **kwargs):
    """In thông báo ra màn hình và thêm vào báo cáo
    
    Args:
        reporter: Đối tượng báo cáo
        noi_dung: Nội dung bước kiểm thử
        status: Trạng thái (PASS/FAIL/INFO/WARNING/ERROR)
        input_data: Dữ liệu đầu vào (nếu có)
        output: Kết quả thực tế/Output (nếu có)
        expected: Kết quả mong đợi (nếu có)
        screenshot_path: Đường dẫn đến ảnh chụp màn hình (nếu có)
        **kwargs: Các tham số bổ sung (sẽ được thêm vào notes)
    """
    timestamp = datetime.now().strftime('[%H:%M:%S]')
    status = status.upper()
    status_text = status if status in ['PASS', 'FAIL', 'INFO', 'WARNING', 'ERROR'] else 'INFO'
    
    # In ra console
    print(f"{timestamp} {noi_dung}")
    
    # Xử lý ghi chú từ các tham số bổ sung
    notes = []
    for key, value in kwargs.items():
        if value and key not in ['screenshot']:  # Bỏ qua screenshot vì đã có tham số riêng
            notes.append(f"{key}: {value}")
    
    # Thêm vào báo cáo Excel
    if reporter:
        reporter.add_step(
            description=noi_dung,
            status=status_text,
            input_data=str(input_data) if input_data else '',
            output=str(output) if output else '',
            expected=str(expected) if expected else '',
            screenshot_path=screenshot_path,
            notes='\n'.join(notes) if notes else ''
        )

    return True

def lay_san_pham_ngau_nhien(driver, so_luong=1):
    """Lấy sản phẩm ngẫu nhiên từ trang chủ hoặc trang tất cả sản phẩm
    
    Args:
        driver: WebDriver instance
        so_luong: Số lượng sản phẩm cần lấy (mặc định 1)
    
    Returns:
        List of dict với keys: url, name, id
    """
    import random
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    
    try:
        # Thử lấy từ trang chủ trước
        driver.get("http://localhost/webbansach/index.php")
        time.sleep(2)
        
        # Tìm tất cả các link sản phẩm
        product_links = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[href*='san-pham.php?id=']"))
        )
        
        # Nếu không có trên trang chủ, thử trang tất cả sản phẩm
        if not product_links or len(product_links) < so_luong:
            driver.get("http://localhost/webbansach/tat-ca-san-pham.php")
            time.sleep(2)
            product_links = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[href*='san-pham.php?id=']"))
            )
        
        if not product_links:
            raise Exception("Không tìm thấy sản phẩm nào")
        
        # Loại bỏ các link trùng lặp
        unique_links = []
        seen_urls = set()
        for link in product_links:
            url = link.get_attribute('href')
            if url and url not in seen_urls:
                seen_urls.add(url)
                unique_links.append(link)
        
        # Chọn ngẫu nhiên
        if len(unique_links) < so_luong:
            selected_links = unique_links
        else:
            selected_links = random.sample(unique_links, so_luong)
        
        # Lấy thông tin sản phẩm
        products = []
        for link in selected_links:
            url = link.get_attribute('href')
            # Extract product ID from URL
            product_id = None
            if 'id=' in url:
                product_id = url.split('id=')[1].split('&')[0]
            
            # Try to get product name
            product_name = "Sản phẩm không xác định"
            try:
                # Try to find product name in the link or nearby elements
                parent = link.find_element(By.XPATH, "./ancestor::div[contains(@class, 'product-card') or contains(@class, 'product')]")
                name_elem = parent.find_elements(By.CSS_SELECTOR, "h3, .product-name, .product-title, a[href*='san-pham.php']")
                if name_elem:
                    product_name = name_elem[0].text.strip()
            except:
                pass
            
            products.append({
                'url': url,
                'name': product_name,
                'id': product_id
            })
        
        return products
        
    except Exception as e:
        print(f"Lỗi khi lấy sản phẩm ngẫu nhiên: {str(e)}")
        # Fallback: trả về sản phẩm mặc định
        return [{
            'url': 'http://localhost/webbansach/san-pham.php?id=12',
            'name': 'Sản phẩm mặc định',
            'id': '12'
        }]
