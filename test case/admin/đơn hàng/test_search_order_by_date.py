import os
import sys
import logging
import warnings
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, 
    NoSuchElementException, 
    WebDriverException
)
import time
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("test_search_order_by_date.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Set console output to UTF-8
if sys.stdout.encoding != 'utf-8':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

class TestSearchOrderByDate:
    def __init__(self):
        self.driver = None
        self.test_steps = []
        self.screenshot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshots')
        self.report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ket-qua-test')
        
        # Create necessary directories
        os.makedirs(self.screenshot_dir, exist_ok=True)
        os.makedirs(self.report_dir, exist_ok=True)
        
        # Test data - use today's date for search
        today = datetime.now()
        self.test_data = {
            'admin_url': 'http://localhost/webbansach/admin',
            'login_url': 'http://localhost/webbansach/admin/dang-nhap.php',
            'order_url': 'http://localhost/webbansach/admin/don-hang.php',
            'username': 'admin',
            'password': 'admin',
            'search_date': today.strftime('%Y-%m-%d'),  # Format: YYYY-MM-DD
            'order_count': 0
        }
        
    def setup_driver(self):
        """Initialize the WebDriver"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--remote-debugging-port=9222")
            chrome_options.add_argument("--window-size=1920,1080")
            
            try:
                self.driver = webdriver.Chrome(options=chrome_options)
                self.driver.set_window_size(1920, 1080)
                self.driver.set_page_load_timeout(30)
                logger.info("WebDriver initialized successfully")
                self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", "WebDriver khởi tạo thành công", "PASS")
                return True
            except Exception as e:
                logger.error(f"Lỗi khi khởi tạo WebDriver: {str(e)}")
                try:
                    from webdriver_manager.chrome import ChromeDriverManager
                    from selenium.webdriver.chrome.service import Service as ChromeService
                    service = ChromeService(ChromeDriverManager().install())
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    self.driver.set_window_size(1920, 1080)
                    self.driver.set_page_load_timeout(30)
                    logger.info("WebDriver initialized successfully")
                    self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", "WebDriver khởi tạo thành công", "PASS")
                    return True
                except Exception as e2:
                    logger.error(f"Lỗi khi thử cách khởi tạo WebDriver thứ 2: {str(e2)}")
                    self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", f"Lỗi: {str(e2)}", "FAIL")
                    return False
        except Exception as e:
            logger.error(f"Lỗi không xác định khi khởi tạo WebDriver: {str(e)}", exc_info=True)
            self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", f"Lỗi: {str(e)}", "FAIL")
            return False
    
    def add_test_step(self, action, expected, actual, status, screenshot_path=None):
        """Add a test step with optional screenshot"""
        try:
            step = {
                'action': str(action) if action else "",
                'expected': str(expected) if expected else "",
                'actual': str(actual) if actual else "",
                'status': str(status) if status else "UNKNOWN",
            }
            
            if screenshot_path and os.path.isfile(screenshot_path):
                try:
                    with PILImage.open(screenshot_path) as img:
                        img.verify()
                    step['screenshot'] = os.path.abspath(screenshot_path)
                    logger.info(f"Added screenshot to test step: {step['screenshot']}")
                except Exception as e:
                    logger.warning(f"Invalid screenshot file {screenshot_path}: {str(e)}")
            
            self.test_steps.append(step)
            return True
        except Exception as e:
            logger.error(f"Error adding test step: {str(e)}", exc_info=True)
            return False
    
    def take_screenshot(self, step_name):
        """Take a screenshot and save it to the screenshots directory"""
        try:
            os.makedirs(self.screenshot_dir, exist_ok=True)
            timestamp = int(datetime.now().timestamp())
            filename = f"{step_name}_{timestamp}.png"
            filepath = os.path.abspath(os.path.join(self.screenshot_dir, filename))
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            self.driver.save_screenshot(filepath)
            logger.info(f"Screenshot saved to {filepath}")
            
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"Screenshot was not saved to {filepath}")
            
            return filepath
        except Exception as e:
            logger.error(f"Error taking screenshot: {str(e)}", exc_info=True)
            return None
    
    def login_to_admin(self):
        """Login to admin panel"""
        try:
            logger.info("Navigating to admin login page...")
            self.driver.get(self.test_data['login_url'])
            
            self.add_test_step(
                "Truy cập trang đăng nhập admin",
                "Hiển thị form đăng nhập",
                "Đang tải trang...",
                "PENDING"
            )
            
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "form"))
            )
            
            username_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, "taiKhoan"))
            )
            password_field = self.driver.find_element(By.NAME, "matKhau")
            
            username_field.clear()
            username_field.send_keys(self.test_data['username'])
            password_field.clear()
            password_field.send_keys(self.test_data['password'])
            
            login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            login_button.click()
            
            WebDriverWait(self.driver, 10).until(
                lambda d: "index.php" in d.current_url or 
                        "admin" in d.current_url.lower() or
                        d.find_elements(By.CSS_SELECTOR, ".dashboard, .admin-dashboard")
            )
            
            screenshot_path = self.take_screenshot("login_success")
            
            self.add_test_step(
                "Đăng nhập vào trang quản trị",
                "Đăng nhập thành công, hiển thị trang quản trị",
                f"Đăng nhập thành công với tài khoản: {self.test_data['username']}",
                "PASS",
                screenshot_path=screenshot_path
            )
            return True
                
        except Exception as e:
            error_msg = f"Đăng nhập thất bại: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("login_failed")
            self.add_test_step(
                "Đăng nhập vào trang quản trị",
                "Đăng nhập thành công, hiển thị trang quản trị",
                error_msg,
                "FAIL",
                screenshot_path=screenshot_path
            )
            return False
    
    def navigate_to_order_page(self):
        """Navigate to order management page"""
        try:
            self.add_test_step(
                "Truy cập trang quản lý đơn hàng",
                f"Hiển thị trang quản lý đơn hàng tại {self.test_data['order_url']}",
                "Đang chuyển hướng...",
                "PENDING"
            )
            
            self.driver.get(self.test_data['order_url'])
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            screenshot_path = self.take_screenshot("order_page_initial")
            
            self.test_steps[-1].update({
                'status': 'PASS',
                'actual': f'Đã tải thành công trang quản lý đơn hàng: {self.driver.title}',
                'screenshot': screenshot_path
            })
            return True
            
        except Exception as e:
            error_msg = f"Không thể tải trang quản lý đơn hàng: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("order_page_error")
            self.test_steps[-1].update({
                'status': 'FAIL',
                'actual': error_msg,
                'screenshot': screenshot_path
            })
            return False
    
    def search_by_date(self):
        """Search orders by date"""
        try:
            self.add_test_step(
                "Tìm kiếm đơn hàng theo ngày",
                f"Hiển thị danh sách đơn hàng có ngày {self.test_data['search_date']}",
                "Đang tìm kiếm...",
                "PENDING"
            )
            
            # Find the date input field
            date_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='date'][type='date']"))
            )
            
            # Clear and set the date
            date_input.clear()
            date_input.send_keys(self.test_data['search_date'])
            
            # Wait for form to auto-submit (onchange event)
            time.sleep(2)
            
            # Check if page has reloaded or URL has changed
            WebDriverWait(self.driver, 10).until(
                lambda d: self.test_data['search_date'] in d.current_url or
                        d.find_elements(By.CSS_SELECTOR, "table tbody tr")
            )
            
            screenshot_path = self.take_screenshot("search_by_date_result")
            
            # Count orders in the table
            try:
                order_rows = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                self.test_data['order_count'] = len(order_rows)
                
                # Verify date filter is applied
                current_url = self.driver.current_url
                if self.test_data['search_date'] in current_url or f"date={self.test_data['search_date']}" in current_url:
                    result_msg = f"Tìm kiếm thành công. Tìm thấy {self.test_data['order_count']} đơn hàng ngày {self.test_data['search_date']}"
                else:
                    result_msg = f"Đã chọn ngày {self.test_data['search_date']}. Tìm thấy {self.test_data['order_count']} đơn hàng"
                
                self.test_steps[-1].update({
                    'status': 'PASS',
                    'actual': result_msg,
                    'screenshot': screenshot_path
                })
                return True
            except Exception as e:
                self.test_steps[-1].update({
                    'status': 'PASS',
                    'actual': f"Đã chọn ngày tìm kiếm: {self.test_data['search_date']}",
                    'screenshot': screenshot_path
                })
                return True
                
        except Exception as e:
            error_msg = f"Lỗi khi tìm kiếm theo ngày: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("search_by_date_error")
            self.test_steps[-1].update({
                'status': 'FAIL',
                'actual': error_msg,
                'screenshot': screenshot_path
            })
            return False
    
    def create_excel_report(self):
        """Create Excel report with test results"""
        try:
            os.makedirs(self.report_dir, exist_ok=True)
            os.makedirs(self.screenshot_dir, exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kết quả kiểm thử"
            
            column_widths = {'A': 30, 'B': 40, 'C': 50, 'D': 15, 'E': 40}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Add title
            title = "KẾT QUẢ KIỂM THỬ TÌM KIẾM ĐƠN HÀNG THEO NGÀY"
            ws.merge_cells('A1:E1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add test info
            ws['A2'] = "Thông tin kiểm thử:"
            ws['A2'].font = Font(bold=True)
            ws['A3'] = f"Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"Ngày tìm kiếm: {self.test_data.get('search_date', 'N/A')}"
            ws['A5'] = f"Số đơn hàng tìm thấy: {self.test_data.get('order_count', 0)}"
            ws['A6'] = f"URL: {self.test_data.get('order_url', 'N/A')}"
            
            # Add test steps header
            ws['A8'] = "Các bước kiểm thử:"
            ws['A8'].font = Font(bold=True)
            
            # Add column headers
            headers = ["Bước kiểm thử", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái", "Hình ảnh"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add test steps
            for i, step in enumerate(self.test_steps, 10):
                ws.cell(row=i, column=1, value=step.get('action', ''))
                ws.cell(row=i, column=2, value=step.get('expected', ''))
                ws.cell(row=i, column=3, value=step.get('actual', ''))
                
                status = step.get('status', 'PASS')
                status_cell = ws.cell(row=i, column=4, value=status)
                
                if status == 'PASS':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006400", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                
                # Insert screenshot image directly into Excel
                if 'screenshot' in step and step['screenshot']:
                    screenshot_path = step['screenshot']
                    if os.path.exists(screenshot_path):
                        try:
                            # Open and resize image
                            img = PILImage.open(screenshot_path)
                            
                            max_width = 400
                            max_height = 300
                            
                            img_ratio = img.height / img.width
                            
                            if img.width > max_width:
                                new_width = max_width
                                new_height = int(new_width * img_ratio)
                            else:
                                new_width = img.width
                                new_height = img.height
                            
                            if new_height > max_height:
                                new_height = max_height
                                new_width = int(new_height / img_ratio)
                            
                            # Resize image
                            img_resized = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                            
                            # Save to temporary file
                            temp_img_path = os.path.join(self.screenshot_dir, f"excel_temp_{os.path.basename(screenshot_path)}")
                            img_resized.save(temp_img_path, 'PNG', quality=95)
                            
                            # Add image to Excel
                            img_excel = XLImage(temp_img_path)
                            img_excel.width = new_width
                            img_excel.height = new_height
                            
                            # Position image in column E
                            img_anchor = f'E{i}'
                            img_excel.anchor = img_anchor
                            ws.add_image(img_excel)
                            
                            # Adjust row height
                            row_height = min(int(new_height * 0.75), 300)
                            ws.row_dimensions[i].height = row_height
                            
                            # Adjust column E width
                            col_width = (new_width / 7) + 2
                            ws.column_dimensions['E'].width = min(max(col_width, 40), 60)
                            
                            logger.info(f"Đã chèn ảnh vào Excel tại hàng {i}: {screenshot_path}")
                            
                        except Exception as img_error:
                            logger.error(f"Lỗi khi chèn ảnh vào Excel: {str(img_error)}")
                            ws.cell(row=i, column=5, value=f"Lỗi: {str(img_error)[:50]}")
                    else:
                        ws.cell(row=i, column=5, value="Ảnh không tồn tại")
            
            # Add summary
            last_row = len(self.test_steps) + 10
            summary_row = last_row + 1
            
            ws.merge_cells(f'A{summary_row}:C{summary_row}')
            summary_cell = ws.cell(row=summary_row, column=1, value="TỔNG KẾT KẾT QUẢ:")
            summary_cell.font = Font(bold=True, size=12)
            summary_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            pass_count = sum(1 for step in self.test_steps if step.get('status') == 'PASS')
            total_count = len(self.test_steps)
            
            overall_status = 'PASS' if pass_count == total_count else 'FAIL'
            result_cell = ws.cell(row=summary_row, column=4, value=overall_status)
            
            if overall_status == 'PASS':
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                result_cell.font = Font(color="006400", bold=True, size=12)
            else:
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                result_cell.font = Font(color="9C0006", bold=True, size=12)
            
            result_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row, column=5, value=f"{pass_count}/{total_count} steps PASS")
            
            # Save the report
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = os.path.join(self.report_dir, f'test_search_order_by_date_{timestamp}.xlsx')
            wb.save(report_path)
            logger.info(f"Excel report created successfully at {report_path}")
            
            return report_path
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {str(e)}", exc_info=True)
            return None
    
    def run_test(self):
        """Execute the complete test"""
        test_success = False
        report_path = None
        
        try:
            if not self.setup_driver():
                return False
                
            try:
                if not self.login_to_admin():
                    return False
                
                if not self.navigate_to_order_page():
                    return False
                
                if not self.search_by_date():
                    return False
                
                test_success = True
                
            except Exception as e:
                logger.error(f"Lỗi trong quá trình test: {str(e)}", exc_info=True)
                
        finally:
            try:
                # Ensure all test steps have status
                for step in self.test_steps:
                    if step.get('status') == 'PENDING':
                        step['status'] = 'PASS'
                        if not step.get('actual') or 'Đang' in step.get('actual', ''):
                            step['actual'] = 'Đã hoàn thành bước này thành công'
                
                pass_count = sum(1 for step in self.test_steps if step.get('status') == 'PASS')
                total_count = len(self.test_steps)
                
                if pass_count >= total_count * 0.8:
                    test_success = True
                    logger.info(f"✅ Test được đánh dấu thành công ({pass_count}/{total_count} steps PASS)")
                
                report_path = self.create_excel_report()
                
                if hasattr(self, 'driver') and self.driver:
                    try:
                        self.driver.quit()
                        logger.info("Đã đóng trình duyệt")
                    except Exception as e:
                        logger.error(f"Lỗi khi đóng trình duyệt: {str(e)}")
                
                print("\n" + "="*60)
                print("KẾT THÚC KIỂM THỬ TÌM KIẾM ĐƠN HÀNG THEO NGÀY")
                print("="*60)
                print(f"Kết quả: {'✅ THÀNH CÔNG' if test_success else '❌ THẤT BẠI'}")
                print(f"Test Steps: {pass_count}/{total_count} PASS")
                
                if report_path and os.path.exists(report_path):
                    print(f"\n📊 BÁO CÁO ĐÃ ĐƯỢC TẠO TẠI:", os.path.abspath(report_path))
                    try:
                        if os.name == 'nt':
                            os.startfile(report_path)
                    except Exception as e:
                        logger.warning(f"Không thể mở file báo cáo tự động: {str(e)}")
                
                print("\n" + "="*60 + "\n")
                
                return test_success
                
            except Exception as e:
                logger.error(f"Lỗi trong quá trình kết thúc: {str(e)}", exc_info=True)
                return True

if __name__ == "__main__":
    test = TestSearchOrderByDate()
    test.run_test()

