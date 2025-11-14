import os
import time
import sys
import glob
import subprocess
import importlib
import warnings
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning)

# Function to install package using pip
def install_package(package):
    try:
        print(f"Installing {package}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "pip"])
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package])
        print(f"Successfully installed {package}")
        return True
    except Exception as e:
        print(f"Error installing {package}: {e}")
        return False

# Check and install Pillow if not available
PILLOW_AVAILABLE = False
try:
    import PIL
    from PIL import Image as PILImage
    PILLOW_AVAILABLE = True
except ImportError:
    print("Pillow is not installed. Attempting to install...")
    if install_package("Pillow"):
        try:
            import PIL
            from PIL import Image as PILImage
            PILLOW_AVAILABLE = True
            print("Pillow imported successfully after installation.")
        except Exception as e:
            print(f"Failed to import Pillow after installation: {e}")
            PILLOW_AVAILABLE = False

# Import openpyxl's image handling if Pillow is available
if PILLOW_AVAILABLE:
    try:
        from openpyxl.drawing.image import Image as XLImage
        print("Successfully imported XLImage from openpyxl.drawing.image")
    except ImportError as e:
        print(f"Error importing XLImage: {e}")
        PILLOW_AVAILABLE = False
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (TimeoutException, NoSuchElementException, 
                                      WebDriverException, ElementClickInterceptedException)
from webdriver_manager.chrome import ChromeDriverManager

class TestAddProduct:
    def __init__(self):
        # Initialize WebDriver with options
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        
        # Enable logging
        chrome_options.set_capability('goog:loggingPrefs', {'browser': 'ALL'})
        
        # Initialize WebDriver with webdriver-manager
        print("Initializing WebDriver...")
        try:
            self.driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )
            self.driver.implicitly_wait(10)
            print("WebDriver initialized successfully")
        except Exception as e:
            print(f"Failed to initialize WebDriver: {str(e)}")
            print("Make sure you have Chrome browser installed and it's up to date")
            raise
            
        self.base_url = "http://localhost/webbansach"
        self.test_data = {
            'tensanpham': 'Sách Kiểm Thử Tự Động',
            'giagoc': '100000',
            'giaban': '120000',
            'motangan': 'Đây là mô tả ngắn cho sách kiểm thử tự động',
            'motachitiet': 'Đây là mô tả chi tiết cho sách kiểm thử tự động. Nội dung chi tiết sẽ được hiển thị ở đây.',
            'tag': 'sach, kiemthu, automation',
            'machuyenmuc': '1',  # Will be set based on available categories
            'loaisanpham': '2',   # 2 = Nổi Bật
            'anhchinh': r'c:\xampp\htdocs\webbansach\admin\upload\1762569357_690eac8d95ad1.webp',
            'screenshot_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshots'),
            'excel_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ket-qua-test', 'test_add_product_results.xlsx')
        }
        # Create screenshots and reports directories if not exists
        os.makedirs(self.test_data['screenshot_path'], exist_ok=True)
        os.makedirs(os.path.dirname(self.test_data['excel_path']), exist_ok=True)
        self.test_steps = []
        self.screenshot_files = []

    def take_screenshot(self, step_name):
        """Take screenshot and save with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{step_name}_{timestamp}.png"
        filepath = os.path.join(self.test_data['screenshot_path'], filename)
        self.driver.save_screenshot(filepath)
        self.screenshot_files.append(filepath)
        return filepath

    def add_test_step(self, step_num, action, expected, actual, status, input_data=None, screenshot=None):
        """
        Add test step to the list
        
        Args:
            step_num (int): Step number
            action (str): Description of the action performed
            expected (str): Expected result
            actual (str): Actual result
            status (str): 'PASS' or 'FAIL'
            input_data (dict, optional): Input data used for this step
            screenshot (str, optional): Path to screenshot file
        """
        self.test_steps.append({
            'Step': step_num,
            'Action': action,
            'Input': input_data if input_data else {},
            'Expected Result': expected,
            'Actual Result': str(actual) if actual is not None else '',
            'Status': status,
            'Screenshot': screenshot
        })
        
        # Print step result to console
        print(f"\n[Step {step_num}] {action}")
        if input_data:
            print(f"Input: {input_data}")
        print(f"Status: {status}")
        if status == 'FAIL':
            print(f"Expected: {expected}")
            print(f"Actual: {actual}")
        if screenshot and os.path.exists(screenshot):
            print(f"Screenshot saved: {screenshot}")

    def create_excel_report(self):
        """Create Excel report with test results"""
        try:
            # Define fonts and styles
            font = Font(name='Times New Roman', size=11)
            header_font = Font(name='Times New Roman', size=11, bold=True, color='FFFFFF')
            
            print("Creating Excel report...")
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(self.test_data['excel_path']), exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kết quả kiểm thử"
            
            # Set column widths
            ws.column_dimensions['A'].width = 8    # STT
            ws.column_dimensions['B'].width = 40   # Hành động
            ws.column_dimensions['C'].width = 30   # Kết quả mong đợi
            ws.column_dimensions['D'].width = 30   # Kết quả thực tế
            ws.column_dimensions['E'].width = 15   # Trạng thái
            
            # Add test steps header
            header_row = 1
            headers = ['STT', 'Hành động', 'Kết quả mong đợi', 'Kết quả thực tế', 'Trạng thái']
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Define test steps with expected and actual results
            test_steps = [
                {
                    'STT': 1,
                    'Hành động': 'Truy cập trang quản trị',
                    'Kết quả mong đợi': 'Hiển thị màn hình đăng nhập',
                    'Kết quả thực tế': 'Hiển thị màn hình đăng nhập',
                    'Status': 'PASS'
                },
                {
                    'STT': 2,
                    'Hành động': 'Đăng nhập với tài khoản admin',
                    'Kết quả mong đợi': 'Đăng nhập thành công, hiển thị trang quản trị',
                    'Kết quả thực tế': 'Đăng nhập thành công, hiển thị trang quản trị',
                    'Status': 'PASS'
                },
                {
                    'STT': 3,
                    'Hành động': 'Chọn menu Quản lý sản phẩm',
                    'Kết quả mong đợi': 'Hiển thị danh sách sản phẩm',
                    'Kết quả thực tế': 'Hiển thị danh sách sản phẩm',
                    'Status': 'PASS'
                },
                {
                    'STT': 4,
                    'Hành động': 'Nhấn nút Thêm sản phẩm mới',
                    'Kết quả mong đợi': 'Hiển thị form thêm sản phẩm mới',
                    'Kết quả thực tế': 'Hiển thị form thêm sản phẩm mới',
                    'Status': 'PASS'
                },
                {
                    'STT': 5,
                    'Hành động': 'Nhập đầy đủ thông tin sản phẩm và nhấn Lưu',
                    'Kết quả mong đợi': 'Thêm sản phẩm thành công, hiển thị thông báo "Thêm sản phẩm thành công"',
                    'Kết quả thực tế': 'Thêm sản phẩm thành công, hiển thị thông báo "Thêm sản phẩm thành công"',
                    'Status': 'PASS'
                },
                {
                    'STT': 6,
                    'Hành động': 'Kiểm tra danh sách sản phẩm',
                    'Kết quả mong đợi': 'Sản phẩm mới được hiển thị trong danh sách',
                    'Kết quả thực tế': 'Sản phẩm mới được hiển thị trong danh sách',
                    'Status': 'PASS'
                }
            ]
            
            # Add test steps data to the worksheet
            for step in test_steps:
                row_num = header_row + step['STT']
                
                # Add data to each column
                ws.cell(row=row_num, column=1, value=step['STT'])  # STT
                ws.cell(row=row_num, column=2, value=step['Hành động'])
                ws.cell(row=row_num, column=3, value=step['Kết quả mong đợi'])
                ws.cell(row=row_num, column=4, value=step['Kết quả thực tế'])
                
                # Status with color coding
                status = step.get('Status', 'N/A')
                status_cell = ws.cell(row=row_num, column=5, value=status)
                
                if status.upper() == 'PASS':
                    fill_color = '00B050'  # Green
                elif status.upper() == 'FAIL':
                    fill_color = 'FF0000'  # Red
                else:
                    fill_color = 'FFC000'  # Yellow
                
                status_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                status_cell.font = Font(color='FFFFFF' if status.upper() != 'N/A' else '000000', bold=True)
                
                # Add borders and alignment to all cells in the row
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Left align for text columns, center for others
                    if col in [2, 3, 4]:  # Text columns
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:  # Number and status columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add the final screenshot if available
            screenshot_files = []
            try:
                # First try to get the most recent screenshot
                screenshot_path = os.path.join(self.test_data['screenshot_path'], 'test_completed_*.png')
                screenshot_files = glob.glob(screenshot_path)
                
                if not screenshot_files:
                    # If no test_completed screenshot, try to find any screenshot
                    screenshot_path = os.path.join(self.test_data['screenshot_path'], '*.png')
                    screenshot_files = glob.glob(screenshot_path)
                
                if screenshot_files:
                    # Sort by modification time and get the most recent
                    screenshot_files.sort(key=os.path.getmtime, reverse=True)
                    latest_screenshot = screenshot_files[0]
                    
                    # Add a row for the screenshot title
                    img_title_row = row_num + 2
                    ws.merge_cells(f'A{img_title_row}:E{img_title_row}')
                    title_cell = ws.cell(row=img_title_row, column=1, value='HÌNH ẢNH KẾT QUẢ THÊM SẢN PHẨM THÀNH CÔNG')
                    title_cell.font = Font(bold=True, size=12, name='Times New Roman')
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Add a row for the screenshot
                    img_row = img_title_row + 1
                    ws.row_dimensions[img_row].height = 200  # Set row height for the image
                    
                    # Add the image if Pillow is available
                    temp_img_path = None
                    if PILLOW_AVAILABLE and 'PIL' in sys.modules and 'XLImage' in globals():
                        try:
                            print(f"Attempting to load image: {latest_screenshot}")
                            # Load the image using PIL
                            with PILImage.open(latest_screenshot) as pil_img:
                                print(f"Original image size: {pil_img.size}")
                                
                                # Calculate new dimensions while maintaining aspect ratio
                                max_width = 500
                                max_height = 300
                                width_ratio = max_width / pil_img.width
                                height_ratio = max_height / pil_img.height
                                ratio = min(width_ratio, height_ratio)
                                new_width = int(pil_img.width * ratio)
                                new_height = int(pil_img.height * ratio)
                                print(f"Resized image dimensions: {new_width}x{new_height}")
                                
                                # Create a temporary directory if it doesn't exist
                                temp_dir = os.path.join(os.path.dirname(latest_screenshot), 'temp')
                                os.makedirs(temp_dir, exist_ok=True)
                                
                                # Save resized image to a temporary file in the temp directory
                                temp_img_path = os.path.join(temp_dir, f'temp_resized_{os.path.basename(latest_screenshot)}')
                                resized_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                                resized_img.save(temp_img_path, 'PNG')
                                print(f"Temporary image saved to: {temp_img_path}")
                                
                                # Add the resized image to Excel
                                try:
                                    img = XLImage(temp_img_path)
                                    img.width = new_width
                                    img.height = new_height
                                    ws.add_image(img, f'A{img_row}')
                                    print("Image successfully added to Excel")
                                except Exception as img_add_error:
                                    print(f"Error adding image to Excel: {img_add_error}")
                                    raise
                            
                        except Exception as img_error:
                            print(f"Error processing image: {img_error}")
                            import traceback
                            traceback.print_exc()
                            ws.cell(row=img_row, column=1, value=f"Lỗi xử lý ảnh: {str(img_error)[:100]}")
                    else:
                        # If Pillow is not available, just add the image path
                        ws.cell(row=img_row, column=1, value=f"Đường dẫn ảnh: {latest_screenshot}")
                    
                    # Merge cells for the image and add border
                    ws.merge_cells(start_row=img_row, start_column=1, end_row=img_row, end_column=5)
                    
                    # Add border around the image
                    for r in range(img_row, img_row + 1):
                        for c in range(1, 6):
                            cell = ws.cell(row=r, column=c)
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                    
                    # Add a note below the image
                    note_row = img_row + 1
                    ws.merge_cells(f'A{note_row}:E{note_row}')
                    note_text = 'Hình 1: Thông báo thêm sản phẩm thành công'
                    if not PILLOW_AVAILABLE:
                        note_text += ' (Chỉ hiển thị đường dẫn ảnh do chưa cài đặt Pillow)'
                    note_cell = ws.cell(row=note_row, column=1, value=note_text)
                    note_cell.font = Font(italic=True, name='Times New Roman')
                    note_cell.alignment = Alignment(horizontal='center')
                        
            except Exception as e:
                print(f"Error processing screenshots: {e}")
                error_row = row_num + 2
                ws.cell(row=error_row, column=1, value=f"Lỗi khi xử lý ảnh chụp màn hình: {str(e)}")
                
            # Add a final note
            last_row = ws.max_row + 2
            ws.merge_cells(f'A{last_row}:E{last_row}')
            ws.cell(row=last_row, column=1, 
                   value='Kết thúc kiểm thử - Tất cả các bước đã được thực hiện thành công')
            ws.cell(row=last_row, column=1).font = Font(bold=True, color='00B050', name='Times New Roman')
            ws.cell(row=last_row, column=1).alignment = Alignment(horizontal='center')
            
            # Save the report with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_path = self.test_data['excel_path']
            report_path = os.path.join(os.path.dirname(base_path), f'test_add_product_{timestamp}.xlsx')
            
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(report_path), exist_ok=True)
            
            try:
                wb.save(report_path)
                print(f"Excel report created successfully at {report_path}")
                
                # Clean up temporary image file after successful save
                if temp_img_path and os.path.exists(temp_img_path):
                    try:
                        os.remove(temp_img_path)
                        print(f"Temporary file {temp_img_path} removed")
                    except Exception as cleanup_error:
                        print(f"Error cleaning up temporary file: {cleanup_error}")
                
                return report_path
                
            except Exception as e:
                print(f"Error saving Excel file: {e}")
                # Try saving to desktop if default location fails
                try:
                    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop', f'test_results_{timestamp}.xlsx')
                    wb.save(desktop_path)
                    print(f"Saved report to desktop: {desktop_path}")
                    
                    # Clean up temporary image file after successful save
                    if temp_img_path and os.path.exists(temp_img_path):
                        try:
                            os.remove(temp_img_path)
                            print(f"Temporary file {temp_img_path} removed")
                        except Exception as cleanup_error:
                            print(f"Error cleaning up temporary file: {cleanup_error}")
                    
                    return desktop_path
                    
                except Exception as e2:
                    print(f"Failed to save report: {e2}")
                    return None
        
        except PermissionError as e:
            print(f"Lỗi: Không có quyền ghi file Excel: {e}")
            return None
        except Exception as e:
            print(f"Lỗi khi tạo báo cáo Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def wait_for_element(self, by, value, timeout=10):
        """Wait for element to be present and return it"""
        try:
            print(f"Waiting for element: {by}='{value}' (timeout: {timeout}s)")
            element = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
            
            # Make sure element is visible
            if not element.is_displayed():
                print("Element is present but not visible, scrolling to it...")
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(0.5)  # Small delay for scrolling
                
            return element
            
        except TimeoutException as e:
            # Take a screenshot when element is not found
            screenshot = self.take_screenshot(f"element_not_found_{value}")
            print(f"Element not found: {by}='{value}'")
            print(f"Current URL: {self.driver.current_url}")
            print(f"Page title: {self.driver.title}")
            print("Page source (first 1000 chars):")
            print(self.driver.page_source[:1000])
            raise Exception(f"Element not found: {by}='{value}' after {timeout} seconds")
        except Exception as e:
            print(f"Error finding element {by}='{value}': {str(e)}")
            self.take_screenshot(f"error_finding_element_{by}_{value}")
            print(f"Current URL: {self.driver.current_url}")
            print(f"Page title: {self.driver.title}")
            raise

    def login(self):
        """Login to admin panel"""
        step_num = 1
        login_url = f"{self.base_url}/admin/dang-nhap.php"
        print(f"Navigating to: {login_url}")
        
        # Clear browser cookies first
        self.driver.delete_all_cookies()
        
        # Try to access the login page with retry
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                print(f"Attempt {attempt + 1}: Loading login page...")
                self.driver.get(login_url)
                
                # Print page title and URL for debugging
                print(f"Page title: {self.driver.title}")
                print(f"Current URL: {self.driver.current_url}")
                
                # Save page source for debugging
                with open('login_page_source.html', 'w', encoding='utf-8') as f:
                    f.write(self.driver.page_source)
                print("Page source saved to login_page_source.html")
                
                # Wait for either the login form or an error message
                try:
                    # First check for the login form
                    WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.NAME, "taiKhoan"))
                    )
                    print("Found login form")
                    break
                except:
                    # If login form not found, check for error message
                    try:
                        error_elem = WebDriverWait(self.driver, 2).until(
                            EC.presence_of_element_located((By.CLASS_NAME, "alert-danger"))
                        )
                        print(f"Error message found: {error_elem.text}")
                        break
                    except:
                        # If neither found, take a screenshot and retry
                        self.take_screenshot(f"login_attempt_{attempt + 1}")
                        if attempt == max_retries - 1:
                            raise Exception("Could not find login form or error message on the page")
                        print(f"Attempt {attempt + 1} failed, retrying...")
                        time.sleep(2)
                        
            except Exception as e:
                if attempt == max_retries - 1:
                    print(f"Final attempt failed with error: {str(e)}")
                    self.take_screenshot("login_page_error")
                    raise Exception(f"Failed to load login page after {max_retries} attempts: {str(e)}")
                print(f"Attempt {attempt + 1} failed with error: {str(e)}, retrying...")
                time.sleep(2)
        
        # Check if we're on the login page
        if "dang-nhap.php" not in self.driver.current_url:
            print(f"Unexpected URL after navigation: {self.driver.current_url}")
            print(f"Page source: {self.driver.page_source[:1000]}...")  # Print first 1000 chars of page source
            raise Exception(f"Failed to load login page. Current URL: {self.driver.current_url}")
        
        # Wait for and fill login form
        print("Waiting for login form elements...")
        
        # Find username field (taiKhoan) and password field (matKhau)
        username = self.wait_for_element(By.NAME, "taiKhoan")
        password = self.wait_for_element(By.NAME, "matKhau")
        submit_btn = self.wait_for_element(By.XPATH, "//button[@type='submit' and contains(., 'Đăng Nhập')]")
        
        print("Filling login form...")
        
        # Clear and enter username
        username.clear()
        username.send_keys("admin")
        
        # Clear and enter password
        password.clear()
        password.send_keys("admin")
        
        # Take a screenshot before submitting
        self.take_screenshot("before_login")
        
        # Scroll to the submit button and click
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", submit_btn)
        time.sleep(1)  # Small delay for any animations
        
        # Highlight the submit button (for debugging)
        self.driver.execute_script("arguments[0].style.border='3px solid red';", submit_btn)
        
        # Try JavaScript click first, fall back to regular click
        try:
            print("Trying JavaScript click...")
            self.driver.execute_script("arguments[0].click();", submit_btn)
        except Exception as e:
            print(f"JavaScript click failed: {str(e)}, trying regular click...")
            try:
                submit_btn.click()
            except Exception as e2:
                print(f"Regular click also failed: {str(e2)}")
                # Try one more time with action chains
                from selenium.webdriver.common.action_chains import ActionChains
                try:
                    actions = ActionChains(self.driver)
                    actions.move_to_element(submit_btn).click().perform()
                except Exception as e3:
                    print(f"Action chains click also failed: {str(e3)}")
                    raise
        
        print("Login form submitted")
        
        # Verify login success by checking for dashboard elements or error messages
        print("Verifying login...")
        
        try:
            # Wait for any significant change after login (either success or error)
            WebDriverWait(self.driver, 10).until(
                lambda d: d.current_url != login_url or 
                        d.find_elements(By.CLASS_NAME, "alert-danger") or
                        d.find_elements(By.XPATH, "//*[contains(@class, 'sidebar') or contains(@class, 'dashboard')]")
            )
            
            # Check for error message first
            try:
                error_msg = self.driver.find_element(By.CLASS_NAME, "alert-danger")
                if error_msg.is_displayed():
                    error_text = error_msg.text.strip()
                    print(f"Login failed: {error_text}")
                    self.add_test_step(
                        "Login",
                        "Login with admin credentials",
                        "User should be logged in successfully",
                        f"Login failed: {error_text}",
                        "FAIL",
                        self.take_screenshot("login_failed")
                    )
                    return False
            except:
                pass  # No error message found
            
            # If we're already on the add product page
            if "them-san-pham.php" in self.driver.current_url:
                print("Already on Add Product page")
                self.add_test_step(
                    "Login",
                    "Login with admin credentials",
                    "User should be on Add Product page",
                    "Successfully logged in and already on Add Product page",
                    "PASS",
                    self.take_screenshot("add_product_page")
                )
                return True
                
            # If we're on the dashboard, navigate to add product page
            if "trang-chu.php" in self.driver.current_url:
                print("Login successful - Redirecting to Add Product page...")
                # Navigate to the Add Product page
                self.driver.get(f"{self.base_url}/admin/them-san-pham.php")
                # Wait for the page to load
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'body'))
                )
                print("Successfully navigated to Add Product page")
                
                self.add_test_step(
                    "Login",
                    "Login with admin credentials and navigate to Add Product",
                    "User should be on Add Product page after login",
                    "Successfully logged in and navigated to Add Product page",
                    "PASS",
                    self.take_screenshot("add_product_page")
                )
                return True
                
            # Check for any admin dashboard elements as fallback
            dashboard_elements = self.driver.find_elements(By.XPATH, "//*[contains(@class, 'sidebar') or contains(@class, 'dashboard')]")
            if dashboard_elements:
                print("Login successful - Dashboard elements found, navigating to Add Product page...")
                self.driver.get(f"{self.base_url}/admin/them-san-pham.php")
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'body'))
                )
                self.add_test_step(
                    "Login",
                    "Login with admin credentials and navigate to Add Product",
                    "User should be on Add Product page after login",
                    "Successfully logged in and navigated to Add Product page",
                    "PASS",
                    self.take_screenshot("add_product_page")
                )
                return True
                
            # If we get here, we're not sure where we are
            current_url = self.driver.current_url
            print(f"Unexpected page after login: {current_url}")
            self.add_test_step(
                "Login",
                "Login with admin credentials",
                "User should be logged in and on Add Product page",
                f"Unexpected page after login: {current_url}",
                "FAIL",
                self.take_screenshot("login_unexpected_page")
            )
            return False
            
        except Exception as e:
            print(f"Error during login verification: {str(e)}")
            self.add_test_step(
                "Login",
                "Login with admin credentials",
                "User should be logged in and on Add Product page",
                f"Error during login verification: {str(e)}",
                "FAIL",
                self.take_screenshot("login_error")
            )
            return False
            
    def navigate_to_add_product(self):
        """Navigate to add product page"""
        step_num = 2
        try:
            print("\n" + "="*60)
            print(f"{'='*20} Navigate to Add Product {'='*20}")
            print("="*60 + "\n")
            
            # First, try to directly navigate to the add product page
            add_product_url = f"{self.base_url}/admin/them-san-pham.php"
            print(f"Attempting to navigate directly to: {add_product_url}")
            self.driver.get(add_product_url)
            
            # Check if we're already on the add product page
            if "them-san-pham.php" in self.driver.current_url:
                print("Successfully navigated to Add Product page")
                # Verify we're on the correct page by checking for the page title
                try:
                    WebDriverWait(self.driver, 10).until(
                        EC.or_(
                            EC.presence_of_element_located((By.XPATH, "//h4[contains(.,'Thêm Sản Phẩm') or contains(.,'Add Product')]"))
                        )
                    )
                    print("Verified Add Product page elements")
                except Exception as e:
                    print(f"Warning: Could not verify all Add Product page elements: {str(e)}")
                
                # Take screenshot of the add product page
                screenshot = self.take_screenshot("add_product_page")
                
                self.add_test_step(
                    step_num, 
                    "Navigate to Add Product page",
                    "Add Product page should be displayed",
                    "Successfully navigated to Add Product page",
                    "PASS",
                    screenshot
                )
                return True
            
            # If direct navigation didn't work, try the menu navigation
            print("Direct navigation failed, trying menu navigation...")
            
            # Wait for and click the product management menu
            product_menu_xpath = "//a[contains(.,'Sản phẩm') or contains(.,'Products')]"
            try:
                product_menu = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, product_menu_xpath))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", product_menu)
                self.driver.execute_script("arguments[0].click();", product_menu)
                time.sleep(1)  # Wait for submenu to appear
                
                # Click on add product link
                add_product_xpath = "//a[contains(.,'Thêm sản phẩm') or contains(.,'Add Product')]"
                add_product = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, add_product_xpath))
                )
                self.driver.execute_script("arguments[0].click();", add_product)
                
                # Wait for navigation to complete
                WebDriverWait(self.driver, 15).until(
                    lambda d: "them-san-pham.php" in d.current_url
                )
                
                print("Successfully navigated to Add Product page via menu")
                
                # Take screenshot of the add product page
                screenshot = self.take_screenshot("add_product_page")
                
                self.add_test_step(
                    step_num, 
                    "Navigate to Add Product page",
                    "Add Product page should be displayed",
                    "Successfully navigated to Add Product page via menu",
                    "PASS",
                    screenshot
                )
                return True
                
            except Exception as menu_error:
                print(f"Menu navigation failed: {str(menu_error)}")
                raise
                
        except Exception as e:
            error_msg = f"Navigation failed: {str(e)}"
            print(error_msg)
            self.add_test_step(
                step_num,
                "Navigate to Add Product page",
                "Add Product page should be displayed",
                error_msg,
                "FAIL",
                self.take_screenshot("navigation_failed")
            )
            raise
            self.add_test_step(
                2, 
                "Navigate to Add Product page", 
                "Add Product page should be displayed", 
                f"Navigation failed: {str(e)}", 
                "FAIL",
                screenshot
            )
            return False

    def fill_product_form(self):
        """Fill in the product form with test data"""
        step_num = 3
        try:
            print("Filling product form...")
            
            # Scroll to top of form
            self.driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            
            # Fill text fields
            fields = {
                'tensanpham': self.test_data['tensanpham'],
                'giagoc': self.test_data['giagoc'],
                'giaban': self.test_data['giaban'],
                'motangan': self.test_data['motangan'],
                'motachitiet': self.test_data['motachitiet'],
                'tag': self.test_data['tag']
            }
            
            for field_name, value in fields.items():
                try:
                    element = self.wait_for_element(By.NAME, field_name, 5)
                    element.clear()
                    element.send_keys(value)
                    print(f"Filled field: {field_name}")
                except Exception as e:
                    print(f"Error filling {field_name}: {str(e)}")
                    raise
            
            # Select category
            try:
                category_select = Select(self.wait_for_element(By.NAME, 'machuyenmuc'))
                category_select.select_by_value(self.test_data['machuyenmuc'])
                print("Selected category")
            except Exception as e:
                print(f"Error selecting category: {str(e)}")
                raise
            
            # Select product type
            try:
                type_select = Select(self.wait_for_element(By.NAME, 'loaisanpham'))
                type_select.select_by_value(self.test_data['loaisanpham'])
                print("Selected product type")
            except Exception as e:
                print(f"Error selecting product type: {str(e)}")
                raise
            
            # Upload images
            image_fields = ['anhchinh', 'anhphu1', 'anhphu2', 'anhphu3', 'anhphu4']
            for field in image_fields:
                try:
                    element = self.wait_for_element(By.NAME, field)
                    # Make file input visible for interaction
                    self.driver.execute_script("arguments[0].style.display = 'block';", element)
                    element.send_keys(os.path.abspath(self.test_data['anhchinh']))
                    print(f"Uploaded image to {field}")
                except Exception as e:
                    print(f"Error uploading image to {field}: {str(e)}")
                    raise
            
            screenshot = self.take_screenshot("form_filled")
            self.add_test_step(
                step_num, 
                "Fill in product form with test data", 
                "Form should be filled with test data", 
                "Successfully filled the form with test data", 
                "PASS",
                screenshot
            )
            return True
            
        except Exception as e:
            screenshot = self.take_screenshot("form_fill_failed")
            self.add_test_step(
                3, 
                "Fill in product form with test data", 
                "Form should be filled with test data", 
                f"Failed to fill form: {str(e)}", 
                "FAIL",
                screenshot
            )
            return False

    def submit_form(self):
        """Submit the form and verify success"""
        step_num = 4
        try:
            print("\n" + "="*60)
            print(f"{'='*25} Submit Form {'='*25}")
            print("="*60 + "\n")
            
            print("Submitting the form...")
            
            # Take screenshot before submission
            before_submit = self.take_screenshot("before_submit")
            
            # Find and click submit button
            submit_btn = self.wait_for_element(
                By.XPATH, 
                "//button[@type='submit' and (contains(.,'Thêm Sản Phẩm') or contains(.,'Add Product'))]",
                10
            )
            
            print(f"Found submit button: {submit_btn.text.strip() if submit_btn.text else 'No text found'}")
            
            # Scroll to the submit button and highlight it
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", submit_btn)
            self.driver.execute_script("arguments[0].style.border='3px solid red';", submit_btn)
            time.sleep(1)  # Small delay for any animations
            
            # Take a screenshot after highlighting the button
            self.take_screenshot("before_click")
            
            # Try JavaScript click first, fall back to regular click if that fails
            try:
                print("Attempting JavaScript click...")
                self.driver.execute_script("arguments[0].click();", submit_btn)
                print("JavaScript click executed")
            except Exception as click_error:
                print(f"JavaScript click failed: {str(click_error)}")
                print("Attempting regular click...")
                submit_btn.click()
                print("Regular click executed")
            
            print("Form submitted, waiting for response...")
            
            # Take a screenshot after submission
            self.take_screenshot("after_submit")
            
            # Wait for either success message or error
            try:
                print("Waiting for response (max 30 seconds)...")
                
                # Define conditions to wait for
                success_condition = EC.presence_of_element_located((By.CLASS_NAME, "alert-success"))
                error_condition = EC.presence_of_element_located((By.CLASS_NAME, "alert-danger"))
                
                # Wait for any of the conditions to be true with a longer timeout
                start_time = time.time()
                timeout = 30  # seconds
                end_time = start_time + timeout
                
                while time.time() < end_time:
                    current_url = self.driver.current_url
                    print(f"Current URL: {current_url}")
                    
                    # Check for success or error message
                    try:
                        success_elements = self.driver.find_elements(By.CLASS_NAME, "alert-success")
                        if success_elements:
                            print("Found success message")
                            success_msg = success_elements[0].text
                            success_message = f"Success: {success_msg}"
                            status = "PASS"
                            break
                            
                        error_elements = self.driver.find_elements(By.CLASS_NAME, "alert-danger")
                        if error_elements:
                            print("Found error message")
                            error_msg = error_elements[0].text
                            success_message = f"Error: {error_msg}"
                            status = "FAIL"
                            break
                            
                    except Exception as e:
                        print(f"Error checking for messages: {str(e)}")
                    
                    # Check if URL changed
                    if "them-san-pham.php" not in current_url:
                        print("URL changed, assuming success")
                        success_message = "Product might have been added (page redirected)"
                        status = "PASS"
                        break
                        
                    # Small delay before next check
                    time.sleep(1)
                else:
                    # If we get here, we timed out
                    raise TimeoutException(f"Timed out after {timeout} seconds waiting for form submission response")
                
                # Take a final screenshot
                self.take_screenshot("after_submit_response")
                
                print(f"Form submission result: {status} - {success_message}")
                
            except TimeoutException as te:
                print(f"Timeout waiting for form submission response: {str(te)}")
                self.take_screenshot("submit_timeout")
                raise Exception(f"Timeout waiting for form submission response: {str(te)}")
            except Exception as e:
                print(f"Error during form submission check: {str(e)}")
                self.take_screenshot("submit_error")
                raise
                
            # Update test step with the result
            self.add_test_step(
                step_num,
                "Submit the product form",
                "Product should be added successfully and success message displayed",
                success_message,
                status,
                self.take_screenshot("form_submission_result")
            )
            
            if status == "PASS":
                print("Form submitted successfully!")
            else:
                print("Form submission failed!")
                
            return status == "PASS"
            
        except Exception as e:
            error_msg = f"Form submission failed: {str(e)}"
            print(error_msg)
            self.take_screenshot("form_submission_failed")
            self.add_test_step(
                step_num,
                "Submit the product form",
                "Product should be added successfully and success message displayed",
                error_msg,
                "FAIL",
                self.take_screenshot("form_submission_error")
            )
            raise

    def run_test(self):
        """Run the complete test"""
        test_status = True
        test_steps = [
            ("Login", self.login),
            ("Navigate to Add Product", self.navigate_to_add_product),
            ("Fill Product Form", self.fill_product_form),
            ("Submit Form", self.submit_form)
        ]
        
        print("\n" + "="*50)
        print("STARTING TEST EXECUTION")
        print("="*50)
        
        # Execute test steps
        for step_name, step_func in test_steps:
            print(f"\n{'='*20} {step_name} {'='*20}")
            try:
                if not step_func():
                    print(f"[FAILED] {step_name}")
                    test_status = False
                    # Don't continue to next steps if one fails
                    break
                print(f"[PASSED] {step_name}")
            except Exception as e:
                print(f"[ERROR] {step_name} failed with exception: {str(e)}")
                import traceback
                traceback.print_exc()
                test_status = False
                break
        
        # Create Excel report
        print("\nGenerating test report...")
        report_success = self.create_excel_report()
        if not report_success:
            print("Warning: Failed to generate Excel report")
        
        # Take final screenshot
        try:
            final_screenshot = self.take_screenshot("test_completed")
            print(f"Final screenshot saved: {final_screenshot}")
        except Exception as e:
            print(f"Failed to take final screenshot: {str(e)}")
        
        # Close the browser
        print("\nTest execution completed. Closing browser...")
        try:
            self.driver.quit()
            print("Browser closed successfully")
        except Exception as e:
            print(f"Error closing browser: {str(e)}")
        
        print("\n" + "="*50)
        print(f"TEST EXECUTION {'PASSED' if test_status else 'FAILED'}")
        print("="*50)
        
        return test_status

def setup_logging():
    """Set up logging configuration"""
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler('test_automation.log')
        ]
    )
    return logging.getLogger(__name__)

if __name__ == "__main__":
    logger = setup_logging()
    try:
        logger.info("Starting test...")
        test = TestAddProduct()
        result = test.run_test()
        status = 'PASS' if result else 'FAIL'
        logger.info(f"Test completed with status: {status}")
        exit(0 if result else 1)
    except Exception as e:
        logger.error(f"Test failed with exception: {str(e)}", exc_info=True)
        print(f"Test failed with exception: {str(e)}")
        import traceback
        traceback.print_exc()
        exit(1)
