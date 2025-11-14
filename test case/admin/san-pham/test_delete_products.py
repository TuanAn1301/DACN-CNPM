import os
import sys
import logging
import time
import re
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
    ElementClickInterceptedException
)
import warnings

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("test_delete_products.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Set console output to UTF-8
if sys.stdout.encoding != 'utf-8':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

class TestDeleteProducts:
    def __init__(self):
        self.driver = None
        self.test_steps = []
        self.screenshot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshots')
        self.report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ket-qua-test')
        self.excel_path = os.path.join(self.report_dir, f'test_delete_products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        # Create necessary directories
        os.makedirs(self.screenshot_dir, exist_ok=True)
        os.makedirs(self.report_dir, exist_ok=True)
        
        # Test data
        self.test_data = {
            'admin_url': 'http://localhost/webbansach/admin',
            'products_url': 'http://localhost/webbansach/admin/san-pham.php',
            'username': 'admin',
            'password': 'admin',
            'max_products_to_delete': 1  # Chỉ xóa 1 sản phẩm có ID nhỏ nhất
        }
    
    def setup_driver(self):
        """Initialize the WebDriver"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--remote-debugging-port=9222")
            chrome_options.add_argument("--window-size=1920,1080")
            
            try:
                self.driver = webdriver.Chrome(options=chrome_options)
                self.driver.set_window_size(1920, 1080)
                self.driver.set_page_load_timeout(30)
                logger.info("WebDriver initialized successfully")
                self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", "WebDriver khởi tạo thành công", "PASS")
                return True
            except Exception as e:
                logger.error(f"Lỗi khi khởi tạo WebDriver: {str(e)}")
                try:
                    from webdriver_manager.chrome import ChromeDriverManager
                    from selenium.webdriver.chrome.service import Service as ChromeService
                    service = ChromeService(ChromeDriverManager().install())
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    self.driver.set_window_size(1920, 1080)
                    self.driver.set_page_load_timeout(30)
                    logger.info("WebDriver initialized successfully with webdriver_manager")
                    self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", "WebDriver khởi tạo thành công với webdriver_manager", "PASS")
                    return True
                except Exception as e2:
                    logger.error(f"Lỗi khi thử cách khởi tạo WebDriver thứ 2: {str(e2)}")
                    self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", f"Lỗi: {str(e2)}", "FAIL")
                    return False
        except Exception as e:
            logger.error(f"Lỗi không xác định khi khởi tạo WebDriver: {str(e)}", exc_info=True)
            self.add_test_step("Khởi tạo WebDriver", "WebDriver khởi tạo thành công", f"Lỗi: {str(e)}", "FAIL")
            return False
    
    def add_test_step(self, action, expected, actual, status, screenshot_path=None):
        """Add a test step with optional screenshot"""
        try:
            step = {
                'action': str(action) if action else "",
                'expected': str(expected) if expected else "",
                'actual': str(actual) if actual else "",
                'status': str(status) if status else "UNKNOWN",
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            if screenshot_path and os.path.isfile(screenshot_path):
                step['screenshot'] = os.path.abspath(screenshot_path)
                logger.info(f"Added screenshot to test step: {step['screenshot']}")
            
            self.test_steps.append(step)
            return True
        except Exception as e:
            logger.error(f"Error adding test step: {str(e)}", exc_info=True)
            return False
    
    def take_screenshot(self, step_name):
        """Take a screenshot and save it to the screenshots directory"""
        try:
            os.makedirs(self.screenshot_dir, exist_ok=True)
            timestamp = int(datetime.now().timestamp())
            filename = f"{step_name}_{timestamp}.png"
            filepath = os.path.abspath(os.path.join(self.screenshot_dir, filename))
            
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            self.driver.save_screenshot(filepath)
            logger.info(f"Screenshot saved to {filepath}")
            
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"Screenshot was not saved to {filepath}")
            
            if hasattr(self, 'test_steps') and self.test_steps:
                self.test_steps[-1]['screenshot'] = filepath
                
            return filepath
            
        except Exception as e:
            logger.error(f"Error taking screenshot: {str(e)}", exc_info=True)
            return None
    
    def login_to_admin(self):
        """Login to admin panel"""
        try:
            logger.info("Navigating to admin login page...")
            self.driver.get(self.test_data['admin_url'])
            
            try:
                # Wait for login form
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "form"))
                )
                
                # Find login form elements
                username_field = self.driver.find_element(By.NAME, "taiKhoan")
                password_field = self.driver.find_element(By.NAME, "matKhau")
                
                # Enter credentials
                username_field.clear()
                username_field.send_keys(self.test_data['username'])
                password_field.clear()
                password_field.send_keys(self.test_data['password'])
                
                # Submit form
                login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
                login_button.click()
                
                # Wait for login to complete
                WebDriverWait(self.driver, 10).until(
                    lambda d: "index.php" in d.current_url or 
                            "admin" in d.current_url.lower() or
                            d.find_elements(By.CSS_SELECTOR, ".dashboard, .admin-dashboard")
                )
                
                # Take screenshot after successful login
                screenshot_path = self.take_screenshot("login_success")
                
                self.add_test_step(
                    "Đăng nhập vào trang quản trị",
                    "Đăng nhập thành công, hiển thị trang quản trị",
                    "Đăng nhập thành công, hiển thị trang quản trị",
                    "PASS",
                    screenshot_path=screenshot_path
                )
                return True
                
            except Exception as e:
                error_screenshot = self.take_screenshot("login_error")
                logger.error(f"Login form error: {str(e)}")
                raise
                
        except Exception as e:
            error_msg = f"Đăng nhập thất bại: {str(e)}"
            logger.error(error_msg, exc_info=True)
            error_screenshot = self.take_screenshot("login_failed")
            self.add_test_step(
                "Đăng nhập vào trang quản trị",
                "Đăng nhập thành công, hiển thị trang quản trị",
                error_msg,
                "FAIL",
                screenshot_path=error_screenshot
            )
            return False
    
    def get_product_ids(self):
        """Get all product IDs from the products page, sorted from smallest to largest"""
        try:
            self.driver.get(self.test_data['products_url'])
            
            # Wait for products table to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            # Wait a bit for table to fully render
            time.sleep(1)
            
            # Find all product rows
            product_rows = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
            product_ids = []
            
            logger.info(f"Tìm thấy {len(product_rows)} dòng sản phẩm trong bảng")
            
            for row in product_rows:
                try:
                    product_id = None
                    
                    # Method 1: Tìm icon thùng rác trong cột "Hành Động" và lấy ID từ onclick
                    try:
                        # Tìm icon thùng rác trong hàng
                        trash_icon_selectors = [
                            "i[class*='trash']",
                            "i[class*='delete']",
                            "*[class*='trash']",
                            "*[class*='delete']",
                            "i.fa-trash, i.fa-trash-alt, i.fa-trash-o"
                        ]
                        
                        for selector in trash_icon_selectors:
                            try:
                                trash_icon = row.find_element(By.CSS_SELECTOR, selector)
                                # Tìm phần tử cha (link) chứa icon thùng rác
                                delete_link = trash_icon.find_element(By.XPATH, "./ancestor::a[1] | ./parent::a")
                                onclick_text = delete_link.get_attribute('onclick')
                                if onclick_text:
                                    # Extract ID from onclick: xoa('123') or xoa("123")
                                    match = re.search(r"xoa\(['\"]?(\d+)['\"]?\)", onclick_text)
                                    if match:
                                        product_id = int(match.group(1))
                                        break
                            except:
                                continue
                        
                        # Nếu không tìm thấy qua icon, thử tìm link trong cột "Hành Động" (cột cuối)
                        if product_id is None:
                            try:
                                cells = row.find_elements(By.TAG_NAME, "td")
                                if cells:
                                    # Cột cuối thường là cột "Hành Động"
                                    action_cell = cells[-1]
                                    delete_link = action_cell.find_element(By.CSS_SELECTOR, "a[onclick*='xoa'], a[href*='xoa']")
                                    onclick_text = delete_link.get_attribute('onclick')
                                    if onclick_text:
                                        match = re.search(r"xoa\(['\"]?(\d+)['\"]?\)", onclick_text)
                                        if match:
                                            product_id = int(match.group(1))
                            except:
                                pass
                    except:
                        pass
                    
                    # Method 2: Try to find the delete button which contains the product ID (fallback)
                    if product_id is None:
                        try:
                            delete_btn = row.find_element(By.CSS_SELECTOR, "a[onclick*='xoa']")
                            onclick_text = delete_btn.get_attribute('onclick')
                            if onclick_text:
                                # Extract ID from onclick: xoa('123') or xoa("123")
                                match = re.search(r"xoa\(['\"]?(\d+)['\"]?\)", onclick_text)
                                if match:
                                    product_id = int(match.group(1))
                        except:
                            pass
                    
                    # Method 3: Try to get ID from href attribute (if delete link has href)
                    if product_id is None:
                        try:
                            delete_link = row.find_element(By.CSS_SELECTOR, "a[href*='xoa']")
                            href = delete_link.get_attribute('href')
                            if href:
                                match = re.search(r'[?&]id=(\d+)', href)
                                if match:
                                    product_id = int(match.group(1))
                        except:
                            pass
                    
                    # Method 4: Try to get ID from data attribute
                    if product_id is None:
                        try:
                            data_id = row.get_attribute('data-id')
                            if data_id and data_id.isdigit():
                                product_id = int(data_id)
                        except:
                            pass
                    
                    # Method 5: Try to get ID from first cell (if ID is in first column)
                    if product_id is None:
                        try:
                            cells = row.find_elements(By.TAG_NAME, "td")
                            if cells:
                                first_cell_text = cells[0].text.strip()
                                if first_cell_text.isdigit():
                                    product_id = int(first_cell_text)
                        except:
                            pass
                    
                    if product_id is not None:
                        product_ids.append(product_id)
                        logger.debug(f"Tìm thấy sản phẩm ID: {product_id}")
                        
                except Exception as e:
                    logger.debug(f"Không thể lấy ID từ dòng: {str(e)}")
                    continue
            
            # Remove duplicates and sort IDs in ascending order (from smallest to largest)
            product_ids = list(set(product_ids))  # Remove duplicates
            product_ids.sort()  # Sort ascending (smallest to largest)
            
            logger.info(f"Đã lấy được {len(product_ids)} ID sản phẩm: {product_ids}")
            
            # Chỉ lấy sản phẩm có ID nhỏ nhất
            if product_ids:
                smallest_id = product_ids[0]  # ID nhỏ nhất
                product_ids = [smallest_id]
                logger.info(f"Đã chọn sản phẩm có ID nhỏ nhất để xóa: {smallest_id}")
            else:
                logger.warning("Không tìm thấy sản phẩm nào")
            
            return product_ids
            
        except Exception as e:
            error_msg = f"Lỗi khi lấy danh sách ID sản phẩm: {str(e)}"
            logger.error(error_msg, exc_info=True)
            self.take_screenshot("get_product_ids_error")
            return []
    
    def delete_product(self, product_id):
        """Delete a product by ID (products are deleted in order from smallest to largest ID)"""
        try:
            logger.info(f"Bắt đầu xóa sản phẩm ID: {product_id}")
            
            # Navigate to products page if not already there
            if "san-pham.php" not in self.driver.current_url:
                self.driver.get(self.test_data['products_url'])
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                time.sleep(1)  # Wait for table to fully load
            
            # Find the delete button (trash icon) in the "Hành Động" (Action) column for this product
            delete_btn = None
            
            # Method 1: Tìm icon thùng rác trong cột "Hành Động" của hàng chứa product_id
            # Tìm hàng chứa product_id, sau đó tìm icon thùng rác trong cột "Hành Động"
            try:
                # Tìm hàng chứa product_id
                product_row = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, f"//tr[.//text()[contains(., '{product_id}')] or .//td[contains(text(), '{product_id}')]]")
                    )
                )
                
                # Tìm icon thùng rác trong cột "Hành Động" của hàng đó
                # Các selector có thể cho icon thùng rác
                trash_selectors = [
                    (By.CSS_SELECTOR, "i.fa-trash, i.fa-trash-alt, i.fa-trash-o"),
                    (By.CSS_SELECTOR, "i[class*='trash']"),
                    (By.CSS_SELECTOR, "i[class*='delete']"),
                    (By.CSS_SELECTOR, "svg[class*='trash'], svg[class*='delete']"),
                    (By.XPATH, ".//i[contains(@class, 'trash')]"),
                    (By.XPATH, ".//i[contains(@class, 'delete')]"),
                    (By.XPATH, ".//*[contains(@class, 'trash')]"),
                    (By.XPATH, ".//*[contains(@class, 'delete')]"),
                ]
                
                for by, selector in trash_selectors:
                    try:
                        # Tìm trong hàng chứa product_id
                        delete_icon = product_row.find_element(by, selector)
                        # Tìm phần tử cha (link hoặc button) chứa icon
                        delete_btn = delete_icon.find_element(By.XPATH, "./ancestor::a | ./ancestor::button | ./parent::a | ./parent::button")
                        if delete_btn:
                            break
                    except:
                        continue
                
                # Nếu không tìm thấy qua icon, thử tìm link/button trong cột "Hành Động"
                if not delete_btn:
                    # Tìm cột "Hành Động" trong hàng
                    action_cell = product_row.find_element(By.XPATH, ".//td[contains(., 'Hành') or contains(., 'Động') or contains(., 'Action')] | .//td[last()]")
                    # Tìm link hoặc button trong cột đó
                    delete_btn = action_cell.find_element(By.XPATH, ".//a[contains(@onclick, 'xoa') or contains(@href, 'xoa')] | .//button[contains(@onclick, 'xoa')] | .//a[.//i[contains(@class, 'trash')]] | .//a[.//i[contains(@class, 'delete')]]")
                    
            except Exception as e:
                logger.debug(f"Method 1 failed: {str(e)}")
                pass
            
            # Method 2: Tìm trực tiếp bằng XPath với onclick chứa xoa('ID')
            if not delete_btn:
                try:
                    delete_btn = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, f"//tr[.//text()[contains(., '{product_id}')]]//a[contains(@onclick, \"xoa('{product_id}')\")] | //tr[.//td[contains(text(), '{product_id}')]]//a[contains(@onclick, 'xoa')]")
                        )
                    )
                except:
                    pass
            
            # Method 3: Tìm icon thùng rác trong hàng chứa product_id (cách tổng quát hơn)
            if not delete_btn:
                try:
                    # Tìm tất cả các hàng
                    rows = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                    for row in rows:
                        # Kiểm tra xem hàng có chứa product_id không
                        row_text = row.text
                        if str(product_id) in row_text:
                            # Tìm icon thùng rác trong hàng này
                            try:
                                trash_icon = row.find_element(By.CSS_SELECTOR, "i[class*='trash'], i[class*='delete'], *[class*='trash'], *[class*='delete']")
                                # Tìm phần tử cha có thể click được
                                delete_btn = trash_icon.find_element(By.XPATH, "./ancestor::a[1] | ./ancestor::button[1] | ./parent::a | ./parent::button")
                                if delete_btn:
                                    break
                            except:
                                # Nếu không tìm thấy icon, thử tìm link trong cột cuối (thường là cột Hành Động)
                                try:
                                    cells = row.find_elements(By.TAG_NAME, "td")
                                    if cells:
                                        last_cell = cells[-1]  # Cột cuối thường là cột Hành Động
                                        delete_btn = last_cell.find_element(By.TAG_NAME, "a")
                                        if delete_btn:
                                            break
                                except:
                                    continue
                except Exception as e:
                    logger.debug(f"Method 3 failed: {str(e)}")
                    pass
            
            if not delete_btn:
                raise NoSuchElementException(f"Không tìm thấy nút xóa (icon thùng rác) cho sản phẩm ID: {product_id} trong cột Hành Động")
            
            # Scroll to the element
            self.driver.execute_script("arguments[0].scrollIntoView(true);", delete_btn)
            time.sleep(1)  # Small delay for scrolling
            
            # Take screenshot before deletion
            before_delete_screenshot = self.take_screenshot(f"before_delete_{product_id}")
            
            # Click the delete button
            delete_btn.click()
            
            # Handle the confirmation alert
            try:
                WebDriverWait(self.driver, 5).until(EC.alert_is_present())
                alert = self.driver.switch_to.alert
                alert.accept()
                logger.info(f"Accepted delete confirmation for product {product_id}")
            except:
                logger.warning("No confirmation alert found")
            
            # Wait for deletion to complete
            time.sleep(2)  # Wait for the page to update
            
            # Take screenshot after deletion
            after_delete_screenshot = self.take_screenshot(f"after_delete_{product_id}")
            
            # Check if deletion was successful
            try:
                # Check for success message
                success_msg = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "alert-success"))
                )
                success_text = success_msg.text.strip()
                logger.info(f"Delete successful: {success_text}")
                
                self.add_test_step(
                    f"Xóa sản phẩm ID {product_id}",
                    "Xóa sản phẩm thành công, hiển thị thông báo xác nhận",
                    success_text,
                    "PASS",
                    screenshot_path=after_delete_screenshot
                )
                return True
                
            except:
                # If no success message, check if the product is still in the table
                try:
                    self.driver.find_element(By.XPATH, f"//a[contains(@onclick, \"xoa('{product_id}')\")]")
                    # If we get here, the product still exists
                    error_msg = f"Không tìm thấy thông báo xác nhận xóa sản phẩm {product_id}"
                except:
                    # Product was deleted successfully
                    self.add_test_step(
                        f"Xóa sản phẩm ID {product_id}",
                        "Xóa sản phẩm thành công",
                        "Sản phẩm đã được xóa khỏi danh sách",
                        "PASS",
                        screenshot_path=after_delete_screenshot
                    )
                    return True
                
                logger.error(error_msg)
                self.add_test_step(
                    f"Xóa sản phẩm ID {product_id}",
                    "Xóa sản phẩm thành công, hiển thị thông báo xác nhận",
                    error_msg,
                    "FAIL",
                    screenshot_path=after_delete_screenshot
                )
                return False
                
        except Exception as e:
            error_msg = f"Lỗi khi xóa sản phẩm {product_id}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            error_screenshot = self.take_screenshot(f"delete_error_{product_id}")
            self.add_test_step(
                f"Xóa sản phẩm ID {product_id}",
                "Xóa sản phẩm thành công",
                error_msg,
                "FAIL",
                screenshot_path=error_screenshot
            )
            return False
    
    def create_excel_report(self):
        """Create an Excel report with test results"""
        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Results"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Set column headers
            headers = ["STT", "Hành động", "Kỳ vọng", "Thực tế", "Trạng thái", "Thời gian"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add test steps
            for row_num, step in enumerate(self.test_steps, 2):
                ws.cell(row=row_num, column=1, value=row_num-1)  # STT
                ws.cell(row=row_num, column=2, value=step.get('action', ''))
                ws.cell(row=row_num, column=3, value=step.get('expected', ''))
                ws.cell(row=row_num, column=4, value=step.get('actual', ''))
                
                # Status with color coding
                status_cell = ws.cell(row=row_num, column=5, value=step.get('status', ''))
                if step.get('status') == 'PASS':
                    status_cell.fill = pass_fill
                else:
                    status_cell.fill = fail_fill
                
                ws.cell(row=row_num, column=6, value=step.get('timestamp', ''))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                
                # Find the maximum length of content in the column
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width with some padding
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 50)  # Max width 50
            
            # Save the workbook
            wb.save(self.excel_path)
            logger.info(f"Báo cáo đã được lưu tại: {os.path.abspath(self.excel_path)}")
            return True
            
        except Exception as e:
            logger.error(f"Lỗi khi tạo báo cáo Excel: {str(e)}", exc_info=True)
            return False
    
    def run_test(self):
        """Run the test"""
        try:
            # Initialize WebDriver
            if not self.setup_driver():
                logger.error("Không thể khởi tạo WebDriver")
                return False
            
            try:
                # Step 1: Login to admin
                if not self.login_to_admin():
                    logger.error("Đăng nhập thất bại, dừng kiểm thử")
                    return False
                
                # Step 2: Truy cập trang quản lý sản phẩm
                self.add_test_step(
                    "Truy cập trang quản lý sản phẩm",
                    "Truy cập thành công vào http://localhost/webbansach/admin/san-pham.php",
                    "Đang truy cập...",
                    "PENDING"
                )
                
                logger.info("Truy cập trang quản lý sản phẩm...")
                self.driver.get(self.test_data['products_url'])
                
                # Wait for products table to load
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "table"))
                )
                
                screenshot_path = self.take_screenshot("products_page")
                
                self.test_steps[-1].update({
                    'status': 'PASS',
                    'actual': f'Đã truy cập thành công vào trang quản lý sản phẩm: {self.driver.current_url}',
                    'screenshot': screenshot_path
                })
                logger.info(f"✅ Đã truy cập thành công vào trang quản lý sản phẩm: {self.driver.current_url}")
                
                # Step 3: Tìm sản phẩm có ID nhỏ nhất
                self.add_test_step(
                    "Tìm sản phẩm có ID nhỏ nhất",
                    "Tìm thấy sản phẩm có ID nhỏ nhất",
                    "Đang tìm kiếm...",
                    "PENDING"
                )
                
                product_ids = self.get_product_ids()
                
                if not product_ids:
                    self.test_steps[-1].update({
                        'status': 'FAIL',
                        'actual': 'Không tìm thấy sản phẩm nào để xóa'
                    })
                    logger.error("Không tìm thấy sản phẩm nào để xóa")
                    return False
                
                smallest_id = product_ids[0]
                self.test_steps[-1].update({
                    'status': 'PASS',
                    'actual': f'Đã tìm thấy sản phẩm có ID nhỏ nhất: {smallest_id}'
                })
                logger.info(f"✅ Đã tìm thấy sản phẩm có ID nhỏ nhất: {smallest_id}")
                
                # Step 4: Xóa sản phẩm có ID nhỏ nhất
                logger.info("="*60)
                logger.info(f"Bắt đầu xóa sản phẩm có ID nhỏ nhất: {smallest_id}")
                logger.info("="*60)
                
                if self.delete_product(smallest_id):
                    deleted_count = 1
                    logger.info(f"✅ Đã xóa thành công sản phẩm có ID nhỏ nhất: {smallest_id}")
                else:
                    deleted_count = 0
                    logger.error(f"❌ Xóa thất bại sản phẩm có ID: {smallest_id}")
                
                logger.info("="*60)
                logger.info(f"Kết quả: {'✅ THÀNH CÔNG' if deleted_count > 0 else '❌ THẤT BẠI'}")
                logger.info("="*60)
                
                # Step 5: Generate report
                self.add_test_step(
                    "Tạo báo cáo kết quả",
                    "Tạo báo cáo Excel thành công",
                    f"Đã xóa sản phẩm có ID nhỏ nhất: {smallest_id}" if deleted_count > 0 else f"Xóa thất bại sản phẩm có ID: {smallest_id}",
                    "PASS" if deleted_count > 0 else "FAIL"
                )
                
                # Create Excel report
                if self.create_excel_report():
                    logger.info(f"Đã tạo báo cáo tại: {os.path.abspath(self.excel_path)}")
                else:
                    logger.error("Không thể tạo báo cáo")
                
                return deleted_count > 0
                
            except Exception as e:
                error_msg = f"Lỗi không mong muốn: {str(e)}"
                logger.error(error_msg, exc_info=True)
                self.take_screenshot("unexpected_error")
                return False
                
            finally:
                # Close the browser
                try:
                    self.driver.quit()
                    logger.info("Đã đóng trình duyệt")
                except:
                    pass
                
        except Exception as e:
            logger.error(f"Lỗi nghiêm trọng: {str(e)}", exc_info=True)
            return False

if __name__ == "__main__":
    test = TestDeleteProducts()
    result = test.run_test()
    
    # Show final result
    if result:
        print("\n" + "="*60)
        print("Kết quả: ✅ THÀNH CÔNG")
        print("Đã xóa thành công sản phẩm có ID nhỏ nhất trong trang quản lý sản phẩm")
        print("="*60)
    else:
        print("\n" + "="*60)
        print("Kết quả: ❌ THẤT BẠI")
        print("Xóa sản phẩm có ID nhỏ nhất không thành công")
        print("="*60)
    
    sys.exit(0 if result else 1)
