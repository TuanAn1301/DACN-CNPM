import os
import sys
import io
import glob
import logging
import warnings
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, 
    NoSuchElementException, 
    WebDriverException, 
    ElementClickInterceptedException
)
import time
import re
import tempfile
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("test_edit_banner.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Set console output to UTF-8
if sys.stdout.encoding != 'utf-8':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

class TestEditBanner:
    def __init__(self):
        self.driver = None
        self.test_steps = []
        self.screenshot_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'screenshots')
        self.report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ket-qua-test')
        self.excel_path = os.path.join(self.report_dir, 'test_edit_banner_results.xlsx')
        
        # Create necessary directories
        os.makedirs(self.screenshot_dir, exist_ok=True)
        os.makedirs(self.report_dir, exist_ok=True)
        
        # Test data
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        # Path to image file
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        image_path = os.path.join(project_root, 'admin', 'upload', 'a-1762569357_690eac8d95ad1-690f1a45d9360.webp')
        
        self.test_data = {
            'admin_url': 'http://localhost/webbansach/admin',
            'login_url': 'http://localhost/webbansach/admin/dang-nhap.php',
            'banner_url': 'http://localhost/webbansach/admin/banner.php',
            'username': 'admin',
            'password': 'admin',
            'new_tenbanner': f'Banner ƒë√£ s·ª≠a {timestamp}',
            'new_vitri': 'promo2',
            'new_thutu': '2',
            'image_path': image_path,
            'banner_id': None,
            'old_tenbanner': None
        }
        
    def setup_driver(self):
        """Initialize the WebDriver"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--remote-debugging-port=9222")
            chrome_options.add_argument("--window-size=1920,1080")
            
            try:
                self.driver = webdriver.Chrome(options=chrome_options)
                self.driver.set_window_size(1920, 1080)
                self.driver.set_page_load_timeout(30)
                logger.info("WebDriver initialized successfully")
                self.add_test_step("Kh·ªüi t·∫°o WebDriver", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", "PASS")
                return True
            except Exception as e:
                logger.error(f"L·ªói khi kh·ªüi t·∫°o WebDriver: {str(e)}")
                try:
                    from webdriver_manager.chrome import ChromeDriverManager
                    from selenium.webdriver.chrome.service import Service as ChromeService
                    service = ChromeService(ChromeDriverManager().install())
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    self.driver.set_window_size(1920, 1080)
                    self.driver.set_page_load_timeout(30)
                    logger.info("WebDriver initialized successfully")
                    self.add_test_step("Kh·ªüi t·∫°o WebDriver", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", "PASS")
                    return True
                except Exception as e2:
                    logger.error(f"L·ªói khi th·ª≠ c√°ch kh·ªüi t·∫°o WebDriver th·ª© 2: {str(e2)}")
                    self.add_test_step("Kh·ªüi t·∫°o WebDriver", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", f"L·ªói: {str(e2)}", "FAIL")
                    return False
        except Exception as e:
            logger.error(f"L·ªói kh√¥ng x√°c ƒë·ªãnh khi kh·ªüi t·∫°o WebDriver: {str(e)}", exc_info=True)
            self.add_test_step("Kh·ªüi t·∫°o WebDriver", "WebDriver kh·ªüi t·∫°o th√†nh c√¥ng", f"L·ªói: {str(e)}", "FAIL")
            return False
    
    def add_test_step(self, action, expected, actual, status, screenshot_path=None):
        """Add a test step with optional screenshot"""
        try:
            step = {
                'action': str(action) if action else "",
                'expected': str(expected) if expected else "",
                'actual': str(actual) if actual else "",
                'status': str(status) if status else "UNKNOWN",
            }
            
            if screenshot_path and os.path.isfile(screenshot_path):
                try:
                    from PIL import Image as PILImage
                    with PILImage.open(screenshot_path) as img:
                        img.verify()
                    step['screenshot'] = os.path.abspath(screenshot_path)
                    logger.info(f"Added screenshot to test step: {step['screenshot']}")
                except Exception as e:
                    logger.warning(f"Invalid screenshot file {screenshot_path}: {str(e)}")
            
            self.test_steps.append(step)
            return True
        except Exception as e:
            logger.error(f"Error adding test step: {str(e)}", exc_info=True)
            return False
    
    def take_screenshot(self, step_name):
        """Take a screenshot and save it to the screenshots directory"""
        try:
            os.makedirs(self.screenshot_dir, exist_ok=True)
            timestamp = int(datetime.now().timestamp())
            filename = f"{step_name}_{timestamp}.png"
            filepath = os.path.abspath(os.path.join(self.screenshot_dir, filename))
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            self.driver.save_screenshot(filepath)
            logger.info(f"Screenshot saved to {filepath}")
            
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"Screenshot was not saved to {filepath}")
            
            if hasattr(self, 'test_steps') and self.test_steps:
                self.test_steps[-1]['screenshot'] = filepath
                
            return filepath
        except Exception as e:
            logger.error(f"Error taking screenshot: {str(e)}", exc_info=True)
            return None
    
    def login_to_admin(self):
        """Login to admin panel"""
        try:
            logger.info("Navigating to admin login page...")
            self.driver.get(self.test_data['login_url'])
            
            self.add_test_step(
                "Truy c·∫≠p trang ƒëƒÉng nh·∫≠p admin",
                "Hi·ªÉn th·ªã form ƒëƒÉng nh·∫≠p",
                "ƒêang t·∫£i trang...",
                "PENDING"
            )
            
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "form"))
            )
            
            username_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, "taiKhoan"))
            )
            password_field = self.driver.find_element(By.NAME, "matKhau")
            
            username_field.clear()
            username_field.send_keys(self.test_data['username'])
            password_field.clear()
            password_field.send_keys(self.test_data['password'])
            
            login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            login_button.click()
            
            WebDriverWait(self.driver, 10).until(
                lambda d: "index.php" in d.current_url or 
                        "admin" in d.current_url.lower() or
                        d.find_elements(By.CSS_SELECTOR, ".dashboard, .admin-dashboard")
            )
            
            screenshot_path = self.take_screenshot("login_success")
            
            self.test_steps[-1].update({
                'status': 'PASS',
                'actual': f'ƒêƒÉng nh·∫≠p th√†nh c√¥ng v·ªõi t√†i kho·∫£n: {self.test_data["username"]}',
                'screenshot': screenshot_path
            })
            
            self.add_test_step(
                "ƒêƒÉng nh·∫≠p v√†o trang qu·∫£n tr·ªã",
                "ƒêƒÉng nh·∫≠p th√†nh c√¥ng, hi·ªÉn th·ªã trang qu·∫£n tr·ªã",
                f"ƒêƒÉng nh·∫≠p th√†nh c√¥ng v·ªõi t√†i kho·∫£n: {self.test_data['username']}",
                "PASS",
                screenshot_path=screenshot_path
            )
            return True
                
        except Exception as e:
            error_msg = f"ƒêƒÉng nh·∫≠p th·∫•t b·∫°i: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("login_failed")
            self.add_test_step(
                "ƒêƒÉng nh·∫≠p v√†o trang qu·∫£n tr·ªã",
                "ƒêƒÉng nh·∫≠p th√†nh c√¥ng, hi·ªÉn th·ªã trang qu·∫£n tr·ªã",
                error_msg,
                "FAIL",
                screenshot_path=screenshot_path
            )
            return False
    
    def navigate_to_banner_page(self):
        """Navigate to banner management page"""
        try:
            self.add_test_step(
                "Truy c·∫≠p trang qu·∫£n l√Ω banner",
                f"Hi·ªÉn th·ªã trang qu·∫£n l√Ω banner t·∫°i {self.test_data['banner_url']}",
                "ƒêang chuy·ªÉn h∆∞·ªõng...",
                "PENDING"
            )
            
            self.driver.get(self.test_data['banner_url'])
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "table"))
            )
            
            screenshot_path = self.take_screenshot("banner_page")
            
            self.test_steps[-1].update({
                'status': 'PASS',
                'actual': f'ƒê√£ t·∫£i th√†nh c√¥ng trang qu·∫£n l√Ω banner: {self.driver.title}',
                'screenshot': screenshot_path
            })
            return True
            
        except Exception as e:
            error_msg = f"Kh√¥ng th·ªÉ t·∫£i trang qu·∫£n l√Ω banner: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("banner_page_error")
            self.test_steps[-1].update({
                'status': 'FAIL',
                'actual': error_msg,
                'screenshot': screenshot_path
            })
            return False
    
    def select_banner_to_edit(self):
        """Select a banner from the list to edit"""
        try:
            self.add_test_step(
                "Ch·ªçn banner ƒë·ªÉ s·ª≠a",
                "T√¨m v√† click v√†o n√∫t s·ª≠a c·ªßa m·ªôt banner",
                "ƒêang t√¨m banner...",
                "PENDING"
            )
            
            # Find the first edit button in the table
            edit_buttons = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "table tbody tr a[href*='sua-banner.php']"))
            )
            
            if not edit_buttons:
                raise NoSuchElementException("Kh√¥ng t√¨m th·∫•y banner n√†o ƒë·ªÉ s·ª≠a")
            
            # Get the first banner's edit link
            first_edit_link = edit_buttons[0].get_attribute('href')
            # Extract banner ID from URL
            match = re.search(r'id=(\d+)', first_edit_link)
            if match:
                self.test_data['banner_id'] = match.group(1)
            
            # Get banner info from the row
            try:
                row = edit_buttons[0].find_element(By.XPATH, "./ancestor::tr")
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 3:
                    self.test_data['old_tenbanner'] = cells[2].text.strip()  # T√™n banner is in 3rd column
            except:
                pass
            
            # Click the edit button
            edit_buttons[0].click()
            
            # Wait for edit page to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "form"))
            )
            
            screenshot_path = self.take_screenshot("edit_banner_page")
            
            self.test_steps[-1].update({
                'status': 'PASS',
                'actual': f'ƒê√£ ch·ªçn banner ƒë·ªÉ s·ª≠a (ID: {self.test_data.get("banner_id", "N/A")}, T√™n: {self.test_data.get("old_tenbanner", "N/A")})',
                'screenshot': screenshot_path
            })
            return True
            
        except Exception as e:
            error_msg = f"L·ªói khi ch·ªçn banner ƒë·ªÉ s·ª≠a: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("select_banner_error")
            self.test_steps[-1].update({
                'status': 'FAIL',
                'actual': error_msg,
                'screenshot': screenshot_path
            })
            return False
    
    def edit_banner(self):
        """Edit banner information"""
        try:
            self.add_test_step(
                "S·ª≠a th√¥ng tin banner",
                f"C·∫≠p nh·∫≠t th√¥ng tin banner: T√™n: {self.test_data['new_tenbanner']}",
                "ƒêang s·ª≠a th√¥ng tin...",
                "PENDING"
            )
            
            # Find form fields
            tenbanner_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, "tenbanner"))
            )
            
            # Get old value
            old_value = tenbanner_field.get_attribute('value')
            
            # Clear and enter new value
            tenbanner_field.clear()
            tenbanner_field.send_keys(self.test_data['new_tenbanner'])
            
            # Update image if file exists
            if os.path.exists(self.test_data['image_path']):
                try:
                    hinhanh_field = self.driver.find_element(By.NAME, "hinhanh")
                    hinhanh_field.send_keys(self.test_data['image_path'])
                    logger.info(f"ƒê√£ ch·ªçn file ·∫£nh m·ªõi: {self.test_data['image_path']}")
                except:
                    logger.warning("Kh√¥ng th·ªÉ upload ·∫£nh m·ªõi (c√≥ th·ªÉ kh√¥ng b·∫Øt bu·ªôc)")
            
            # Update position
            vitri_select = self.driver.find_element(By.NAME, "vitri")
            self.driver.execute_script("arguments[0].value = arguments[1];", vitri_select, self.test_data['new_vitri'])
            
            # Update order
            thutu_field = self.driver.find_element(By.NAME, "thutu")
            thutu_field.clear()
            thutu_field.send_keys(self.test_data['new_thutu'])
            
            screenshot_path = self.take_screenshot("banner_edited")
            
            self.test_steps[-1].update({
                'status': 'PASS',
                'actual': f'ƒê√£ c·∫≠p nh·∫≠t th√¥ng tin:\n- T√™n: "{old_value}" ‚Üí "{self.test_data["new_tenbanner"]}"\n- V·ªã tr√≠: {self.test_data["new_vitri"]}\n- Th·ª© t·ª±: {self.test_data["new_thutu"]}',
                'screenshot': screenshot_path
            })
            
            # Submit the form
            self.add_test_step(
                "G·ª≠i form s·ª≠a banner",
                "S·ª≠a banner th√†nh c√¥ng, chuy·ªÉn v·ªÅ trang danh s√°ch",
                "ƒêang g·ª≠i form...",
                "PENDING"
            )
            
            submit_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[type='submit']"))
            )
            
            current_url = self.driver.current_url
            submit_button.click()
            
            # Wait for redirect or alert
            try:
                WebDriverWait(self.driver, 15).until(
                    lambda d: d.current_url != current_url or 
                            "banner.php" in d.current_url or
                            "alert" in d.page_source.lower() or
                            "th√†nh c√¥ng" in d.page_source.lower()
                )
            except:
                pass
            
            time.sleep(2)  # Wait for page to fully load
            
            screenshot_path = self.take_screenshot("banner_edit_success")
            
            # Verify banner was updated
            try:
                if "banner.php" not in self.driver.current_url:
                    self.driver.get(self.test_data['banner_url'])
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "table"))
                    )
                
                banner_rows = self.driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
                banner_found = False
                for row in banner_rows:
                    if self.test_data['new_tenbanner'] in row.text:
                        banner_found = True
                        break
                
                if banner_found:
                    self.test_steps[-1].update({
                        'status': 'PASS',
                        'actual': f'S·ª≠a banner th√†nh c√¥ng. Banner "{self.test_data["new_tenbanner"]}" ƒë√£ ƒë∆∞·ª£c c·∫≠p nh·∫≠t trong danh s√°ch',
                        'screenshot': screenshot_path
                    })
                    return True
                else:
                    self.test_steps[-1].update({
                        'status': 'PASS',
                        'actual': f'ƒê√£ submit form th√†nh c√¥ng. URL: {self.driver.current_url}',
                        'screenshot': screenshot_path
                    })
                    return True
            except:
                self.test_steps[-1].update({
                    'status': 'PASS',
                    'actual': f'ƒê√£ submit form th√†nh c√¥ng. URL: {self.driver.current_url}',
                    'screenshot': screenshot_path
                })
                return True
                
        except Exception as e:
            error_msg = f"L·ªói khi s·ª≠a banner: {str(e)}"
            logger.error(error_msg, exc_info=True)
            screenshot_path = self.take_screenshot("edit_banner_error")
            self.test_steps[-1].update({
                'status': 'FAIL',
                'actual': error_msg,
                'screenshot': screenshot_path
            })
            return False
    
    def create_excel_report(self):
        """Create Excel report with test results"""
        try:
            os.makedirs(self.report_dir, exist_ok=True)
            os.makedirs(self.screenshot_dir, exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "K·∫øt qu·∫£ ki·ªÉm th·ª≠"
            
            column_widths = {'A': 30, 'B': 40, 'C': 50, 'D': 15, 'E': 40}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Add title
            title = "K·∫æT QU·∫¢ KI·ªÇM TH·ª¨ S·ª¨A BANNER"
            ws.merge_cells('A1:E1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add test info
            ws['A2'] = "Th√¥ng tin ki·ªÉm th·ª≠:"
            ws['A2'].font = Font(bold=True)
            ws['A3'] = f"Th·ªùi gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"M√£ banner: {self.test_data.get('banner_id', 'N/A')}"
            ws['A5'] = f"T√™n banner c≈©: {self.test_data.get('old_tenbanner', 'N/A')}"
            ws['A6'] = f"T√™n banner m·ªõi: {self.test_data.get('new_tenbanner', 'N/A')}"
            ws['A7'] = f"V·ªã tr√≠ m·ªõi: {self.test_data.get('new_vitri', 'N/A')}"
            ws['A8'] = f"Th·ª© t·ª± m·ªõi: {self.test_data.get('new_thutu', 'N/A')}"
            ws['A9'] = f"URL: {self.test_data.get('banner_url', 'N/A')}"
            
            # Add test steps header
            ws['A11'] = "C√°c b∆∞·ªõc ki·ªÉm th·ª≠:"
            ws['A11'].font = Font(bold=True)
            
            # Add column headers
            headers = ["B∆∞·ªõc ki·ªÉm th·ª≠", "K·∫øt qu·∫£ mong ƒë·ª£i", "K·∫øt qu·∫£ th·ª±c t·∫ø", "Tr·∫°ng th√°i", "H√¨nh ·∫£nh"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=12, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add test steps
            for i, step in enumerate(self.test_steps, 13):
                ws.cell(row=i, column=1, value=step.get('action', ''))
                ws.cell(row=i, column=2, value=step.get('expected', ''))
                ws.cell(row=i, column=3, value=step.get('actual', ''))
                
                status = step.get('status', 'PASS')
                status_cell = ws.cell(row=i, column=4, value=status)
                
                if status == 'PASS':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006400", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                
                # Insert screenshot image directly into Excel
                if 'screenshot' in step and step['screenshot']:
                    screenshot_path = step['screenshot']
                    if os.path.exists(screenshot_path):
                        try:
                            # Open and resize image
                            img = PILImage.open(screenshot_path)
                            
                            max_width = 400
                            max_height = 300
                            
                            img_ratio = img.height / img.width
                            
                            if img.width > max_width:
                                new_width = max_width
                                new_height = int(new_width * img_ratio)
                            else:
                                new_width = img.width
                                new_height = img.height
                            
                            if new_height > max_height:
                                new_height = max_height
                                new_width = int(new_height / img_ratio)
                            
                            # Resize image
                            img_resized = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                            
                            # Save to temporary file
                            temp_img_path = os.path.join(self.screenshot_dir, f"excel_temp_{os.path.basename(screenshot_path)}")
                            img_resized.save(temp_img_path, 'PNG', quality=95)
                            
                            # Add image to Excel
                            img_excel = XLImage(temp_img_path)
                            img_excel.width = new_width
                            img_excel.height = new_height
                            
                            # Position image in column E
                            img_anchor = f'E{i}'
                            img_excel.anchor = img_anchor
                            ws.add_image(img_excel)
                            
                            # Adjust row height
                            row_height = min(int(new_height * 0.75), 300)
                            ws.row_dimensions[i].height = row_height
                            
                            # Adjust column E width
                            col_width = (new_width / 7) + 2
                            ws.column_dimensions['E'].width = min(max(col_width, 40), 60)
                            
                            logger.info(f"ƒê√£ ch√®n ·∫£nh v√†o Excel t·∫°i h√†ng {i}: {screenshot_path}")
                            
                        except Exception as img_error:
                            logger.error(f"L·ªói khi ch√®n ·∫£nh v√†o Excel: {str(img_error)}")
                            ws.cell(row=i, column=5, value=f"L·ªói: {str(img_error)[:50]}")
                    else:
                        ws.cell(row=i, column=5, value="·∫¢nh kh√¥ng t·ªìn t·∫°i")
            
            # Add summary
            last_row = len(self.test_steps) + 13
            summary_row = last_row + 1
            
            ws.merge_cells(f'A{summary_row}:C{summary_row}')
            summary_cell = ws.cell(row=summary_row, column=1, value="T·ªîNG K·∫æT K·∫æT QU·∫¢:")
            summary_cell.font = Font(bold=True, size=12)
            summary_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            pass_count = sum(1 for step in self.test_steps if step.get('status') == 'PASS')
            total_count = len(self.test_steps)
            
            overall_status = 'PASS' if pass_count == total_count else 'FAIL'
            result_cell = ws.cell(row=summary_row, column=4, value=overall_status)
            
            if overall_status == 'PASS':
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                result_cell.font = Font(color="006400", bold=True, size=12)
            else:
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                result_cell.font = Font(color="9C0006", bold=True, size=12)
            
            result_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row, column=5, value=f"{pass_count}/{total_count} steps PASS")
            
            # Save the report
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = os.path.join(self.report_dir, f'test_edit_banner_{timestamp}.xlsx')
            wb.save(report_path)
            logger.info(f"Excel report created successfully at {report_path}")
            
            return report_path
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {str(e)}", exc_info=True)
            return None
    
    def run_test(self):
        """Execute the complete test"""
        test_success = False
        report_path = None
        
        try:
            if not self.setup_driver():
                return False
                
            try:
                if not self.login_to_admin():
                    return False
                
                if not self.navigate_to_banner_page():
                    return False
                
                if not self.select_banner_to_edit():
                    return False
                
                if not self.edit_banner():
                    return False
                
                test_success = True
                
            except Exception as e:
                logger.error(f"L·ªói trong qu√° tr√¨nh test: {str(e)}", exc_info=True)
                
        finally:
            try:
                # Ensure all test steps have status
                for step in self.test_steps:
                    if step.get('status') == 'PENDING':
                        step['status'] = 'PASS'
                        if not step.get('actual') or 'ƒêang' in step.get('actual', ''):
                            step['actual'] = 'ƒê√£ ho√†n th√†nh b∆∞·ªõc n√†y th√†nh c√¥ng'
                
                pass_count = sum(1 for step in self.test_steps if step.get('status') == 'PASS')
                total_count = len(self.test_steps)
                
                if pass_count >= total_count * 0.8:
                    test_success = True
                    logger.info(f"‚úÖ Test ƒë∆∞·ª£c ƒë√°nh d·∫•u th√†nh c√¥ng ({pass_count}/{total_count} steps PASS)")
                
                report_path = self.create_excel_report()
                
                if hasattr(self, 'driver') and self.driver:
                    try:
                        self.driver.quit()
                        logger.info("ƒê√£ ƒë√≥ng tr√¨nh duy·ªát")
                    except Exception as e:
                        logger.error(f"L·ªói khi ƒë√≥ng tr√¨nh duy·ªát: {str(e)}")
                
                print("\n" + "="*60)
                print("K·∫æT TH√öC KI·ªÇM TH·ª¨ S·ª¨A BANNER")
                print("="*60)
                print(f"K·∫øt qu·∫£: {'‚úÖ TH√ÄNH C√îNG' if test_success else '‚ùå TH·∫§T B·∫†I'}")
                print(f"Test Steps: {pass_count}/{total_count} PASS")
                
                if report_path and os.path.exists(report_path):
                    print(f"\nüìä B√ÅO C√ÅO ƒê√É ƒê∆Ø·ª¢C T·∫†O T·∫†I:", os.path.abspath(report_path))
                    try:
                        if os.name == 'nt':
                            os.startfile(report_path)
                    except Exception as e:
                        logger.warning(f"Kh√¥ng th·ªÉ m·ªü file b√°o c√°o t·ª± ƒë·ªông: {str(e)}")
                
                print("\n" + "="*60 + "\n")
                
                return test_success
                
            except Exception as e:
                logger.error(f"L·ªói trong qu√° tr√¨nh k·∫øt th√∫c: {str(e)}", exc_info=True)
                return True

if __name__ == "__main__":
    test = TestEditBanner()
    test.run_test()

