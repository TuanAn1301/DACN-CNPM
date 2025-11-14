import time
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from PIL import Image
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

class TestReporter:
    def __init__(self, test_name):
        self.test_name = test_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Kết quả đăng nhập"
        self.current_row = 1
        self.screenshot_count = 0
        self.test_steps = []
        self.setup_worksheet()
        
    def setup_worksheet(self):
        # Định dạng tiêu đề
        headers = ["Bước thực hiện", "Mô tả", "Dữ liệu đầu vào", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái"]
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Đặt chiều rộng cột
            if col_num == 1:  # Cột Bước thực hiện
                self.ws.column_dimensions['A'].width = 15
            elif col_num == 2:  # Cột Mô tả
                self.ws.column_dimensions['B'].width = 40
            elif col_num == 3:  # Cột Dữ liệu đầu vào
                self.ws.column_dimensions['C'].width = 30
            elif col_num == 4:  # Cột Kết quả mong đợi
                self.ws.column_dimensions['D'].width = 30
            elif col_num == 5:  # Cột Kết quả thực tế
                self.ws.column_dimensions['E'].width = 40
            else:  # Cột Trạng thái
                self.ws.column_dimensions['F'].width = 15
                
        self.current_row += 1
        
    def add_step(self, description, status='PASS', input_data='', expected='', notes=''):
        """Thêm một bước thực hiện vào báo cáo"""
        step_num = len(self.test_steps) + 1
        self.test_steps.append({
            'step': step_num,
            'description': description,
            'status': status.upper(),
            'input': input_data,
            'expected': expected,
            'notes': notes,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
        # Thêm vào worksheet
        self.ws.append([
            step_num,
            description,
            str(input_data)[:300],  # Giới hạn độ dài để tránh lỗi
            str(expected)[:300],
            str(notes)[:300],
            status.upper()
        ])
        
        # Định dạng dòng vừa thêm
        for col in range(1, 7):
            cell = self.ws.cell(row=self.current_row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Tô màu ô trạng thái
            if col == 6:  # Cột Trạng thái
                if status.upper() == 'PASS':
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Xanh lá nhạt
                elif status.upper() == 'FAIL':
                    cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Đỏ nhạt
                elif status.upper() == 'ERROR':
                    cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")  # Cam nhạt
                else:  # INFO, WARNING, v.v.
                    cell.fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")  # Tím nhạt
        
        self.current_row += 1
        
    def add_screenshot(self, driver, description=""):
        """Chụp màn hình và thêm vào báo cáo"""
        try:
            # Tạo thư mục screenshots nếu chưa tồn tại
            if not os.path.exists('screenshots'):
                os.makedirs('screenshots')
                
            # Tạo tên file ảnh
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"screenshots/screenshot_{self.screenshot_count:02d}_{timestamp}.png"
            
            # Chụp màn hình và lưu file
            driver.save_screenshot(filename)
            
            # Thêm ảnh vào báo cáo
            img = XLImage(filename)
            # Thu nhỏ ảnh nếu cần
            img.width = 600
            img.height = 400
            
            # Thêm ảnh vào ô bên dưới bảng
            img_cell = f'G{self.current_row}'
            self.ws.add_image(img, img_cell)
            
            # Điều chỉnh chiều cao hàng để hiển thị ảnh
            self.ws.row_dimensions[self.current_row].height = 300
            
            # Thêm mô tả ảnh
            self.ws[f'G{self.current_row-1}'] = f"Hình ảnh: {description}"
            
            self.screenshot_count += 1
            return filename
            
        except Exception as e:
            print(f"Lỗi khi thêm ảnh vào báo cáo: {str(e)}")
            return None
    
    def save_report(self, folder_path=None):
        """Lưu báo cáo ra file Excel"""
        try:
            # Nếu không chỉ định folder_path, tạo folder "kết quả test" trong thư mục chứa file này
            if folder_path is None:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                folder_path = os.path.join(current_dir, "kết quả test")
            
            # Tạo thư mục nếu chưa tồn tại
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                
            # Tạo tên file báo cáo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(folder_path, f"bao_cao_dang_nhap_{timestamp}.xlsx")
            
            # Lưu file
            self.wb.save(filename)
            print(f"\nĐã lưu báo cáo tại: {os.path.abspath(filename)}")
            return filename
            
        except Exception as e:
            print(f"Lỗi khi lưu báo cáo: {str(e)}")
            return None

def in_thong_bao(buoc, trang_thai=""):
    """Hiển thị thông báo từng bước thực hiện"""
    if trang_thai.upper() == "OK":
        print(f"[✓] {buoc}")
    elif trang_thai.upper() == "LOI":
        print(f"[✗] {buoc}")
    else:
        print(f"[ ] {buoc}")
    sys.stdout.flush()  # Đảm bảo in ra ngay lập tức

def kiem_tra_phan_tu(driver, loai, gia_tri, timeout=10):
    """Kiểm tra sự tồn tại của phần tử"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((loai, gia_tri))
        )
        return True
    except TimeoutException:
        return False

def kiem_tra_dang_nhap_thanh_cong(driver):
    """Kiểm tra xem đăng nhập có thành công không"""
    try:
        # Kiểm tra xem có thông báo lỗi không
        thong_bao_loi = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-danger, .error-message")
        if thong_bao_loi:
            return False, f"Lỗi: {thong_bao_loi[0].text}"
            
        # Kiểm tra URL hiện tại
        if "index.php" in driver.current_url:
            return True, "Đăng nhập thành công!"
            
        # Kiểm tra xem có thông báo thành công không
        thong_bao_thanh_cong = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-success")
        if thong_bao_thanh_cong and "thành công" in thong_bao_thanh_cong[0].text.lower():
            return True, "Đăng nhập thành công!"
            
        return False, "Không xác định được trạng thái đăng nhập"
        
    except Exception as e:
        return False, f"Lỗi khi kiểm tra đăng nhập: {str(e)}"

def main():
    # Thông tin đăng nhập
    ten_dang_nhap = 'quan'
    mat_khau = '1'
    url_dang_nhap = 'http://localhost/webbansach/dang-nhap.php'
    
    # Khởi tạo báo cáo
    reporter = TestReporter("Kiểm thử đăng nhập thành công")
    reporter.add_step("Khởi tạo", "INFO", "", "Khởi tạo môi trường kiểm thử")
    reporter.add_step("Thông tin đăng nhập", "INFO", 
                     f"Tên đăng nhập: {ten_dang_nhap}\nMật khẩu: {'*' * len(mat_khau)}", 
                     f"Truy cập {url_dang_nhap}")
    
    # Tùy chọn cho Chrome
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')  # Mở cửa sổ lớn nhất
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # Tắt thông báo console
    
    print("\n=== BẮT ĐẦU TỰ ĐỘNG ĐĂNG NHẬP ===\n")
    print(f"Tài khoản: {ten_dang_nhap}")
    print(f"Mật khẩu: {'*' * len(mat_khau)}\n")

    driver = None
    try:
        # Khởi tạo Chrome WebDriver
        in_thong_bao("Đang khởi tạo trình duyệt...")
        reporter.add_step("Khởi tạo trình duyệt", "INFO", "Chrome WebDriver", "Trình duyệt khởi động thành công")
        
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            in_thong_bao("Đã khởi tạo trình duyệt thành công", "OK")
            reporter.add_step("Cài đặt WebDriver", "PASS", "ChromeDriverManager", "Cài đặt ChromeDriver thành công")
        except Exception as e:
            error_msg = f"Lỗi khi khởi tạo trình duyệt: {str(e)}"
            in_thong_bao(error_msg, "LOI")
            reporter.add_step("Cài đặt WebDriver", "ERROR", "ChromeDriverManager", error_msg)
            
            # Thử cách khác nếu cách trên thất bại
            try:
                in_thong_bao("Thử khởi tạo trình duyệt với Chrome mặc định...")
                driver = webdriver.Chrome(options=options)
                in_thong_bao("Đã khởi tạo trình duyệt thành công (dùng Chrome mặc định)", "OK")
                reporter.add_step("Khởi tạo trình duyệt", "PASS", "Chrome mặc định", "Khởi tạo trình duyệt thành công")
            except Exception as e2:
                error_msg = f"Không thể khởi tạo trình duyệt: {str(e2)}"
                in_thong_bao(error_msg, "LOI")
                reporter.add_step("Khởi tạo trình duyệt", "ERROR", "Chrome mặc định", error_msg)
                reporter.save_report()
                return

        # Mở trang đăng nhập
        in_thong_bao(f"Đang mở trang đăng nhập: {url_dang_nhap}")
        try:
            driver.get(url_dang_nhap)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            in_thong_bao("Đã tải trang đăng nhập", "OK")
            reporter.add_step("Truy cập trang đăng nhập", "PASS", url_dang_nhap, "Tải trang đăng nhập thành công")
            
            # Chụp màn hình trang đăng nhập
            screenshot_path = reporter.add_screenshot(driver, "Trang đăng nhập")
            if screenshot_path:
                in_thong_bao(f"Đã lưu ảnh màn hình trang đăng nhập: {screenshot_path}")
                
        except TimeoutException:
            error_msg = "Lỗi: Không thể tải trang đăng nhập. Kiểm tra kết nối hoặc URL."
            in_thong_bao(error_msg, "LOI")
            reporter.add_step("Truy cập trang đăng nhập", "FAIL", url_dang_nhap, error_msg)
            reporter.save_report()
            return
        except Exception as e:
            error_msg = f"Lỗi khi tải trang: {str(e)}"
            in_thong_bao(error_msg, "LOI")
            reporter.add_step("Truy cập trang đăng nhập", "ERROR", url_dang_nhap, error_msg)
            reporter.save_report()
            return

        # Tìm và điền thông tin đăng nhập
        in_thong_bao("Đang điền thông tin đăng nhập...")
        reporter.add_step("Bắt đầu điền thông tin", "INFO", "", "Điền thông tin đăng nhập")
        
        try:
            # Tìm và điền tên đăng nhập (sử dụng ID 'email1')
            try:
                o_tai_khoan = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "email1"))
                )
                o_tai_khoan.clear()
                o_tai_khoan.send_keys(ten_dang_nhap)
                in_thong_bao("   - Đã nhập tên đăng nhập", "OK")
                reporter.add_step("Nhập tên đăng nhập", "PASS", f"ID: email1, Tên đăng nhập: {ten_dang_nhap}", 
                                "Nhập tên đăng nhập thành công")
                time.sleep(1)  # Chờ một chút để đảm báo dữ liệu được nhập
            except Exception as e:
                in_thong_bao(f"   - Không tìm thấy ô tên đăng nhập (ID: email1): {str(e)}", "LOI")
                # Thử tìm bằng name nếu không tìm thấy bằng ID
                try:
                    o_tai_khoan = driver.find_element(By.NAME, "taikhoan")
                    o_tai_khoan.clear()
                    o_tai_khoan.send_keys(ten_dang_nhap)
                    in_thong_bao("   - Đã nhập tên đăng nhập (tìm bằng name)", "OK")
                    reporter.add_step("Nhập tên đăng nhập", "PASS", f"NAME: taikhoan, Tên đăng nhập: {ten_dang_nhap}", 
                                    "Nhập tên đăng nhập thành công (tìm bằng name)")
                    time.sleep(1)
                except Exception as e2:
                    error_msg = f"Không thể tìm thấy ô tên đăng nhập: {str(e2)}"
                    in_thong_bao(f"   - {error_msg}", "LOI")
                    reporter.add_step("Nhập tên đăng nhập", "FAIL", "Tìm kiếm bằng ID và NAME", error_msg)
                    raise

            # Tìm và điền mật khẩu (sử dụng ID 'login-password')
            try:
                o_mat_khau = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "login-password"))
                )
                o_mat_khau.clear()
                o_mat_khau.send_keys(mat_khau)
                in_thong_bao("   - Đã nhập mật khẩu", "OK")
                reporter.add_step("Nhập mật khẩu", "PASS", f"ID: login-password, Mật khẩu: {'*' * len(mat_khau)}", 
                                "Nhập mật khẩu thành công")
                time.sleep(1)  # Chờ một chút để đảm bảo dữ liệu được nhập
            except Exception as e:
                in_thong_bao(f"   - Không tìm thấy ô mật khẩu (ID: login-password): {str(e)}", "LOI")
                # Thử tìm bằng name nếu không tìm thấy bằng ID
                try:
                    o_mat_khau = driver.find_element(By.NAME, "matkhau")
                    o_mat_khau.clear()
                    o_mat_khau.send_keys(mat_khau)
                    in_thong_bao("   - Đã nhập mật khẩu (tìm bằng name)", "OK")
                    reporter.add_step("Nhập mật khẩu", "PASS", f"NAME: matkhau, Mật khẩu: {'*' * len(mat_khau)}", 
                                    "Nhập mật khẩu thành công (tìm bằng name)")
                    time.sleep(1)
                except Exception as e2:
                    error_msg = f"Không thể tìm thấy ô mật khẩu: {str(e2)}"
                    in_thong_bao(f"   - {error_msg}", "LOI")
                    reporter.add_step("Nhập mật khẩu", "FAIL", "Tìm kiếm bằng ID và NAME", error_msg)
                    raise

            # Nhấn nút đăng nhập
            try:
                # Thử tìm nút đăng nhập bằng nhiều cách khác nhau
                try:
                    # Cách 1: Tìm bằng name="dangnhap"
                    nut_dang_nhap = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.NAME, "dangnhap"))
                    )
                    locator_type = "NAME: dangnhap"
                except:
                    # Cách 2: Tìm bằng CSS selector cho nút submit
                    nut_dang_nhap = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit'][name='dangnhap']"))
                    )
                    locator_type = "CSS: input[type='submit'][name='dangnhap']"
                
                # Cuộn đến nút đăng nhập để đảm bảo nó hiển thị
                driver.execute_script("arguments[0].scrollIntoView(true);", nut_dang_nhap)
                time.sleep(1)
                
                # Chụp màn hình trước khi click
                screenshot_before = reporter.add_screenshot(driver, "Trước khi đăng nhập")
                
                # Thử click bằng JavaScript nếu click thông thường không hoạt động
                try:
                    nut_dang_nhap.click()
                    click_method = "click() thông thường"
                except:
                    driver.execute_script("arguments[0].click();", nut_dang_nhap)
                    click_method = "JavaScript click()"
                
                in_thong_bao("   - Đã nhấn nút đăng nhập", "OK")
                reporter.add_step("Nhấn nút đăng nhập", "PASS", 
                                f"Tìm bằng: {locator_type}\nPhương thức click: {click_method}", 
                                "Nhấn nút đăng nhập thành công")
                
                # Chờ chuyển trang hoặc cập nhật giao diện
                in_thong_bao("   - Đang chờ xử lý đăng nhập...")
                time.sleep(3)  # Tăng thời gian chờ để đảm bảo trang đã xử lý xong
                
                # Chụp màn hình sau khi click
                screenshot_after = reporter.add_screenshot(driver, "Sau khi đăng nhập")
                
                # Kiểm tra xem có thông báo lỗi không
                try:
                    thong_bao_loi = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-danger, .error-message")
                    if thong_bao_loi and "Sai tài khoản hoặc mật khẩu" in thong_bao_loi[0].text:
                        error_msg = f"Lỗi đăng nhập: {thong_bao_loi[0].text}"
                        in_thong_bao(f"   - {error_msg}", "LOI")
                        
                        # Chụp màn hình lỗi
                        screenshot_error = reporter.add_screenshot(driver, "Thông báo lỗi đăng nhập")
                        
                        # Thêm bước báo lỗi vào báo cáo
                        reporter.add_step("Kiểm tra đăng nhập", "FAIL", "", 
                                        "Đăng nhập thành công vào hệ thống", 
                                        error_msg)
                        
                        # Lưu báo cáo và kết thúc
                        report_path = reporter.save_report()
                        if report_path:
                            in_thong_bao(f"\nĐã lưu báo cáo đầy đủ tại: {os.path.abspath(report_path)}")
                        
                        # Thoát với mã lỗi
                        sys.exit(1)
                except Exception as e:
                    in_thong_bao(f"   - Lỗi khi kiểm tra thông báo: {str(e)}", "WARNING")
                
            except Exception as e:
                error_msg = f"Lỗi khi nhấn nút đăng nhập: {str(e)}"
                in_thong_bao(f"   - {error_msg}", "LOI")
                reporter.add_step("Nhấn nút đăng nhập", "ERROR", "", "Nhấn nút đăng nhập thành công", error_msg)
                
                # Lưu báo cáo trước khi thoát
                report_path = reporter.save_report()
                if report_path:
                    in_thong_bao(f"\nĐã lưu báo cáo đầy đủ tại: {os.path.abspath(report_path)}")
                
                raise

            # Kiểm tra kết quả đăng nhập
            in_thong_bao("\nĐang kiểm tra kết quả đăng nhập...")
            
            # Chờ một chút để trang xử lý
            time.sleep(2)
            
            # Chụp màn hình trạng thái hiện tại
            reporter.add_screenshot(driver, "Trạng thái sau khi đăng nhập")
            
            # Kiểm tra kết quả đăng nhập
            thanh_cong, thong_bao = kiem_tra_dang_nhap_thanh_cong(driver)
            
            if thanh_cong:
                success_msg = f"Chào mừng {ten_dang_nhap} đã đăng nhập thành công!"
                in_thong_bao(f"ĐĂNG NHẬP THÀNH CÔNG!", "OK")
                in_thong_bao(f"   - {success_msg}")
                in_thong_bao(f"   - URL hiện tại: {driver.current_url}")
                
                # Thêm thông tin thành công vào báo cáo
                reporter.add_step("Xác nhận đăng nhập", "PASS", "", 
                                "Đăng nhập thành công vào hệ thống", 
                                success_msg)
                
            else:
                error_msg = f"ĐĂNG NHẬP THẤT BẠI: {thong_bao}"
                in_thong_bao(error_msg, "LOI")
                
                # Chụp màn hình lỗi
                try:
                    screenshot_path = reporter.add_screenshot(driver, "Màn hình lỗi đăng nhập")
                    if screenshot_path:
                        in_thong_bao(f"   - Đã lưu ảnh màn hình lỗi: {screenshot_path}")
                except Exception as e:
                    in_thong_bao(f"   - Lỗi khi chụp màn hình: {str(e)}", "WARNING")
                
                # Thêm thông tin thất bại vào báo cáo
                reporter.add_step("Xác nhận đăng nhập", "FAIL", "", 
                                "Đăng nhập thành công vào hệ thống", 
                                thong_bao)
                    
        except Exception as e:
            in_thong_bao(f"Lỗi trong quá trình đăng nhập: {str(e)}", "LOI")

        # Lưu báo cáo trước khi đóng
        report_path = reporter.save_report()
        if report_path:
            in_thong_bao(f"\nĐã lưu báo cáo đầy đủ tại: {os.path.abspath(report_path)}")
            
            # Mở file báo cáo tự động
            try:
                if sys.platform.startswith('win'):
                    os.startfile(report_path)
                elif sys.platform.startswith('darwin'):
                    os.system(f'open "{report_path}"')
                else:
                    os.system(f'xdg-open "{report_path}"')
            except Exception as e:
                in_thong_bao(f"Không thể mở file báo cáo tự động: {str(e)}", "WARNING")
        
        # Giữ trình duyệt mở trong 10 giây trước khi đóng
        in_thong_bao("\nSẽ đóng trình duyệt sau 10 giây nữa...")
        for i in range(10, 0, -1):
            print(f"\rĐóng sau {i} giây...", end="")
            time.sleep(1)
        print("\r" + " "*30 + "\r", end="")  # Xóa dòng đếm ngược

    except WebDriverException as e:
        error_msg = str(e)
        if "net::ERR_CONNECTION_REFUSED" in error_msg:
            full_error = "LỖI: Không thể kết nối đến máy chủ. Vui lòng kiểm tra:\n"
            full_error += "1. Đã khởi động XAMPP (Apache và MySQL) chưa?\n"
            full_error += "2. URL đăng nhập có chính xác không?\n"
            full_error += f"   URL hiện tại: {url_dang_nhap}"
            
            in_thong_bao(f"\n{full_error}", "LOI")
            reporter.add_step("Lỗi kết nối", "ERROR", url_dang_nhap, "Kết nối đến máy chủ thành công", full_error)
        else:
            in_thong_bao(f"\nLỖI TRÌNH DUYỆT: {error_msg}", "LOI")
            reporter.add_step("Lỗi trình duyệt", "ERROR", "", "Thao tác với trình duyệt", error_msg)
            
        # Lưu báo cáo trước khi thoát
        report_path = reporter.save_report()
        if report_path:
            in_thong_bao(f"\nĐã lưu báo cáo lỗi tại: {os.path.abspath(report_path)}")
            
    except Exception as e:
        error_msg = f"CÓ LỖI XẢY RA: {str(e)}"
        in_thong_bao(f"\n{error_msg}", "LOI")
        reporter.add_step("Lỗi không xác định", "ERROR", "", "Thực hiện test case", error_msg)
        
        # Lưu báo cáo trước khi thoát
        report_path = reporter.save_report()
        if report_path:
            in_thong_bao(f"\nĐã lưu báo cáo lỗi tại: {os.path.abspath(report_path)}")
        
    finally:
        # Đóng trình duyệt
        try:
            if driver:
                # Chụp màn hình cuối cùng trước khi đóng
                try:
                    reporter.add_screenshot(driver, "Màn hình cuối cùng")
                except Exception as e:
                    in_thong_bao(f"Lỗi khi chụp màn hình cuối: {str(e)}", "WARNING")
                
                # Lưu báo cáo cuối cùng
                try:
                    report_path = reporter.save_report()
                    if report_path:
                        in_thong_bao(f"\nĐã lưu báo cáo cuối cùng tại: {os.path.abspath(report_path)}")
                except Exception as e:
                    in_thong_bao(f"Lỗi khi lưu báo cáo cuối: {str(e)}", "WARNING")
                
                # Đóng trình duyệt
                driver.quit()
                in_thong_bao("\nĐã đóng trình duyệt")
                reporter.add_step("Kết thúc kiểm thử", "INFO", "", "Đóng trình duyệt và kết thúc")
        except Exception as e:
            error_msg = f"Lỗi khi đóng trình duyệt: {str(e)}"
            in_thong_bao(error_msg, "LOI")
            reporter.add_step("Lỗi khi đóng trình duyệt", "ERROR", "", "Đóng trình duyệt", error_msg)
        
    print("\n=== KẾT THÚC CHƯƠNG TRÌNH ===")

if __name__ == "__main__":
    main()
