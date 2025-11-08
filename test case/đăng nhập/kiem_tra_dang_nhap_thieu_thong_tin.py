import os
import time
import sys
import io
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

class TestReporter:
    def __init__(self, test_name, result_dir):
        self.test_name = test_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Kết quả kiểm thử"
        self.current_row = 1
        self.screenshot_count = 0
        self.result_dir = result_dir
        self.setup_worksheet()
        
    def setup_worksheet(self):
        # Thiết lập tiêu đề và định dạng cho bảng tính
        self.ws.column_dimensions['A'].width = 20  # Bước thực hiện
        self.ws.column_dimensions['B'].width = 40  # Dữ liệu đầu vào
        self.ws.column_dimensions['C'].width = 30  # Kết quả mong đợi
        self.ws.column_dimensions['D'].width = 30  # Kết quả thực tế
        self.ws.column_dimensions['E'].width = 15  # Trạng thái
        self.ws.column_dimensions['F'].width = 60  # Ảnh chụp màn hình
        
        # Tạo tiêu đề báo cáo
        self.ws.merge_cells('A1:F1')
        title_cell = self.ws.cell(row=1, column=1, value=f'BÁO CÁO KIỂM THỬ: {self.test_name.upper()}')
        title_cell.font = Font(size=14, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Thêm tiêu đề cột
        headers = ['Bước thực hiện', 'Dữ liệu đầu vào', 'Kết quả mong đợi', 'Kết quả thực tế', 'Trạng thái', 'Ảnh chụp màn hình']
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        self.current_row = 3
    
    def add_step(self, description, input_data='', expected='', actual='', status='INFO', screenshot=None):
        """Thêm một bước kiểm thử vào báo cáo"""
        row_data = [description, input_data, expected, actual, status]
        
        # Thêm dữ liệu vào bảng tính
        for col_num, value in enumerate(row_data, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Định dạng ô trạng thái
            if col_num == 5:  # Cột trạng thái
                if status == 'PASS':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100', bold=True)
                elif status == 'FAIL':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006', bold=True)
                elif status == 'WARNING':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    cell.font = Font(color='9C5700', bold=True)
                elif status == 'ERROR':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='FF0000', bold=True)
                else:  # INFO
                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Thêm ảnh chụp màn hình nếu có
        if screenshot:
            try:
                # Lưu ảnh vào file tạm
                img_path = os.path.join(self.result_dir, f'screenshot_{self.screenshot_count}.png')
                screenshot.save(img_path)
                
                # Thêm ảnh vào Excel
                img = XLImage(img_path)
                img.width = 200
                img.height = 150
                
                # Đặt vị trí ảnh
                img_anchor = f'F{self.current_row}'
                img.anchor = img_anchor
                self.ws.add_image(img)
                
                # Điều chỉnh chiều cao hàng
                self.ws.row_dimensions[self.current_row].height = 120
                
                self.screenshot_count += 1
            except Exception as e:
                print(f"Lỗi khi thêm ảnh vào báo cáo: {str(e)}")
        
        self.current_row += 1
    
    def save_report(self):
        """Lưu báo cáo ra file Excel"""
        # Tự động điều chỉnh độ rộng cột
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Bỏ qua cột ảnh
            if column_letter == 'F':
                continue
                
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # Lưu file
        report_path = os.path.join(self.result_dir, f'ket_qua_kiem_thu_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        self.wb.save(report_path)
        return report_path

def in_thong_bao(buoc, trang_thai=""):
    """Hiển thị thông báo từng bước thực hiện"""
    if trang_thai.upper() == "OK":
        print(f"[✓] {buoc}")
    elif trang_thai.upper() in ["LOI", "ERROR", "FAIL"]:
        print(f"[✗] {buoc}")
    else:
        print(f"[ ] {buoc}")
    sys.stdout.flush()

def chup_man_hinh(driver, file_path):
    """Chụp màn hình và lưu vào file"""
    try:
        # Chụp màn hình và lưu vào file
        driver.save_screenshot(file_path)
        
        # Mở ảnh bằng PIL để xử lý
        img = PILImage.open(file_path)
        
        # Nếu cần resize hoặc xử lý ảnh, có thể thực hiện ở đây
        # Ví dụ: img = img.resize((800, 600), PILImage.ANTIALIAS)
        
        # Lưu lại ảnh
        img.save(file_path)
        
        return img
    except Exception as e:
        print(f"Lỗi khi chụp màn hình: {str(e)}")
        return None

def kiem_tra_thong_bao_loi(driver, reporter=None):
    """Kiểm tra thông báo lỗi trên trang"""
    try:
        # Kiểm tra thông báo lỗi
        thong_bao = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "alert-danger"))
        )
        
        # Chụp ảnh thông báo lỗi nếu có reporter
        if reporter and thong_bao.is_displayed():
            error_screenshot = os.path.join(os.path.dirname(driver.get_screenshot_as_file("temp.png")), "error_screenshot.png")
            error_img = chup_man_hinh(driver, error_screenshot)
            if error_img:
                reporter.add_step("Phát hiện thông báo lỗi",
                                input_data="Hệ thống hiển thị thông báo lỗi",
                                expected="Thông báo lỗi hiển thị rõ ràng",
                                actual=thong_bao.text.strip(),
                                status="PASS" if thong_bao.text.strip() else "FAIL",
                                screenshot=error_img)
        
        return thong_bao.text.strip()
    except TimeoutException:
        return ""

def thuc_hien_test_case(driver, ten_tc, ten_dang_nhap, mat_khau, result_dir, reporter):
    """Thực hiện một test case đăng nhập"""
    print(f"\n{'='*50}")
    print(f"THỰC HIỆN TEST CASE: {ten_tc}")
    print(f"Tài khoản: {ten_dang_nhap if ten_dang_nhap else '(để trống)'}")
    print(f"Mật khẩu: {'*' * len(mat_khau) if mat_khau else '(để trống)'}")
    print(f"{'='*50}\n")
    
    # Thêm thông tin test case vào báo cáo
    reporter.add_step("Bắt đầu test case",
                     input_data=f"Tên test case: {ten_tc}\nTài khoản: {ten_dang_nhap if ten_dang_nhap else '(để trống)'}\nMật khẩu: {'*' * len(mat_khau) if mat_khau else '(để trống)'}",
                     expected="Bắt đầu thực hiện test case",
                     status="INFO")
    
    try:
        # Làm mới trang đăng nhập
        driver.get('http://localhost/webbansach/dang-nhap.php')
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(1)
        
        # Chụp màn hình trước khi đăng nhập
        screenshot_before = os.path.join(result_dir, "truoc_khi_dang_nhap.png")
        before_img = chup_man_hinh(driver, screenshot_before)
        
        # Thêm bước vào báo cáo
        expected_result = "Hiển thị trang đăng nhập với form điền thông tin"
        reporter.add_step("Truy cập trang đăng nhập", 
                         input_data=f"URL: {driver.current_url}",
                         expected=expected_result,
                         actual="Đã tải trang đăng nhập" if "dang-nhap" in driver.current_url else "Không phải trang đăng nhập",
                         status="PASS" if "dang-nhap" in driver.current_url else "FAIL",
                         screenshot=before_img)
        
        if "dang-nhap" not in driver.current_url:
            return False
        
        # Điền tên đăng nhập (nếu có)
        if ten_dang_nhap:
            try:
                o_tai_khoan = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "email1"))
                )
                o_tai_khoan.clear()
                o_tai_khoan.send_keys(ten_dang_nhap)
                in_thong_bao("   - Đã nhập tên đăng nhập", "OK")
                reporter.add_step("Nhập tên đăng nhập", 
                                input_data=f"Tên đăng nhập: {ten_dang_nhap}",
                                expected="Nhập được tên đăng nhập vào trường tương ứng",
                                actual="Đã nhập tên đăng nhập thành công",
                                status="PASS")
            except Exception as e:
                error_msg = f"Lỗi khi nhập tên đăng nhập: {str(e)}"
                in_thong_bao(f"   - {error_msg}", "LOI")
                reporter.add_step("Nhập tên đăng nhập", 
                                input_data=f"Tên đăng nhập: {ten_dang_nhap}",
                                expected="Nhập được tên đăng nhập vào trường tương ứng",
                                actual=error_msg,
                                status="FAIL")
                return False
        
        # Điền mật khẩu (nếu có)
        if mat_khau:
            try:
                o_mat_khau = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "login-password"))
                )
                o_mat_khau.clear()
                o_mat_khau.send_keys(mat_khau)
                in_thong_bao("   - Đã nhập mật khẩu", "OK")
                reporter.add_step("Nhập mật khẩu", 
                                input_data="Mật khẩu: ********",
                                expected="Nhập được mật khẩu vào trường tương ứng",
                                actual="Đã nhập mật khẩu thành công",
                                status="PASS")
            except Exception as e:
                error_msg = f"Lỗi khi nhập mật khẩu: {str(e)}"
                in_thong_bao(f"   - {error_msg}", "LOI")
                reporter.add_step("Nhập mật khẩu", 
                                input_data="Mật khẩu: ********",
                                expected="Nhập được mật khẩu vào trường tương ứng",
                                actual=error_msg,
                                status="FAIL")
                return False
        
        # Chụp màn hình trước khi nhấn đăng nhập
        screenshot_before_submit = os.path.join(result_dir, "truoc_khi_nhan_dang_nhap.png")
        before_submit_img = chup_man_hinh(driver, screenshot_before_submit)
        
        try:
            # Nhấn nút đăng nhập
            nut_dang_nhap = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.NAME, "dangnhap"))
            )
            
            # Cuộn đến nút và click
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'})", nut_dang_nhap)
            time.sleep(1)
            
            # Thêm bước nhấn nút đăng nhập
            reporter.add_step("Nhấn nút đăng nhập",
                            input_data="Nhấn nút 'Đăng nhập' hoặc tương đương",
                            expected="Gửi yêu cầu đăng nhập đến máy chủ",
                            status="INFO",
                            screenshot=before_submit_img)
            
            try:
                nut_dang_nhap.click()
            except:
                driver.execute_script("arguments[0].click();", nut_dang_nhap)
            
            in_thong_bao("   - Đã gửi yêu cầu đăng nhập", "OK")
            
            # Chờ xử lý
            in_thong_bao("\nĐang kiểm tra kết quả...")
            time.sleep(2)  # Chờ hiển thị thông báo lỗi
            
            # Chụp màn hình sau khi đăng nhập
            screenshot_after = os.path.join(result_dir, "sau_khi_dang_nhap.png")
            after_img = chup_man_hinh(driver, screenshot_after)
            
            # Kiểm tra kết quả
            thong_bao = kiem_tra_thong_bao_loi(driver, reporter)
            url_hien_tai = driver.current_url
            in_thong_bao(f"   - URL hiện tại: {url_hien_tai}")
            
            # Xác định kết quả mong đợi dựa trên test case
            expected_result = ""
            if not ten_dang_nhap and not mat_khau:
                expected_result = "Hiển thị thông báo yêu cầu nhập tên đăng nhập và mật khẩu"
            elif not ten_dang_nhap:
                expected_result = "Hiển thị thông báo yêu cầu nhập tên đăng nhập"
            elif not mat_khau:
                expected_result = "Hiển thị thông báo yêu cầu nhập mật khẩu"
            
            # Đánh giá kết quả
            test_passed = ("dang-nhap" in url_hien_tai) or bool(thong_bao)
            
            # Thêm kết quả vào báo cáo
            reporter.add_step("Kiểm tra kết quả đăng nhập",
                            input_data=f"Tài khoản: {ten_dang_nhap if ten_dang_nhap else '(để trống)'}\nMật khẩu: {'*' * len(mat_khau) if mat_khau else '(để trống)'}",
                            expected=expected_result,
                            actual=f"Thông báo: {thong_bao if thong_bao else 'Không có thông báo'}\nURL: {url_hien_tai}",
                            status="PASS" if test_passed else "FAIL",
                            screenshot=after_img)
            
            if test_passed:
                in_thong_bao(f"[✓] TEST CASE THÀNH CÔNG: {ten_tc}", "OK")
                if thong_bao:
                    in_thong_bao(f"   - Thông báo: {thong_bao}")
                return True
            else:
                in_thong_bao(f"[✗] TEST CASE THẤT BẠI: {ten_tc}", "LOI")
                in_thong_bao(f"   - Không phát hiện lỗi như mong đợi")
                return False
                
        except Exception as e:
            in_thong_bao(f"[✗] LỖI KHI THỰC HIỆN ĐĂNG NHẬP: {str(e)}", "LOI")
            return False
            
    except Exception as e:
        error_msg = f"Lỗi khi thực hiện test case: {str(e)}"
        in_thong_bao(f"[✗] {error_msg}", "LOI")
        
        # Chụp màn hình lỗi
        error_screenshot = os.path.join(result_dir, "loi_khi_thuc_hien.png")
        error_img = chup_man_hinh(driver, error_screenshot)
        
        # Thêm lỗi vào báo cáo
        reporter.add_step("Lỗi khi thực hiện test case",
                         input_data=f"Tài khoản: {ten_dang_nhap if ten_dang_nhap else '(để trống)'}\nMật khẩu: {'*' * len(mat_khau) if mat_khau else '(để trống)'}",
                         expected="Thực hiện test case không gặp lỗi",
                         actual=error_msg,
                         status="ERROR",
                         screenshot=error_img)
        
        # Lưu mã nguồn trang để debug
        error_page = os.path.join(result_dir, "ma_nguon_loi.html")
        with open(error_page, "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        in_thong_bao(f"   - Đã lưu mã nguồn trang: {error_page}")
        
        return False

def main():
    # Tạo cấu trúc thư mục kết quả
    base_dir = os.path.abspath(os.path.join('test_results', 'dang_nhap'))
    test_type = 'thieu_thong_tin'
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Tạo các thư mục cần thiết
    result_dir = os.path.join(base_dir, test_type, f"test_run_{timestamp}")
    os.makedirs(result_dir, exist_ok=True)
    
    # Khởi tạo báo cáo
    reporter = TestReporter("Kiểm tra đăng nhập thiếu thông tin", result_dir)
    
    # Tùy chọn cho Chrome
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-notifications')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    print("\n=== KIỂM TRA ĐĂNG NHẬP THIẾU THÔNG TIN ===\n")
    print("Mục đích: Kiểm tra xử lý khi thiếu thông tin đăng nhập\n")

    driver = None
    try:
        # Khởi tạo trình duyệt
        in_thong_bao("Đang khởi tạo trình duyệt...")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(30)  # Thời gian chờ tải trang tối đa 30 giây
        in_thong_bao("Đã khởi tạo trình duyệt thành công", "OK")
        
        # Danh sách các test case
        test_cases = [
            {
                "ten_tc": "Thiếu mật khẩu",
                "ten_dang_nhap": "tai_khoan_test",
                "mat_khau": ""
            },
            {
                "ten_tc": "Thiếu tên đăng nhập",
                "ten_dang_nhap": "",
                "mat_khau": "mat_khau_123"
            }
        ]
        
        # Thực hiện từng test case
        for i, test_case in enumerate(test_cases, 1):
            in_thong_bao(f"\n{'='*20} TEST CASE {i}/{len(test_cases)}: {test_case['ten_tc']} {'='*20}")
            
            # Tạo thư mục riêng cho mỗi test case
            test_case_name = test_case["ten_tc"].lower().replace(" ", "_")
            test_case_dir = os.path.join(result_dir, f"tc_{i:02d}_{test_case_name}")
            os.makedirs(test_case_dir, exist_ok=True)
            
            # Thêm tiêu đề test case vào báo cáo
            reporter.ws.merge_cells(f'A{reporter.current_row}:F{reporter.current_row}')
            title_cell = reporter.ws.cell(row=reporter.current_row, column=1, 
                                        value=f"TEST CASE {i}: {test_case['ten_tc'].upper()}")
            title_cell.font = Font(bold=True, size=12, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            reporter.current_row += 1
            
            # Thực hiện test case
            test_result = thuc_hien_test_case(
                driver=driver,
                ten_tc=test_case["ten_tc"],
                ten_dang_nhap=test_case["ten_dang_nhap"],
                mat_khau=test_case["mat_khau"],
                result_dir=test_case_dir,
                reporter=reporter
            )
            
            # Thêm dòng phân cách giữa các test case
            reporter.current_row += 1
            
            # Chờ giữa các test case
            if i < len(test_cases):
                in_thong_bao(f"\nChuẩn bị cho test case tiếp theo...")
                time.sleep(2)
        
        # Lưu báo cáo
        report_path = reporter.save_report()
        in_thong_bao("\nTất cả các test case đã hoàn thành!", "OK")
        in_thong_bao(f"Đã lưu báo cáo tại: {report_path}", "OK")
        
    except WebDriverException as e:
        error_msg = f"Lỗi trình duyệt: {str(e)}"
        in_thong_bao(f"\n{error_msg}", "LOI")
        
        # Thêm lỗi vào báo cáo
        reporter.add_step("Lỗi trình duyệt",
                         expected="Trình duyệt hoạt động bình thường",
                         actual=error_msg,
                         status="ERROR")
        
        if "net::ERR_CONNECTION_REFUSED" in str(e):
            in_thong_bao("Vui lòng kiểm tra kết nối đến máy chủ.", "LOI")
    
    except Exception as e:
        error_msg = f"Lỗi không xác định: {str(e)}"
        in_thong_bao(f"\n{error_msg}", "LOI")
        
        # Thêm lỗi vào báo cáo
        reporter.add_step("Lỗi không xác định",
                         expected="Chương trình chạy không có lỗi",
                         actual=error_msg,
                         status="ERROR")
    
    finally:
        # Đóng trình duyệt
        try:
            if driver:
                # Chụp màn hình cuối cùng trước khi đóng
                final_screenshot = os.path.join(result_dir, "ket_thuc_test.png")
                final_img = chup_man_hinh(driver, final_screenshot)
                if final_img:
                    reporter.add_final_screenshot(final_img)
                
                # Lưu báo cáo cuối cùng
                report_path = reporter.save_report()
                in_thong_bao(f"\nĐã lưu báo cáo tại: {report_path}", "OK")
                
                print("\nSẽ đóng trình duyệt sau 10 giây nữa...")
                for i in range(10, 0, -1):
                    print(f"\rĐóng sau {i} giây...", end="")
                    time.sleep(1)
                print("\r" + " "*30 + "\r", end="")  # Xóa dòng đếm ngược
                driver.quit()
                in_thong_bao("Đã đóng trình duyệt")
        except Exception as e:
            error_msg = f"Lỗi khi đóng trình duyệt: {str(e)}"
            in_thong_bao(error_msg, "LOI")
            
            # Thêm lỗi vào báo cáo
            reporter.add_step("Lỗi khi đóng trình duyệt",
                             expected="Đóng trình duyệt thành công",
                             actual=error_msg,
                             status="ERROR")
        
    print("\n=== KẾT THÚC KIỂM TRA ===")

if __name__ == "__main__":
    main()
