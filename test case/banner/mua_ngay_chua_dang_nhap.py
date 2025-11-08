import os
import time
import sys
import io
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import traceback

class TestReporter:
    def __init__(self, test_name):
        self.test_name = test_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Kết quả kiểm thử"
        self.current_row = 1
        self.screenshot_count = 0
        self.test_steps = []
        self.setup_worksheet()
        
    def setup_worksheet(self):
        # Đặt tiêu đề cho các cột
        headers = ["Bước thực hiện", "Dữ liệu đầu vào", "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái", "Ghi chú"]
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            
        # Đặt chiều rộng cột
        self.ws.column_dimensions['A'].width = 30  # Bước thực hiện
        self.ws.column_dimensions['B'].width = 25  # Dữ liệu đầu vào
        self.ws.column_dimensions['C'].width = 30  # Kết quả mong đợi
        self.ws.column_dimensions['D'].width = 30  # Kết quả thực tế
        self.ws.column_dimensions['E'].width = 15  # Trạng thái
        self.ws.column_dimensions['F'].width = 40  # Ghi chú
        
        self.current_row += 1
        
    def add_step(self, description, status='PASS', input_data='', expected='', notes=''):
        """Thêm bước kiểm thử vào báo cáo"""
        # Lưu thông tin bước vào danh sách
        step_info = {
            'description': description,
            'status': status,
            'input': input_data,
            'expected': expected,
            'notes': notes,
            'screenshot': None
        }
        self.test_steps.append(step_info)
        
        # Thêm vào worksheet
        self.ws.append([
            description,
            input_data if isinstance(input_data, str) else str(input_data),
            expected,
            '',  # Kết quả thực tế sẽ được cập nhật sau
            status,
            notes
        ])
        
        # Định dạng ô trạng thái
        status_cell = self.ws.cell(row=self.current_row, column=5)
        if status == 'PASS':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status in ['FAIL', 'ERROR']:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
        # Căn giữa các ô
        for col in range(1, 7):
            self.ws.cell(row=self.current_row, column=col).alignment = Alignment(
                vertical='center', 
                wrap_text=True
            )
            
        self.current_row += 1
        
    def add_screenshot(self, driver, description=""):
        """Chụp màn hình và thêm vào báo cáo"""
        try:
            # Tạo thư mục screenshots nếu chưa tồn tại
            screenshots_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screenshots")
            if not os.path.exists(screenshots_dir):
                os.makedirs(screenshots_dir)
                
            # Tạo tên file ảnh
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            filename = os.path.join(screenshots_dir, f"screenshot_{self.screenshot_count}_{timestamp}.png")
            
            # Chụp màn hình và lưu vào file
            driver.save_screenshot(filename)
            
            # Thêm ảnh vào báo cáo
            if os.path.exists(filename):
                # Thêm dòng mô tả ảnh
                self.ws.cell(row=self.current_row, column=1, value=f"Hình ảnh: {description}")
                self.ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=6)
                self.current_row += 1
                
                # Thêm ảnh vào báo cáo
                img = PILImage.open(filename)
                img.thumbnail((800, 600))  # Giảm kích thước ảnh
                img_io = io.BytesIO()
                img.save(img_io, format='PNG')
                img_io.seek(0)
                
                img_for_excel = XLImage(img_io)
                self.ws.add_image(img_for_excel, f'A{self.current_row}')
                
                # Điều chỉnh chiều cao hàng để chứa ảnh
                self.ws.row_dimensions[self.current_row].height = img.height * 0.75  # Tỉ lệ phù hợp
                
                # Tăng số thứ tự ảnh
                self.screenshot_count += 1
                self.current_row += 1
                
                # Thêm dòng trống sau ảnh
                self.current_row += 1
                
                return filename
            return None
            
        except Exception as e:
            print(f"Lỗi khi thêm ảnh vào báo cáo: {str(e)}")
            return None
    
    def save_report(self, folder_path='test_results'):
        """Lưu báo cáo ra file Excel"""
        try:
            # Tạo thư mục nếu chưa tồn tại
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                
            # Tạo tên file báo cáo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = os.path.join(folder_path, f"ket_qua_kiem_thu_{timestamp}.xlsx")
            
            # Lưu file
            self.wb.save(report_filename)
            print(f"\nĐã lưu báo cáo tại: {os.path.abspath(report_filename)}")
            return report_filename
            
        except Exception as e:
            print(f"Lỗi khi lưu báo cáo: {str(e)}")
            return None

def in_thong_bao(buoc, trang_thai="", reporter=None, input_data="", expected="", notes=""):
    """Hiển thị thông báo từng bước thực hiện và thêm vào báo cáo nếu có reporter"""
    # Hiển thị ra console
    if trang_thai.upper() == "OK":
        print(f"[✓] {buoc}")
        status = "PASS"
    elif trang_thai.upper() in ["LOI", "LỖI"]:
        print(f"[✗] {buoc}")
        status = "ERROR" if trang_thai.upper() == "LỖI" else "FAIL"
    else:
        print(f"[ ] {buoc}")
        status = "INFO"
    
    sys.stdout.flush()
    
    # Thêm vào báo cáo nếu có reporter
    if reporter:
        reporter.add_step(
            description=buoc,
            status=status,
            input_data=input_data,
            expected=expected,
            notes=notes
        )

def chup_man_hinh(driver, ten_file=None, reporter=None, description=""):
    """Chụp màn hình và lưu vào file hoặc thêm vào báo cáo"""
    try:
        if reporter and not ten_file:
            # Nếu có reporter và không chỉ định tên file, thêm ảnh trực tiếp vào báo cáo
            return reporter.add_screenshot(driver, description)
        else:
            # Nếu chỉ định tên file, lưu ảnh ra file
            driver.save_screenshot(ten_file)
            return ten_file
    except Exception as e:
        error_msg = f"Lỗi khi chụp màn hình: {str(e)}"
        in_thong_bao(error_msg, "LỖI", reporter)
        return None

def kiem_tra_phan_tu(driver, loai, gia_tri, timeout=10):
    """Kiểm tra sự tồn tại của phần tử"""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((loai, gia_tri))
        )
        return True
    except TimeoutException:
        return False

def kiem_tra_dang_nhap_thanh_cong(driver):
    """Kiểm tra xem đăng nhập có thành công không"""
    try:
        # Kiểm tra xem có thông báo lỗi không
        thong_bao_loi = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-danger, .error-message")
        if thong_bao_loi:
            return False, f"Lỗi: {thong_bao_loi[0].text}"
            
        # Kiểm tra URL hiện tại
        if "index.php" in driver.current_url:
            return True, "Đăng nhập thành công!"
            
        # Kiểm tra xem có thông báo thành công không
        thong_bao_thanh_cong = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-success")
        if thong_bao_thanh_cong and "thành công" in thong_bao_thanh_cong[0].text.lower():
            return True, "Đăng nhập thành công!"
            
        # Kiểm tra xem có nút đăng xuất không
        if kiem_tra_phan_tu(driver, By.XPATH, "//a[contains(@href, 'dang-xuat') or contains(text(), 'Đăng xuất')]"):
            return True, "Đăng nhập thành công!"
            
        return False, "Không xác định được trạng thái đăng nhập"
        
    except Exception as e:
        return False, f"Lỗi khi kiểm tra đăng nhập: {str(e)}"

def wait_and_click(driver, locator, description="element", timeout=10):
    """Đợi và click vào phần tử với khả năng chịu lỗi"""
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(locator)
        )
        # Cuộn đến phần tử và đánh dấu
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'})", element)
        driver.execute_script("arguments[0].style.border='3px solid red';", element)
        time.sleep(0.5)
        
        # Thử click bằng JavaScript
        driver.execute_script("arguments[0].click();", element)
        time.sleep(1)  # Đợi một chút sau khi click
        return True
    except Exception as e:
        in_thong_bao(f"Không thể click {description}: {str(e)}", "LỖI")
        return False

def dang_nhap(driver, tai_khoan, mat_khau, thu_muc_ket_qua):
    """Thực hiện đăng nhập vào hệ thống"""
    try:
        in_thong_bao("Bắt đầu quá trình đăng nhập...")
        
        # Tìm và điền tên đăng nhập
        try:
            username_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "username"))
            )
            username_field.clear()
            username_field.send_keys(tai_khoan)
            in_thong_bao("Đã điền tên đăng nhập", "OK")
        except Exception as e:
            in_thong_bao(f"Lỗi khi điền tên đăng nhập: {str(e)}", "LỖI")
            return False
        
        # Tìm và điền mật khẩu
        try:
            password_field = driver.find_element(By.NAME, "password")
            password_field.clear()
            password_field.send_keys(mat_khau)
            in_thong_bao("Đã điền mật khẩu", "OK")
        except Exception as e:
            in_thong_bao(f"Lỗi khi điền mật khẩu: {str(e)}", "LỖI")
            return False
        
        # Chụp màn hình trước khi đăng nhập
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_dang_nhap.png"))
        
        # Tìm và nhấn nút đăng nhập
        if not wait_and_click(driver, (By.CSS_SELECTOR, "button[type='submit']"), "nút đăng nhập"):
            in_thong_bao("Không tìm thấy nút đăng nhập", "LỖI")
            return False
        
        # Chờ đăng nhập xử lý
        time.sleep(3)
        
        # Kiểm tra đăng nhập thành công
        thanh_cong, thong_bao = kiem_tra_dang_nhap_thanh_cong(driver)
        if thanh_cong:
            in_thong_bao("Đăng nhập thành công!", "OK")
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "dang_nhap_thanh_cong.png"))
            return True
        else:
            in_thong_bao(f"Đăng nhập thất bại: {thong_bao}", "LỖI")
            return False
            
    except Exception as e:
        in_thong_bao(f"Lỗi trong quá trình đăng nhập: {str(e)}", "LỖI")
        return False

def click_mua_ngay(driver, thu_muc_ket_qua, reporter=None):
    try:
        in_thong_bao("Đang tìm nút Mua ngay trên banner...", reporter=reporter,
                    expected="Tìm thấy và click vào nút Mua ngay")
        
        # Danh sách các cách tìm nút Mua ngay
        locators = [
            (By.CSS_SELECTOR, "a.btn-outlined--primary"),
            (By.CSS_SELECTOR, ".single-slide .btn-outlined--primary"),
            (By.XPATH, "//a[contains(translate(., 'MUA NGAY', 'mua ngay'), 'mua ngay')]"),
            (By.XPATH, "//a[contains(@class, 'btn-outlined')]"),
            (By.CSS_SELECTOR, ".home-content a[class*='btn']")
        ]
        
        for locator in locators:
            try:
                elements = driver.find_elements(*locator)
                for element in elements:
                    try:
                        if element.is_displayed() and element.is_enabled():
                            text = element.text.lower()
                            if 'mua' in text and 'ngay' in text:
                                # Cuộn đến nút và đánh dấu
                                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                                driver.execute_script("arguments[0].style.border='3px solid red';", element)
                                time.sleep(1)
                                
                                # Chụp màn hình trước khi click
                                chup_man_hinh(driver, reporter=reporter, description="Trước khi click Mua ngay")
                                
                                # Click bằng JavaScript
                                driver.execute_script("arguments[0].click();", element)
                                in_thong_bao("Đã nhấn nút Mua ngay", "OK", reporter,
                                            input_data=element.text,
                                            expected="Chuyển hướng đến trang đăng nhập")
                                time.sleep(3)  # Chờ chuyển trang
                                return True
                    except:
                        continue
            except:
                continue
                
        error_msg = "Không tìm thấy nút Mua ngay phù hợp"
        in_thong_bao(error_msg, "LỖI", reporter,
                    expected="Tìm thấy và click vào nút Mua ngay",
                    notes=error_msg)
        return False
        
    except Exception as e:
        error_msg = f"Lỗi khi click nút Mua ngay: {str(e)}"
        in_thong_bao(error_msg, "LỖI", reporter,
                    expected="Tìm thấy và click vào nút Mua ngay",
                    notes=error_msg)
        return False

def dien_thong_tin_giao_hang(driver, thu_muc_ket_qua):
    try:
        in_thong_bao("6. Đang điền thông tin giao hàng...")
        
        # Thông tin giao hàng mẫu
        thong_tin = {
            'sonha': '123 Đường ABC',
            'thonxom': 'Tổ 1',
            'phuongxa': 'Phường 10',
            'huyen': 'Quận 5',
            'tinhthanh': 'TP. Hồ Chí Minh'
        }
        
        # Điền thông tin vào các trường
        for field, value in thong_tin.items():
            try:
                element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, f'.{field}'))
                )
                element.clear()
                element.send_keys(value)
                in_thong_bao(f"   - Đã điền {field}: {value}", "OK")
                time.sleep(0.5)
            except Exception as e:
                in_thong_bao(f"   - Không tìm thấy trường {field}: {str(e)}", "LỖI")
                raise
        
        # Chụp màn hình sau khi điền thông tin
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "dien_thong_tin_giao_hang.png"))
        in_thong_bao("   - Đã điền đầy đủ thông tin giao hàng", "OK")
        return True
        
    except Exception as e:
        in_thong_bao(f"Lỗi khi điền thông tin giao hàng: {str(e)}", "LỖI")
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_dien_thong_tin.png"))
        return False

def xac_nhan_don_hang(driver, thu_muc_ket_qua, reporter=None):
    try:
        in_thong_bao("7. Đang xác nhận đơn hàng...", reporter=reporter,
                    expected="Xác nhận đơn hàng thành công")
        
        # Cuộn xuống phần xác nhận đơn hàng
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        
        # Tích vào ô đồng ý điều khoản
        try:
            checkbox = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.ID, "accept_terms2"))
            )
            if not checkbox.is_selected():
                driver.execute_script("arguments[0].click();", checkbox)
                in_thong_bao("   - Đã tích vào ô đồng ý điều khoản", "OK", reporter)
        except Exception as e:
            in_thong_bao(f"   - Không tìm thấy ô đồng ý điều khoản: {str(e)}", "CẢNH BÁO", reporter)
        
        # Nhấn nút đặt hàng
        try:
            nut_dat_hang = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ".dathang"))
            )
            # Cuộn đến nút đặt hàng
            driver.execute_script("arguments[0].scrollIntoView(true);", nut_dat_hang)
            time.sleep(1)
            
            # Đánh dấu nút đặt hàng
            driver.execute_script("arguments[0].style.border='3px solid red';", nut_dat_hang)
            
            # Chụp màn hình trước khi click (không lưu vào Excel)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_nhan_dat_hang.png"))
            
            # Click bằng JavaScript
            driver.execute_script("arguments[0].click();", nut_dat_hang)
            in_thong_bao("   - Đã nhấn nút đặt hàng", "OK", reporter)
            time.sleep(3)  # Chờ xử lý đơn hàng
            
            # Kiểm tra đặt hàng thành công
            if "hoan-thanh-thanh-toan" in driver.current_url or "success" in driver.current_url:
                in_thong_bao("   - Đặt hàng thành công!", "THÀNH CÔNG", reporter)
                
                # Chỉ chụp màn hình khi thành công và thêm vào báo cáo Excel
                if reporter:
                    # Chờ cho trang tải xong
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    time.sleep(2)  # Đợi thêm 2 giây để đảm bảo trang đã tải xong
                    
                    # Chụp màn hình và thêm vào báo cáo
                    screenshot_path = chup_man_hinh(
                        driver, 
                        reporter=reporter, 
                        description="Xác nhận đơn hàng thành công"
                    )
                    
                    if screenshot_path:
                        in_thong_bao("   - Đã lưu ảnh xác nhận đơn hàng vào báo cáo", "OK", reporter)
                    else:
                        in_thong_bao("   - Không thể lưu ảnh xác nhận đơn hàng", "CẢNH BÁO", reporter)
                
                return True
            else:
                # Kiểm tra xem có thông báo lỗi không
                try:
                    thong_bao_loi = driver.find_elements(By.CSS_SELECTOR, ".alert.alert-danger, .error-message")
                    if thong_bao_loi:
                        error_msg = f"Lỗi khi đặt hàng: {thong_bao_loi[0].text}"
                        in_thong_bao(f"   - {error_msg}", "LỖI", reporter, notes=error_msg)
                    else:
                        error_msg = "Không xác định được kết quả đặt hàng"
                        in_thong_bao(f"   - {error_msg}", "CẢNH BÁO", reporter, notes=error_msg)
                except Exception as e:
                    error_msg = f"Lỗi khi kiểm tra thông báo: {str(e)}"
                    in_thong_bao(f"   - {error_msg}", "LỖI", reporter, notes=error_msg)
                
                # Chụp màn hình lỗi nhưng không thêm vào Excel
                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_dat_hang.png"))
                return False
                
        except Exception as e:
            error_msg = f"Lỗi khi nhấn nút đặt hàng: {str(e)}"
            in_thong_bao(f"   - {error_msg}", "LỖI", reporter, notes=error_msg)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_nhan_dat_hang.png"))
            return False
            
    except Exception as e:
        error_msg = f"Lỗi khi xác nhận đơn hàng: {str(e)}"
        in_thong_bao(error_msg, "LỖI", reporter, notes=error_msg)
        return False

def main():
    # Tạo thư mục kết quả
    thu_muc_ket_qua = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_results")
    if not os.path.exists(thu_muc_ket_qua):
        os.makedirs(thu_muc_ket_qua)
    in_thong_bao(f"Thư mục kết quả: {thu_muc_ket_qua}")
    
    # Khởi tạo báo cáo
    reporter = TestReporter("Kiểm thử tính năng Mua ngay chưa đăng nhập")
    reporter.add_step("Bắt đầu kiểm thử", "INFO", "", "Khởi tạo môi trường kiểm thử")
    
    # Tùy chọn cho Chrome
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')  # Mở cửa sổ lớn nhất
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # Tắt thông báo console
    
    # Khởi tạo trình duyệt
    in_thong_bao("Đang khởi tạo trình duyệt...")
    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(30)
        in_thong_bao("Đã khởi tạo trình duyệt thành công", "OK", reporter,
                    input_data="Chrome WebDriver",
                    expected="Trình duyệt khởi động thành công")
    except Exception as e:
        error_msg = f"Lỗi khi khởi tạo trình duyệt: {str(e)}"
        in_thong_bao(error_msg, "LỖI", reporter,
                    input_data="Chrome WebDriver",
                    expected="Trình duyệt khởi động thành công",
                    notes=error_msg)
        
        # Thử cách khác nếu cách trên thất bại
        try:
            in_thong_bao("Thử khởi tạo trình duyệt với Chrome mặc định...", reporter=reporter)
            driver = webdriver.Chrome(options=options)
            in_thong_bao("Đã khởi tạo trình duyệt thành công (dùng Chrome mặc định)", "OK", reporter,
                        input_data="Chrome mặc định",
                        expected="Trình duyệt khởi động thành công")
        except Exception as e2:
            error_msg = f"Không thể khởi tạo trình duyệt: {str(e2)}"
            in_thong_bao(error_msg, "LỖI", reporter,
                        input_data="Chrome mặc định",
                        expected="Trình duyệt khởi động thành công",
                        notes=error_msg)
            in_thong_bao("Vui lòng kiểm tra cài đặt Chrome và ChromeDriver", "LỖI", reporter)
            reporter.save_report(thu_muc_ket_qua)
            return
    
    try:
        # Bước 1: Mở trang chủ
        in_thong_bao("1. Đang mở trang chủ...", reporter=reporter,
                    expected="Trang chủ được tải thành công")
        driver.get("http://localhost/webbansach/")
        time.sleep(2)
        
        # Chụp màn hình trang chủ
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_chu.png"))
        in_thong_bao("   - Đã mở trang chủ thành công", "OK", reporter)
        
        # Bước 2: Thực hiện mua ngay từ banner
        in_thong_bao("2. Đang thực hiện mua ngay từ banner...", reporter=reporter,
                    expected="Tìm thấy và click vào nút Mua ngay")
        if not click_mua_ngay(driver, thu_muc_ket_qua, reporter):
            in_thong_bao("Không thể thực hiện test case do không tìm thấy nút Mua ngay", "LỖI", reporter)
            reporter.save_report(thu_muc_ket_qua)
            driver.quit()
            return
        
        # Chờ một chút để trang xử lý
        time.sleep(3)
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "sau_khi_nhan_mua_ngay.png"))
        
        # Bước 3: Kiểm tra chuyển hướng đến trang giỏ hàng hoặc đăng nhập
        in_thong_bao("3. Kiểm tra chuyển hướng...", reporter=reporter,
                    expected="Chuyển hướng đến trang đăng nhập khi chưa đăng nhập")
        time.sleep(2)
        
        if "gio-hang" in driver.current_url:
            in_thong_bao("   - Đã chuyển đến trang giỏ hàng", "OK", reporter)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_gio_hang.png"))
            
            # Bước 3.1: Nhấn nút Thanh Toán trong giỏ hàng
            in_thong_bao("3.1. Đang nhấn nút Thanh Toán...", reporter=reporter,
                        expected="Chuyển hướng đến trang đăng nhập khi chưa đăng nhập")
            try:
                # Tìm nút Thanh Toán
                nut_thanh_toan = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, "checkout-btn"))
                )
                
                # Cuộn đến nút Thanh Toán
                driver.execute_script("arguments[0].scrollIntoView(true);", nut_thanh_toan)
                time.sleep(1)
                
                # Đánh dấu nút Thanh Toán
                driver.execute_script("arguments[0].style.border='3px solid red';", nut_thanh_toan)
                
                # Chụp màn hình trước khi click
                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_nhan_thanh_toan.png"))
                
                # Click vào nút Thanh Toán
                nut_thanh_toan.click()
                in_thong_bao("   - Đã nhấn nút Thanh Toán", "OK", reporter)
                time.sleep(3)  # Chờ chuyển hướng
                
                # Kiểm tra đã chuyển đến trang đăng nhập chưa
                if "dang-nhap" in driver.current_url:
                    in_thong_bao("   - Đã chuyển đến trang đăng nhập", "OK", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_dang_nhap.png"))
                else:
                    in_thong_bao(f"   - Cảnh báo: Không chuyển đến trang đăng nhập. URL hiện tại: {driver.current_url}", "CẢNH BÁO", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "khong_chuyen_den_dang_nhap.png"))
                    
            except Exception as e:
                in_thong_bao(f"   - Lỗi khi nhấn nút Thanh Toán: {str(e)}", "LỖI", reporter)
                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_nhan_thanh_toan.png"))
                return
                
        elif "dang-nhap" in driver.current_url:
            in_thong_bao("   - Đã chuyển trực tiếp đến trang đăng nhập", "OK", reporter)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_dang_nhap.png"))
            
        else:
            in_thong_bao(f"   - Cảnh báo: Không chuyển đến trang giỏ hàng hoặc đăng nhập. URL hiện tại: {driver.current_url}", "CẢNH BÁO", reporter)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "khong_chuyen_huong_dung.png"))
            return
        
        # Bước 4: Thực hiện đăng nhập
        in_thong_bao("4. Đang thực hiện đăng nhập...", reporter=reporter,
                    expected="Đăng nhập thành công")
        TAI_KHOAN = "quân"
        MAT_KHAU = "1"
        
        try:
            # Điền thông tin đăng nhập
            in_thong_bao("   - Đang điền thông tin đăng nhập...", reporter=reporter)
            
            # Tìm và điền tên đăng nhập
            try:
                # Thử tìm bằng ID trước (email1)
                o_tai_khoan = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "email1"))
                )
                o_tai_khoan.clear()
                o_tai_khoan.send_keys(TAI_KHOAN)
                in_thong_bao("   - Đã nhập tên đăng nhập", "OK", reporter)
                time.sleep(1)
            except Exception as e:
                try:
                    # Nếu không tìm thấy bằng ID, thử tìm bằng name
                    o_tai_khoan = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.NAME, "taikhoan"))
                    )
                    o_tai_khoan.clear()
                    o_tai_khoan.send_keys(TAI_KHOAN)
                    in_thong_bao("   - Đã nhập tên đăng nhập (tìm bằng name)", "OK", reporter)
                    time.sleep(1)
                except Exception as e2:
                    in_thong_bao(f"   - Không tìm thấy ô tên đăng nhập: {str(e2)}", "LỖI", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "khong_tim_thay_tai_khoan.png"))
                    raise
            
            # Tìm và điền mật khẩu
            try:
                # Thử tìm bằng ID trước (login-password)
                o_mat_khau = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "login-password"))
                )
                o_mat_khau.clear()
                o_mat_khau.send_keys(MAT_KHAU)
                in_thong_bao("   - Đã nhập mật khẩu", "OK", reporter)
                time.sleep(1)
            except Exception as e:
                try:
                    # Nếu không tìm thấy bằng ID, thử tìm bằng name
                    o_mat_khau = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.NAME, "matkhau"))
                    )
                    o_mat_khau.clear()
                    o_mat_khau.send_keys(MAT_KHAU)
                    in_thong_bao("   - Đã nhập mật khẩu (tìm bằng name)", "OK", reporter)
                    time.sleep(1)
                except Exception as e2:
                    in_thong_bao(f"   - Không tìm thấy ô mật khẩu: {str(e2)}", "LỖI", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "khong_tim_thay_mat_khau.png"))
                    raise
            
            # Chụp màn hình trước khi đăng nhập
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_dang_nhap.png"))
            
            # Nhấn nút đăng nhập
            try:
                # Thử tìm nút đăng nhập bằng name
                try:
                    nut_dang_nhap = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.NAME, "dangnhap"))
                    )
                except:
                    # Nếu không tìm thấy bằng name, thử tìm bằng CSS selector
                    nut_dang_nhap = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "input[type='submit'][value='Đăng Nhập']"))
                    )
                
                # Cuộn đến nút đăng nhập
                driver.execute_script("arguments[0].scrollIntoView(true);", nut_dang_nhap)
                time.sleep(1)
                
                # Đánh dấu nút đăng nhập
                driver.execute_script("arguments[0].style.border='3px solid red';", nut_dang_nhap)
                
                # Chụp màn hình trước khi click
                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_nhan_dang_nhap.png"))
                
                # Click bằng JavaScript
                driver.execute_script("arguments[0].click();", nut_dang_nhap)
                in_thong_bao("   - Đã nhấn nút đăng nhập", "OK", reporter)
                time.sleep(3)  # Chờ đăng nhập xử lý
                
                # Kiểm tra đăng nhập thành công
                if kiem_tra_dang_nhap_thanh_cong(driver):
                    in_thong_bao("   - Đăng nhập thành công", "OK", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "dang_nhap_thanh_cong.png"))
                    
                    # Quay lại trang chủ sau khi đăng nhập
                    in_thong_bao("   - Đang quay về trang chủ...", "THÔNG BÁO", reporter)
                    driver.get("http://localhost/webbansach/")
                    time.sleep(2)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_chu_sau_dang_nhap.png"))
                    
                    # Thực hiện lại các bước mua hàng
                    in_thong_bao("   - Đang thực hiện mua ngay từ banner...", "THÔNG BÁO", reporter)
                    if not click_mua_ngay(driver, thu_muc_ket_qua, reporter):
                        in_thong_bao("   - Không thể thực hiện mua ngay", "LỖI", reporter)
                        return
                        
                    time.sleep(3)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "sau_khi_nhan_mua_ngay_lan_2.png"))
                    
                    # Kiểm tra chuyển hướng đến trang thanh toán
                    if "thanh-toan" not in driver.current_url:
                        in_thong_bao("   - Chưa chuyển đến trang thanh toán, kiểm tra giỏ hàng...", "CẢNH BÁO", reporter)
                        # Nếu đang ở trang giỏ hàng, nhấn nút Thanh Toán
                        if "gio-hang" in driver.current_url:
                            try:
                                nut_thanh_toan = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.CLASS_NAME, "checkout-btn"))
                                )
                                driver.execute_script("arguments[0].scrollIntoView(true);", nut_thanh_toan)
                                time.sleep(1)
                                driver.execute_script("arguments[0].style.border='3px solid red';", nut_thanh_toan)
                                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "truoc_khi_nhan_thanh_toan_lan_2.png"))
                                nut_thanh_toan.click()
                                in_thong_bao("   - Đã nhấn nút Thanh Toán lần 2", "OK", reporter)
                                time.sleep(3)
                            except Exception as e:
                                in_thong_bao(f"   - Lỗi khi nhấn nút Thanh Toán lần 2: {str(e)}", "LỖI", reporter)
                                return
                    
                    # Kiểm tra lại xem đã chuyển đến trang thanh toán chưa
                    if "thanh-toan" not in driver.current_url:
                        in_thong_bao("   - Không thể chuyển đến trang thanh toán", "LỖI", reporter)
                        return
                        
                else:
                    in_thong_bao("   - Đăng nhập không thành công", "LỖI", reporter)
                    chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "dang_nhap_that_bai.png"))
                    return
                    
            except Exception as e:
                in_thong_bao(f"   - Lỗi khi nhấn nút đăng nhập: {str(e)}", "LỖI", reporter)
                chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_nhan_nut_dang_nhap.png"))
                raise
                
        except Exception as e:
            in_thong_bao(f"   - Lỗi khi thực hiện đăng nhập: {str(e)}", "LỖI", reporter)
            chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi_dang_nhap.png"))
            return
        
        # Bước 5: Xác nhận đang ở trang thanh toán
        in_thong_bao("5. Đang xác nhận trang thanh toán...", reporter=reporter,
                    expected="Xác nhận đang ở trang thanh toán")
        time.sleep(3)
        
        # Kiểm tra xem đã chuyển đến trang thanh toán chưa
        if "thanh-toan" not in driver.current_url:
            in_thong_bao("   - Cảnh báo: Chưa ở trang thanh toán", "CẢNH BÁO", reporter)
            in_thong_bao(f"   - URL hiện tại: {driver.current_url}", "THÔNG TIN", reporter)
            return
            
        # Chụp màn hình trang thanh toán
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "trang_thanh_toan.png"))
        in_thong_bao("   - Đã xác nhận đang ở trang thanh toán", "OK", reporter)
        
        # Bước 6: Điền thông tin giao hàng
        if not dien_thong_tin_giao_hang(driver, thu_muc_ket_qua):
            in_thong_bao("   - Không thể điền thông tin giao hàng", "LỖI", reporter)
            return
        
        # Bước 7: Xác nhận và hoàn tất đơn hàng
        if not xac_nhan_don_hang(driver, thu_muc_ket_qua, reporter):
            in_thong_bao("Kiểm thử thất bại: Lỗi khi xác nhận đơn hàng", "LỖI", reporter)
            return
        
        in_thong_bao("ĐÃ HOÀN THÀNH QUY TRÌNH MUA HÀNG THÀNH CÔNG!", "THÀNH CÔNG", reporter)
        
    except Exception as e:
        in_thong_bao(f"Lỗi trong quá trình thực hiện: {str(e)}", "LỖI", reporter)
        chup_man_hinh(driver, os.path.join(thu_muc_ket_qua, "loi.png"))
    
    finally:
        # Kết thúc
        in_thong_bao("Hoàn thành kiểm thử", reporter=reporter)
        reporter.save_report(thu_muc_ket_qua)
        driver.quit()

if __name__ == "__main__":
    print("="*80)
    print("KIỂM TRA TÍNH NĂNG MUA NGAY TỪ BANNER (CHƯA ĐĂNG NHẬP)")
    print("="*80)
    main()